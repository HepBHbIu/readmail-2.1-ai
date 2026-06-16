from __future__ import annotations

import imaplib
import json
import re
import threading
import time
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from .classifier import (
    apply_ai_overlay,
    build_export_json,
    classify_email,
    force_operator_review,
    load_buyer_rules,
    make_strong_key,
    norm,
    normalize_subject,
    quality_check,
)
from .ai_client import run_ai_suggestion, test_ai_connection, probe_ai_server, list_ai_models
from .email_parser import select_visible_text
from .config import settings
from .db import (
    compact_db, connect, control_dashboard, deliver_outbox_events, dumps, init_db, search_cases,
    learn_buyer_from_case, outbox_dashboard, reconcile_outbox_events, reset_outbox_errors,
    reset_processing_data, load_buyer_identities, loads, queue_case_event, queue_case_export,
    queue_control_events, queue_ready_cases, record_learning_event, record_process_event,
    process_event_dashboard, list_process_events, clear_process_events, reset_processed_work_data, row_to_dict, save_case,
    utcnow,
    get_import_job_status, get_import_errors, get_quarantined_uids, retry_quarantined_uid,
    record_ai_usage,
)
from .runtime_settings import apply_runtime_settings, get_settings_payload, update_settings_from_panel
from . import runtime_control, ai_cost_ledger, server_core, auth as auth_mod, dashboard as dashboard_mod, search as search_mod
from .evidence_panel import router as evidence_panel_router
from .imap_importer import (
    import_from_eml_dir,
    import_from_imap,
    import_from_imap_raw,
    process_imported_emails,
    list_imap_folders,
    decode_imap_utf7,
)
from .demo_data import generate_demo_data

app = FastAPI(title=settings.app_name, version=settings.app_version)
app.include_router(evidence_panel_router)
BASE_DIR = Path(__file__).resolve().parent
WEB_DIR = BASE_DIR / "web"
app.mount("/static", StaticFiles(directory=WEB_DIR / "static"), name="static")

# ── Auth enforcement middleware ───────────────────────────────────────
# Когда auth включён (require_auth=true ИЛИ allow_lan=true), все не-public endpoints закрыты.
_AUTH_PUBLIC_PREFIXES = ("/static", "/favicon")
_AUTH_PUBLIC_EXACT = {"/", "/login", "/api/health", "/api/auth/login", "/api/auth/status",
                      "/api/auth/me", "/api/auth/change-password"}
# Инженерные endpoints — только admin/developer.
_AUTH_DEVELOPER_PREFIXES = ("/api/ai-trace", "/api/evidence", "/api/metrics/ai-cost",
                            "/api/ai/", "/api/outbox-staging")


def _auth_token(request: Request) -> str | None:
    tok = request.cookies.get("readmail_session")
    if tok:
        return tok
    h = request.headers.get("Authorization") or ""
    if h.lower().startswith("bearer "):
        return h[7:].strip()
    return None


@app.middleware("http")
async def _auth_enforcement(request: Request, call_next):
    if not server_core.auth_required():
        return await call_next(request)
    path = request.url.path
    if path in _AUTH_PUBLIC_EXACT or any(path.startswith(p) for p in _AUTH_PUBLIC_PREFIXES):
        return await call_next(request)
    sess = auth_mod.get_session(_auth_token(request))
    if not sess:
        return JSONResponse({"ok": False, "error": "auth_required"}, status_code=401)
    if sess.get("must_change") and not path.startswith("/api/auth/change-password"):
        return JSONResponse({"ok": False, "error": "must_change_password"}, status_code=403)
    if any(path.startswith(p) for p in _AUTH_DEVELOPER_PREFIXES) and sess.get("role") not in ("admin", "developer"):
        return JSONResponse({"ok": False, "error": "developer_only"}, status_code=403)
    return await call_next(request)


class _LoginIn(BaseModel):
    username: str
    password: str


class _ChangePwIn(BaseModel):
    old_password: str
    new_password: str


_LOGIN_HTML = """<!doctype html><html lang=ru><head><meta charset=utf-8>
<title>Readmail — вход</title><meta name=viewport content="width=device-width,initial-scale=1">
<style>body{font-family:system-ui;background:#0f172a;color:#e2e8f0;display:flex;min-height:100vh;
align-items:center;justify-content:center}form{background:#1e293b;padding:28px;border-radius:12px;width:300px}
input{width:100%;padding:10px;margin:6px 0;border-radius:8px;border:1px solid #334155;background:#0f172a;color:#e2e8f0}
button{width:100%;padding:10px;margin-top:10px;border:0;border-radius:8px;background:#2563eb;color:#fff;cursor:pointer}
.msg{font-size:13px;margin-top:10px;color:#fca5a5}.hint{font-size:12px;color:#94a3b8}</style></head>
<body><form onsubmit="return doLogin(event)"><h3>Readmail</h3>
<input id=u placeholder="Логин" autocomplete=username>
<input id=p type=password placeholder="Пароль" autocomplete=current-password>
<button>Войти</button><div class=msg id=m></div>
<div class=hint id=h></div></form>
<script>
fetch('/api/auth/status').then(r=>r.json()).then(s=>{if(s.bootstrap_required)
document.getElementById('h').textContent='Первый вход: admin / admin (затем смена пароля)';});
async function doLogin(e){e.preventDefault();
const r=await fetch('/api/auth/login',{method:'POST',headers:{'Content-Type':'application/json'},
body:JSON.stringify({username:u.value,password:p.value})});const d=await r.json();
if(d.ok){if(d.must_change){location.href='/?must_change=1';}else{location.href='/';}}
else{document.getElementById('m').textContent=d.hint||d.error||'Ошибка входа';}return false;}
</script></body></html>"""


@app.get("/login", response_class=HTMLResponse)
def login_page() -> Any:
    """Публичная страница входа (если auth включён). Без секретов."""
    return HTMLResponse(_LOGIN_HTML)


@app.get("/api/auth/status")
def api_auth_status() -> dict[str, Any]:
    return {"ok": True, "auth_required": server_core.auth_required(),
            "bootstrap_required": auth_mod.bootstrap_required(),
            "admin_configured": auth_mod.admin_configured()}


@app.post("/api/auth/login")
def api_auth_login(payload: _LoginIn) -> Any:
    res = auth_mod.login(payload.username, payload.password)
    resp = JSONResponse(res, status_code=200 if res.get("ok") else 401)
    if res.get("ok"):
        resp.set_cookie("readmail_session", res["token"], httponly=True, samesite="lax")
    return resp


@app.post("/api/auth/logout")
def api_auth_logout(request: Request) -> Any:
    auth_mod.revoke(_auth_token(request))
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("readmail_session")
    return resp


@app.get("/api/auth/me")
def api_auth_me(request: Request) -> dict[str, Any]:
    return auth_mod.me(_auth_token(request))


@app.post("/api/auth/change-password")
def api_auth_change_password(payload: _ChangePwIn, request: Request) -> Any:
    res = auth_mod.change_password(_auth_token(request), payload.old_password, payload.new_password)
    resp = JSONResponse(res, status_code=200 if res.get("ok") else 400)
    if res.get("ok"):
        resp.set_cookie("readmail_session", res["token"], httponly=True, samesite="lax")
    return resp


def _case_export_from_row(row: dict[str, Any], fields: dict[str, Any] | None = None) -> dict[str, Any]:
    """Build export_json from the same fields that are visible in the case UI."""
    case_data = {
        "buyer_code": row.get("buyer_code"),
        "buyer_name": row.get("buyer_name"),
        "event_type": row.get("event_type"),
        "claim_kind": row.get("claim_kind"),
        "status": row.get("status"),
        "priority": row.get("priority"),
        "confidence": row.get("confidence"),
        "deadline_at": row.get("deadline_at"),
        "thread_key": row.get("thread_key"),
        "strong_key": row.get("strong_key"),
        "weak_key": row.get("weak_key"),
        "is_followup": bool(row.get("is_followup")),
        "ready_for_export": bool(row.get("ready_for_export")),
        "needs_review": bool(row.get("needs_review")),
        "state": row.get("state"),
        "fields": fields if fields is not None else (row.get("fields") or {}),
        "missing": row.get("missing") or [],
        "quality": row.get("quality") or [],
        "payload": row.get("payload") or {},
    }
    email_data = {
        "subject": row.get("subject"),
        "from_addr": row.get("from_addr"),
        "received_at": row.get("received_at"),
        "body_text": row.get("body_text"),
        "snippet": row.get("snippet"),
    }
    return build_export_json(row.get("id"), email_data, case_data)


class ManualUpdate(BaseModel):
    buyer_code: str | None = None
    buyer_name: str | None = None
    event_type: str | None = None
    claim_kind: str | None = None
    status: str | None = None
    priority: str | None = None
    deadline_at: str | None = None
    state: str | None = None
    fields: dict[str, Any] | None = None
    ready_for_export: bool | None = None
    needs_review: bool | None = None


class SettingsUpdate(BaseModel):
    values: dict[str, Any]


class FolderSelectionUpdate(BaseModel):
    folders: list[str]


_IMPORT_LOCK = threading.Lock()
_IMPORT_STOP = threading.Event()
_SCAN_THREAD_STARTED = False
_LAST_IMPORT_RESULT: dict[str, Any] = {"ok": True, "message": "not_started"}

# Pipeline stop events for patterns + AI
_PATTERNS_LOCK = threading.Lock()
_PATTERNS_STOP = threading.Event()
_AI_BATCH_LOCK = threading.Lock()
_AI_BATCH_STOP = threading.Event()

# Глобальный замок писателей БД. SQLite держит ОДНОГО писателя — параллельные
# фоновые задачи (импорт, паттерны, AI-батч, обучение, цикл автопилота) дают
# "database is locked". Все они проходят через этот ОДИН реентрантный замок →
# сериализуются в один поток записи. RLock (не Lock!): цикл автопилота держит
# замок и внутри вызывает под-шаги (импорт→классификация→...) в ТОМ ЖЕ потоке —
# реентрантность спасает от самодедлока. Ручные кнопки идут из других потоков и
# спокойно ждут в очереди, пока крутит автопилот.
_PIPELINE_LOCK = threading.RLock()

# v1.14: operator-facing autopilot. This is deliberately in-memory: if the app restarts,
# autopilot stops safely and the operator must start it again from the panel.
_AUTOPILOT_LOCK = threading.Lock()
_AUTOPILOT_STOP = threading.Event()
_AUTOPILOT_THREAD: threading.Thread | None = None
_AUTOPILOT_STATE: dict[str, Any] = {
    "enabled": False,
    "running_cycle": False,
    "started_at": None,
    "stopped_at": None,
    "last_cycle_started_at": None,
    "last_cycle_finished_at": None,
    "next_cycle_at": None,
    "cycle_count": 0,
    "last_error": None,
    "last_cycle": {},
    "config": {"interval_seconds": 300, "import_limit": 50, "ai_limit": 10, "deliver": False},
}


def _log(stage: str, message: str, *, level: str = "info", case_id: int | None = None, raw_email_id: int | None = None, subject: str | None = None, details: dict[str, Any] | None = None) -> None:
    """Best-effort live timeline log. Never breaks the business pipeline."""
    try:
        with connect() as con:
            record_process_event(con, stage=stage, message=message, level=level, case_id=case_id, raw_email_id=raw_email_id, subject=subject, details=details or {})
    except Exception:
        pass


def _auto_link_followup(con: Any, case_id: int, case_data: dict[str, Any]) -> None:
    """Auto-link a follow-up/reminder to its parent new_return case.

    Searches by thread_key / strong_key. If found: marks linked_event and optionally
    queues a 1C update event with escalated priority if configured.
    """
    try:
        apply_runtime_settings()
        thread_key = case_data.get("thread_key")
        strong_key = case_data.get("strong_key")
        buyer_code = case_data.get("buyer_code")
        event_type = case_data.get("event_type")
        fields = case_data.get("fields") or {}

        parent_row = None
        # 1. Try strong_key match
        if strong_key:
            parent_row = con.execute(
                "SELECT id, state, priority FROM cases WHERE strong_key=? AND event_type='new_return' AND id<>? ORDER BY id ASC LIMIT 1",
                (strong_key, case_id),
            ).fetchone()
        # 2. Try thread_key
        if not parent_row and thread_key:
            parent_row = con.execute(
                "SELECT id, state, priority FROM cases WHERE thread_key=? AND event_type='new_return' AND id<>? ORDER BY id ASC LIMIT 1",
                (thread_key, case_id),
            ).fetchone()
        # 3. For document/service events, link by document number.
        if not parent_row and buyer_code and event_type in {"correction_request", "marking_request", "ready_to_ship"}:
            doc = fields.get("document_number")
            if doc:
                parent_row = con.execute(
                    """SELECT c.id, c.state, c.priority FROM cases c
                       WHERE c.buyer_code=? AND c.event_type='new_return'
                         AND c.fields_json LIKE ? AND c.id<>? ORDER BY c.id ASC LIMIT 1""",
                    (buyer_code, f"%{doc}%", case_id),
                ).fetchone()
        # 3.5 Привязка по № заявки/возврата (return_number) — у ixora/profit/avtoto/favorit
        # это главный ключ диалога (в теме И теле и материнского, и ответа).
        if not parent_row and buyer_code:
            rnum = fields.get("return_number") or fields.get("client_request_number")
            if rnum and len(str(rnum)) >= 4:
                parent_row = con.execute(
                    """SELECT c.id, c.state, c.priority FROM cases c
                       WHERE c.buyer_code=? AND c.event_type='new_return'
                         AND c.fields_json LIKE ? AND c.id<>? ORDER BY c.id ASC LIMIT 1""",
                    (buyer_code, f'%"{rnum}"%', case_id),
                ).fetchone()
        # 4. Try buyer + same business number (weak match)
        if not parent_row and buyer_code:
            claim_num = fields.get("claim_number") or fields.get("client_request_number")
            if claim_num:
                parent_row = con.execute(
                    """SELECT c.id, c.state, c.priority FROM cases c
                       WHERE c.buyer_code=? AND c.event_type='new_return'
                         AND c.fields_json LIKE ? AND c.id<>? ORDER BY c.id ASC LIMIT 1""",
                    (buyer_code, f"%{claim_num}%", case_id),
                ).fetchone()

        if not parent_row:
            return

        parent_id = int(parent_row["id"])
        parent_priority = parent_row["priority"] or "normal"

        # Наследование полей от материнского кейса (B6): followup-ответ («ждём ответ»,
        # «есть решение?») часто не несёт своих данных — паттерн вытаскивает мусор из
        # текста ответа. Если у followup НЕТ своего артикула — это чистый диалог: берём
        # идентификацию позиции от родителя, документ/кол-во дозаполняем при отсутствии.
        try:
            cf = dict(case_data.get("fields") or {})
            prow = con.execute("SELECT fields_json, claim_kind FROM cases WHERE id=?", (parent_id,)).fetchone()
            if prow:
                pf = loads(prow["fields_json"], {}) or {}
                # Та же позиция, что у родителя (нет своего артикула ИЛИ совпадает) → имя/бренд
                # из текста ответа = мусор, перекрываем родительскими (родитель — чистый источник).
                same_item = (not cf.get("part_number")) or (cf.get("part_number") == pf.get("part_number"))
                if same_item:
                    for k in ("part_number", "brand", "product_name"):
                        if pf.get(k):
                            cf[k] = pf[k]
                for k in ("document_number", "document_date", "quantity"):
                    if pf.get(k) and not cf.get(k):
                        cf[k] = pf[k]
                inherited_kind = case_data.get("claim_kind") or prow["claim_kind"]
                con.execute(
                    "UPDATE cases SET fields_json=?, claim_kind=COALESCE(claim_kind,?), updated_at=? WHERE id=?",
                    (dumps(cf), inherited_kind, utcnow(), case_id),
                )
                case_data["fields"] = cf
        except Exception:
            pass

        # Mark this follow-up as linked
        con.execute(
            "UPDATE cases SET state='linked_event', thread_key=COALESCE(thread_key,?), updated_at=? WHERE id=?",
            (f"parent:{parent_id}", utcnow(), case_id),
        )

        # Escalate parent priority if needed
        escalation_map = {"normal": "medium", "medium": "high", "high": "critical", "critical": "critical"}
        new_priority = escalation_map.get(parent_priority, parent_priority)
        if new_priority != parent_priority:
            con.execute("UPDATE cases SET priority=?, updated_at=? WHERE id=?", (new_priority, utcnow(), parent_id))

        # Queue to 1C as "followup_update" event if configured
        if getattr(settings, "send_followups_to_1c", False):
            followup_payload = {
                "schema_version": "readmail-followup-v1",
                "event": {
                    "type": "followup_update",
                    "source_event_type": case_data.get("event_type"),
                    "parent_case_id": parent_id,
                    "followup_case_id": case_id,
                    "priority_escalated": new_priority != parent_priority,
                    "new_priority": new_priority,
                },
                "parent_case_id": parent_id,
                "buyer": {"code": buyer_code, "name": case_data.get("buyer_name")},
                "note": f"Продолжение диалога привязано к кейсу #{parent_id}",
            }
            event_key = f"followup:{case_id}:parent:{parent_id}"
            con.execute(
                "INSERT OR IGNORE INTO outbox(case_id, payload_json, status, created_at, event_type, event_key, channel, business_priority) VALUES (?,?,'new',?,?,?,?,?)",
                (case_id, dumps(followup_payload), utcnow(), "followup_update", event_key, "file", new_priority),
            )

        try:
            record_process_event(
                con,
                stage="classifier",
                level="ok",
                message=f"Диалог #{case_id} привязан к кейсу #{parent_id}, приоритет -> {new_priority}",
                case_id=case_id,
                details={"parent_case_id": parent_id, "new_priority": new_priority},
            )
        except Exception:
            pass
    except Exception as exc:
        try:
            record_process_event(
                con,
                stage="classifier",
                level="warn",
                message=f"Автолинк диалога #{case_id}: ошибка {exc}",
                case_id=case_id,
            )
        except Exception:
            pass


def _set_payload_ref(con: Any, case_id: int, key: str, ref: dict[str, Any]) -> None:
    """Записать перекрёстную ссылку в payload_json кейса (не меняя состояний)."""
    row = con.execute("SELECT payload_json FROM cases WHERE id=?", (case_id,)).fetchone()
    if not row:
        return
    try:
        payload = json.loads(row["payload_json"] or "{}")
    except Exception:
        payload = {}
    payload[key] = ref
    con.execute("UPDATE cases SET payload_json=?, updated_at=? WHERE id=?",
                (json.dumps(payload, ensure_ascii=False), utcnow(), case_id))


def _link_problem_notice(con: Any, case_id: int, case_data: dict[str, Any]) -> None:
    """Мягкая связка «уведомление о проблеме» ↔ «возврат» по документ+артикул.

    problem_notice — пред-претензия (товар принят с дефектом, видеофиксация). Когда по той
    же позиции (накладная+артикул) появляется реальный возврат — связываем их перекрёстно,
    НЕ меняя состояний: возврат остаётся в 1С, уведомление — на своей полке. Оператор видит
    связь с обеих сторон. Срабатывает в любом порядке (notice раньше или возврат раньше)."""
    try:
        et = case_data.get("event_type")
        fields = case_data.get("fields") or {}
        doc = str(fields.get("document_number") or "").strip()
        part = str(fields.get("part_number") or "").strip()
        strong = case_data.get("strong_key")
        if et == "problem_notice":
            target_types = ("new_return", "pre_delivery_refusal")
        elif et in ("new_return", "pre_delivery_refusal"):
            target_types = ("problem_notice",)
        else:
            return
        if not ((doc and part) or strong):
            return
        ph = ",".join("?" for _ in target_types)
        other = None
        if doc and part:
            other = con.execute(
                f"""SELECT id FROM cases WHERE event_type IN ({ph}) AND id<>?
                    AND json_extract(fields_json,'$.document_number')=?
                    AND json_extract(fields_json,'$.part_number')=? ORDER BY id ASC LIMIT 1""",
                (*target_types, case_id, doc, part),
            ).fetchone()
        if not other and strong:
            other = con.execute(
                f"SELECT id FROM cases WHERE event_type IN ({ph}) AND id<>? AND strong_key=? ORDER BY id ASC LIMIT 1",
                (*target_types, case_id, strong),
            ).fetchone()
        if not other:
            return
        other_id = int(other["id"])
        notice_id, return_id = (case_id, other_id) if et == "problem_notice" else (other_id, case_id)
        _set_payload_ref(con, notice_id, "realized_by_return",
                         {"case_id": return_id, "doc": doc, "part": part})
        _set_payload_ref(con, return_id, "preceding_problem_notice",
                         {"case_id": notice_id, "doc": doc, "part": part})
        record_process_event(
            con, stage="classifier", level="info",
            message=f"Связь: уведомление о проблеме #{notice_id} ↔ возврат #{return_id} (накладная {doc}, арт. {part})",
            case_id=case_id,
        )
    except Exception:
        pass


def _processed_hidden_count(con: Any) -> int:
    """Кол-во писем в разделе «Обработанные / не требуют действия» — для бейджа вкладки."""
    try:
        from . import processed_hidden as ph
        return int(ph.build_processed_hidden_summary(con).get("hidden_from_operator") or 0)
    except Exception:
        return 0


def _auto_queue_ready_if_enabled() -> dict[str, Any] | None:
    apply_runtime_settings()
    result: dict[str, Any] = {"ok": True}
    with connect() as con:
        if getattr(settings, "auto_queue_control_events", True):
            result["control_events"] = queue_control_events(con, limit=1000)
        elif getattr(settings, "auto_queue_ready_to_outbox", False):
            result["ready"] = queue_ready_cases(con, limit=500)
        else:
            return None
        if getattr(settings, "auto_deliver_outbox", False):
            result["delivery"] = deliver_outbox_events(con, limit=500)
    return result


def _patterns_dashboard() -> dict[str, Any]:
    """Отчёт: сколько писем без кейсов, сколько needs_review, сколько ready."""
    try:
        with connect() as con:
            raw_no_case = con.execute(
                "SELECT COUNT(*) c FROM raw_emails r LEFT JOIN cases c ON c.raw_email_id=r.id WHERE c.id IS NULL"
            ).fetchone()["c"]
            needs_review = con.execute(
                "SELECT COUNT(*) c FROM cases WHERE state='needs_review' AND event_type='new_return'"
            ).fetchone()["c"]
            ready = con.execute(
                "SELECT COUNT(*) c FROM cases WHERE state='ready_to_1c' AND ready_for_export=1"
            ).fetchone()["c"]
            return {"raw_no_case": raw_no_case, "needs_review": needs_review, "ready": ready}
    except Exception:
        return {"raw_no_case": 0, "needs_review": 0, "ready": 0}


def _run_full_pipeline(limit: int = 100) -> dict[str, Any]:
    """Полный пайплайн: паттерны → AI → очередь.

    1. Находит все raw_emails без кейсов → classify_email (паттерны)
    2. Для needs_review + needs_ai → AI-доработка
    3. Для ready_to_1c → в очередь outbox
    """
    _log("pipeline", "Полный пайплайн: старт")
    result: dict[str, Any] = {"ok": True, "patterns": 0, "ai": 0, "queued": 0}
    buyer_rules = load_buyer_rules()

    try:
        with connect() as con:
            # ── ШАГ 1: Паттерны ── новые письма без кейсов ──
            existing_cases_rows = con.execute(
                """
                SELECT c.id, c.event_type, e.from_addr, e.subject
                FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
                WHERE c.event_type IN ('new_return','followup_dialog','followup_reminder','supplier_decision','unknown')
                ORDER BY c.id
                """
            ).fetchall()
            existing_cases = []
            for ecr in existing_cases_rows:
                d = dict(ecr)
                existing_cases.append({
                    "from_addr": norm(d.get("from_addr", "")),
                    "subject_template": normalize_subject(d.get("subject", "")),
                    "event_type": d.get("event_type"),
                })

            raw_no_case = con.execute(
                """
                SELECT r.* FROM raw_emails r
                LEFT JOIN cases c ON c.raw_email_id=r.id
                WHERE c.id IS NULL
                ORDER BY r.id LIMIT ?
                """,
                (limit,),
            ).fetchall()

            for row in raw_no_case:
                email_data = row_to_dict(row) or {}
                email_data["attachments"] = [
                    dict(a) for a in con.execute(
                        "SELECT filename, content_type, size_bytes FROM attachments WHERE raw_email_id=?",
                        (row["id"],),
                    ).fetchall()
                ]
                email_data["visible_text"] = email_data.get("body_text") or email_data.get("snippet") or ""
                learned = load_buyer_identities(con)
                case_data = classify_email(
                    email_data, buyer_rules,
                    learned_identities=learned,
                    existing_cases=existing_cases,
                )
                case_id = save_case(con, int(row["id"]), case_data)
                case_data["export"]["case_id"] = case_id
                con.execute("UPDATE cases SET export_json=?, updated_at=? WHERE id=?", (dumps(case_data.get("export") or {}), utcnow(), case_id))

                # v2.1 AI-only: наблюдение/промоция паттернов убраны.
                event_type = case_data.get("event_type", "")
                if event_type in ("followup_reminder", "followup_dialog", "supplier_decision", "correction_request", "marking_request", "ready_to_ship"):
                    _auto_link_followup(con, case_id, case_data)
                _link_problem_notice(con, case_id, case_data)

                existing_cases.append({
                    "from_addr": norm(email_data.get("from_addr", "")),
                    "subject_template": normalize_subject(email_data.get("subject", "")),
                    "event_type": event_type,
                })
                result["patterns"] += 1

            # ── ШАГ 2: AI для needs_review ──
            if settings.enable_ai:
                ai_candidates = con.execute(
                    """
                    SELECT c.id FROM cases c
                    WHERE c.state IN ('needs_review','needs_link')
                      AND c.event_type IN ('new_return','unknown')
                      AND (c.needs_ai = 1 OR c.event_type = 'unknown')
                      AND NOT EXISTS (
                        SELECT 1 FROM ai_suggestions s WHERE s.case_id=c.id AND s.accepted=1
                      )
                    ORDER BY c.id
                    LIMIT ?
                    """,
                    (min(limit, 50),),
                ).fetchall()

                for acr in ai_candidates:
                    try:
                        r = _apply_ai_to_case_id(int(acr["id"]), purpose="pipeline_ai_apply")
                        if r.get("applied"):
                            result["ai"] += 1
                    except Exception as exc:
                        _log("pipeline", f"AI для кейса {acr['id']}: {exc}", level="warn")

            # ── ШАГ 3: Queue ready cases ──
            ready_queued = queue_control_events(con, limit=500)
            result["queued"] = int(ready_queued.get("queued") or 0)

        _log("pipeline", "Полный пайплайн: завершён", level="ok",
             details=result)
    except Exception as exc:
        _log("pipeline", "Полный пайплайн: ошибка", level="error", details={"error": str(exc)})
        result["ok"] = False
        result["error"] = str(exc)
    return result


def _safe_import_cycle() -> dict[str, Any]:
    global _LAST_IMPORT_RESULT
    apply_runtime_settings()
    _log("import", "Автоимпорт: старт", details={"folders": settings.folders, "limit": settings.imap_limit})
    if not _IMPORT_LOCK.acquire(blocking=False):
        return {"ok": False, "skipped": True, "reason": "import_already_running"}
    try:
        result = import_from_imap_raw()
        _log("import", "Автоимпорт: завершён", level="ok" if result.get("ok", True) else "warn", details=result)

        # ── ПОЛНЫЙ ПАЙПЛАЙН ВСЕГДА ──
        pipeline = _run_full_pipeline(limit=200)
        result["pipeline"] = pipeline
        _log("pipeline", "Пайплайн обработки: паттерны={patterns}, AI={ai}, очередь={queued}".format(**pipeline),
             level="ok" if pipeline.get("ok") else "warn",
             details=pipeline)

        # ── Outbox (если включено) ──
        queued = _auto_queue_ready_if_enabled()
        if queued:
            result["auto_outbox"] = queued
        _LAST_IMPORT_RESULT = {**result, "finished_at": utcnow()}
        return _LAST_IMPORT_RESULT
    except Exception as exc:
        _log("import", "Автоимпорт: ошибка", level="error", details={"error": str(exc)})
        _LAST_IMPORT_RESULT = {"ok": False, "error": str(exc), "finished_at": utcnow()}
        return _LAST_IMPORT_RESULT
    finally:
        _IMPORT_LOCK.release()




def _extract_attachment_text(file_path: str, filename: str, max_chars: int = 6000) -> str:
    """Текст из вложения (Excel/CSV/TXT, вскрытие zip) — для дочитки ИИ полей из акта.
    Фото/PDF не трогаем (это путь vision). Возвращает '' если нечего извлечь."""
    if not file_path or not Path(file_path).exists():
        return ""
    fn = (filename or "").lower()
    ext = fn.rsplit(".", 1)[-1] if "." in fn else ""

    def _dec(b: bytes) -> str:
        for enc in ("utf-8-sig", "cp1251", "latin-1"):
            try:
                return b.decode(enc)
            except Exception:
                continue
        return b.decode("utf-8", errors="replace")

    def _xlsx_rows(data_or_path: Any) -> str:
        import io as _io
        out: list[str] = []
        # .xlsx/.xlsm → openpyxl; старый OLE2 .xls openpyxl НЕ читает → фолбэк на xlrd.
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_io.BytesIO(data_or_path) if isinstance(data_or_path, bytes) else data_or_path,
                                        data_only=True, read_only=True)
            for ws in list(wb.worksheets)[:2]:
                for i, rd in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 120:
                        break
                    cells = [str(c) for c in rd if c is not None]
                    if cells:
                        out.append(" | ".join(cells))
            if out:
                return "\n".join(out)
        except Exception:
            pass
        try:
            import xlrd
            book = (xlrd.open_workbook(file_contents=data_or_path) if isinstance(data_or_path, bytes)
                    else xlrd.open_workbook(data_or_path))
            for sh in book.sheets()[:2]:
                for r in range(min(sh.nrows, 120)):
                    cells = [str(c) for c in sh.row_values(r) if str(c).strip() not in ("", "None")]
                    if cells:
                        out.append(" | ".join(cells))
        except Exception:
            return ""
        return "\n".join(out)

    def _inner_text(name: str, data: bytes) -> str:
        """Текст из файла ВНУТРИ архива по его расширению (Excel/CSV/TXT/PDF)."""
        ie = name.lower().rsplit(".", 1)[-1] if "." in name else ""
        try:
            if ie in {"xlsx", "xls", "xlsm"}:
                return _xlsx_rows(data)
            if ie in {"csv", "txt"}:
                return _dec(data)
            if ie == "pdf":
                from .email_parser import _extract_pdf_text
                return _extract_pdf_text(data) or ""
        except Exception:
            return ""
        return ""

    try:
        if ext in {"xlsx", "xls", "xlsm"}:
            return _xlsx_rows(file_path)[:max_chars]
        if ext in {"csv", "txt"}:
            return _dec(Path(file_path).read_bytes())[:max_chars]
        if ext == "zip":
            import zipfile
            parts: list[str] = []
            with zipfile.ZipFile(file_path) as zf:
                for zi in zf.infolist():
                    if zi.is_dir():
                        continue
                    t = _inner_text(zi.filename, zf.read(zi.filename))
                    if t.strip():
                        parts.append(t)
                    if sum(len(p) for p in parts) > max_chars:
                        break
            return "\n".join(parts)[:max_chars]
        if ext == "7z":
            try:
                import py7zr  # best-effort: либы нет в образе → graceful ""
                parts: list[str] = []
                with py7zr.SevenZipFile(file_path, mode="r") as z:
                    for name, bio in (z.readall() or {}).items():
                        t = _inner_text(name, bio.read())
                        if t.strip():
                            parts.append(t)
                        if sum(len(p) for p in parts) > max_chars:
                            break
                return "\n".join(parts)[:max_chars]
            except Exception:
                return ""
        if ext == "rar":
            try:
                import rarfile  # best-effort: нужен системный unrar/unar → graceful ""
                parts: list[str] = []
                with rarfile.RarFile(file_path) as rf:
                    for ri in rf.infolist():
                        if ri.isdir():
                            continue
                        t = _inner_text(ri.filename, rf.read(ri))
                        if t.strip():
                            parts.append(t)
                        if sum(len(p) for p in parts) > max_chars:
                            break
                return "\n".join(parts)[:max_chars]
            except Exception:
                return ""
    except Exception:
        return ""
    return ""


def _case_attachment_text(raw_email_id: int, max_chars: int = 8000) -> str:
    """Собрать текст из всех Excel/doc/zip-вложений письма (для дочитки ИИ)."""
    with connect() as con:
        atts = con.execute(
            "SELECT filename, content_type, file_path FROM attachments WHERE raw_email_id=? ORDER BY id",
            (raw_email_id,),
        ).fetchall()
    chunks: list[str] = []
    for a in atts:
        t = _extract_attachment_text(a["file_path"], a["filename"])
        if t.strip():
            chunks.append(f"[{a['filename']}]\n{t}")
        if sum(len(c) for c in chunks) > max_chars:
            break
    return ("\n\n".join(chunks))[:max_chars]


def _load_case_email_for_ai(case_id: int) -> tuple[dict[str, Any], dict[str, Any]] | None:
    with connect() as con:
        row = con.execute(
            """
            SELECT c.*, e.mailbox, e.uid, e.message_id, e.in_reply_to, e.references_json, e.subject, e.from_addr,
                   e.to_addr, e.cc_addr, e.direction, e.folder_seen_json, e.received_at, e.body_text, e.body_html,
                   e.snippet, e.quote_markers
            FROM cases c JOIN raw_emails e ON e.id = c.raw_email_id WHERE c.id=?
            """,
            (case_id,),
        ).fetchone()
        if not row:
            return None
        data = row_to_dict(row) or {}
        attachments = [dict(a) for a in con.execute("SELECT filename, content_type, size_bytes FROM attachments WHERE raw_email_id=?", (data["raw_email_id"],)).fetchall()]
    email_data = {
        "mailbox": data.get("mailbox"), "uid": data.get("uid"), "message_id": data.get("message_id"),
        "in_reply_to": data.get("in_reply_to"), "references": data.get("references") or [],
        "subject": data.get("subject"), "from_addr": data.get("from_addr"), "to_addr": data.get("to_addr"),
        "cc_addr": data.get("cc_addr"), "received_at": data.get("received_at"),
        "body_text": data.get("body_text"), "body_html": data.get("body_html"), "snippet": data.get("snippet"),
        "quote_markers": data.get("quote_markers"), "attachments": attachments,
    }
    case_data = {
        "buyer_code": data.get("buyer_code"), "buyer_name": data.get("buyer_name"),
        "event_type": data.get("event_type"), "claim_kind": data.get("claim_kind"), "status": data.get("status"),
        "priority": data.get("priority"), "confidence": data.get("confidence"), "deadline_at": data.get("deadline_at"),
        "thread_key": data.get("thread_key"), "strong_key": data.get("strong_key"), "weak_key": data.get("weak_key"),
        "is_followup": bool(data.get("is_followup")), "ready_for_export": bool(data.get("ready_for_export")),
        "needs_review": bool(data.get("needs_review")), "state": data.get("state"),
        "fields": data.get("fields") or {}, "missing": data.get("missing") or [], "quality": data.get("quality") or [],
        "payload": data.get("payload") or {}, "export": data.get("export") or {},
    }
    return email_data, case_data


def _is_contentless_reminder(email_data: dict[str, Any]) -> bool:
    """Письмо — пустое напоминание без данных («Добрый день. Ожидаем ответ.»). Такие НЕ отдаём
    ИИ на извлечение полей: данных нет, риск выдумки. Поля из темы уже взяты паттернами."""
    vt = (email_data.get("visible_text") or email_data.get("body_text") or email_data.get("snippet") or "").strip()
    if len(vt) > 400:  # достаточно текста — пусть ИИ смотрит
        return False
    low = vt.lower()
    REMIND = ("ожидаем ответ", "ждем ответ", "ждём ответ", "напоминаем", "просьба ответить",
              "почему игнорируете", "срочно дать ответ", "прошу ответить", "нет ответа",
              "ожидаем решени", "ждем решени", "ждём решени")
    if not any(p in low for p in REMIND):
        return False
    # признаки реальных данных (артикул/таблица/количество) → не пусто, пусть ИИ смотрит
    has_data = bool(re.search(r"\b[a-zа-я]{2,}\d{3,}\b", low)) or "артикул" in low or "|" in vt or "кол-во" in low
    return not has_data


def _apply_ai_to_case_id(
    case_id: int,
    *,
    purpose: str = "autopilot_ai_apply",
    queue_ready: bool = True,
    manual_review_gate: bool = False,
    read_attachments: bool = False,
) -> dict[str, Any]:
    _log("ai", "AI: подготовка кейса", case_id=case_id, details={"purpose": purpose})
    # Тег режима для учёта токенов: full_ai (полный ИИ) vs pattern (паттерн+AI-фолбэк).
    _ai_mode = "full_ai" if purpose in ("manual_full_ai", "autopilot_full_ai") else "pattern"
    try:
        from .db import set_ai_usage_context
        set_ai_usage_context(mode=_ai_mode, kind="text")
    except Exception:
        pass
    loaded = _load_case_email_for_ai(case_id)
    if not loaded:
        _log("ai", "AI: кейс не найден", level="error", case_id=case_id, details={"purpose": purpose})
        return {"ok": False, "case_id": case_id, "error": "case_not_found", "applied": False}
    email_data, case_data = loaded
    # v2.1: авто-дочитка ДЕШЁВОГО текста вложений (Excel/CSV/zip-акты) на первом прогоне —
    # без vision. Бренд/наименование/№ у браков часто только в акте ТОРГ-2 (.xls/.zip).
    if not read_attachments:
        _doc_exts = (".xls", ".xlsx", ".xlsm", ".csv", ".txt", ".zip", ".rar", ".7z")
        for _a in (email_data.get("attachments") or []):
            if str(_a.get("filename") or "").lower().endswith(_doc_exts):
                read_attachments = True
                break
    # Дочитка вложений: подмешиваем текст Excel/doc-актов к телу (бренд/имя у браков — в акте).
    if read_attachments:
        try:
            with connect() as _c:
                _rid_row = _c.execute("SELECT raw_email_id FROM cases WHERE id=?", (case_id,)).fetchone()
            _rid = int(_rid_row["raw_email_id"]) if _rid_row else 0
            att_txt = _case_attachment_text(_rid) if _rid else ""
            if att_txt.strip():
                _base = email_data.get("body_text") or email_data.get("snippet") or ""
                email_data = {**email_data, "body_text": _base + "\n\n[ТЕКСТ ВЛОЖЕНИЙ/АКТА]\n" + att_txt, "visible_text": None}
                _log("ai", "AI: дочитаны вложения", case_id=case_id, details={"att_chars": len(att_txt)})
        except Exception:
            pass
    # Защита от выдумок: пустое письмо-напоминание («Ожидаем ответ») без данных не отдаём ИИ —
    # извлекать нечего, есть риск галлюцинации полей (был кейс #24893). Поля из темы уже взяты паттернами.
    if _is_contentless_reminder(email_data):
        _log("ai", "AI: пропуск — пустое напоминание без данных", level="info", case_id=case_id, subject=email_data.get("subject"))
        return {"ok": True, "case_id": case_id, "applied": False, "skipped": "contentless_reminder"}
    _log("ai", "AI: читает письмо", case_id=case_id, raw_email_id=case_data.get("raw_email_id"), subject=email_data.get("subject"), details={"event_type": case_data.get("event_type"), "claim_kind": case_data.get("claim_kind"), "missing": case_data.get("missing"), "quality": case_data.get("quality")})
    with connect() as con:
        _log("ai", "AI: запрос к модели", case_id=case_id, subject=email_data.get("subject"), details={"provider": settings.ai_provider, "model": settings.ai_model, "purpose": purpose})
        suggestion = run_ai_suggestion(email_data, case_data, con=con, case_id=case_id, purpose=purpose)
        con.execute(
            "INSERT INTO ai_suggestions(case_id, model, prompt_hash, response_json, accepted, created_at) VALUES (?, ?, ?, ?, 0, ?)",
            (case_id, suggestion.get("model"), suggestion.get("prompt_hash"), dumps(suggestion), utcnow()),
        )
    if not suggestion.get("ok") or not isinstance(suggestion.get("response"), dict):
        _log("ai", "AI: нет пригодного JSON", level="warn", case_id=case_id, subject=email_data.get("subject"), details={"ok": suggestion.get("ok"), "error": suggestion.get("error"), "provider": suggestion.get("provider"), "model": suggestion.get("model"), "usage": suggestion.get("usage"), "raw_excerpt": suggestion.get("raw_excerpt"), "hint": "Модель ответила, но JSON не найден или JSON невалидный. Смотри raw_excerpt."})
        return {"ok": False, "case_id": case_id, "suggestion": suggestion, "applied": False}
    _log("ai", "AI: получил JSON-подсказку", level="ok", case_id=case_id, subject=email_data.get("subject"), details={"provider": suggestion.get("provider"), "model": suggestion.get("model"), "cached": bool(suggestion.get("cached")), "usage": suggestion.get("usage"), "keys": sorted(list((suggestion.get("response") or {}).keys()))[:30], "response_preview": suggestion.get("response")})
    updated = apply_ai_overlay(email_data, case_data, suggestion["response"])
    if manual_review_gate:
        updated = force_operator_review(updated)
    learning_result: dict[str, Any] = {"skipped": "not_started"}
    with connect() as con:
        raw_row = con.execute("SELECT raw_email_id FROM cases WHERE id=?", (case_id,)).fetchone()
        if not raw_row:
            return {"ok": False, "case_id": case_id, "error": "case_not_found_after_ai", "applied": False}
        updated_id = save_case(con, int(raw_row["raw_email_id"]), updated)
        updated["export"]["case_id"] = updated_id
        con.execute(
            "UPDATE cases SET export_json=?, updated_at=? WHERE id=?",
            (dumps(updated.get("export") or {}), utcnow(), updated_id),
        )
        con.execute("UPDATE ai_suggestions SET accepted=1 WHERE case_id=? AND prompt_hash=?", (case_id, suggestion.get("prompt_hash")))
        # Только готовые кейсы идут в 1С; остальные остаются в нужных вкладках
        if queue_ready and updated.get("state") == "ready_to_1c" and updated.get("ready_for_export"):
            queue_case_event(con, updated_id)
        # Мягкая связка problem_notice ↔ возврат по документ+артикул (не меняет состояний).
        _link_problem_notice(con, updated_id, updated)
        # v2.1 AI-only: самообучение паттернов убрано (паттернов нет).
    # Брак/некондиция: авто-триггер vision по документам (несколько запросов + переключение на
    # vision-модель). Нет вложений → state=absent, vision не вызывается. Токены пишутся kind=vision
    # с текущим режимом (pattern/full_ai) → видно в отчёте токенов и в AI-логе.
    defect_result = None
    if (updated.get("claim_kind") or case_data.get("claim_kind")) in ("defect", "nonconforming"):
        try:
            defect_result = api_check_defect_docs(case_id)
            _log("ai", "AI: авто-проверка документов брака", case_id=case_id,
                 details={"state": defect_result.get("state"), "n_present": defect_result.get("n_present"),
                          "docs_scanned": defect_result.get("docs_scanned"),
                          "photos_scanned": defect_result.get("photos_scanned")})
        except Exception as exc:  # noqa: BLE001
            _log("ai", f"AI: авто-проверка брака не удалась: {exc}", level="warn", case_id=case_id)
    _log("validator", "AI применён через validator", level="ok" if updated.get("ready_for_export") else "warn", case_id=case_id, subject=email_data.get("subject"), details={"state": updated.get("state"), "ready_for_export": bool(updated.get("ready_for_export")), "missing": updated.get("missing") or [], "quality": updated.get("quality") or []})
    return {
        "ok": True,
        "case_id": case_id,
        "applied": True,
        "defect": defect_result,
        "cached": bool(suggestion.get("cached")),
        "model": suggestion.get("model"),
        "state": updated.get("state"),
        "ready_for_export": bool(updated.get("ready_for_export")),
        "missing": updated.get("missing") or [],
        "quality": updated.get("quality") or [],
        "learning": learning_result,
    }


def _select_ai_candidate_ids(limit: int = 10) -> list[int]:
    if limit <= 0 or not getattr(settings, "enable_ai", False):
        return []
    with connect() as con:
        rows = con.execute(
            """
            SELECT c.id
            FROM cases c
            JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.state IN ('needs_review','needs_link')
              AND c.state NOT IN ('closed','ignored_internal','context_sent')
              AND c.event_type IN ('new_return','unknown','correction_request','marking_request')
              AND NOT EXISTS (
                SELECT 1 FROM ai_suggestions s
                WHERE s.case_id=c.id AND s.accepted=1
              )
            ORDER BY
              CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
              COALESCE(c.deadline_at, e.received_at) ASC,
              c.id ASC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
    return [int(r["id"]) for r in rows]


def _ai_decision_snapshot(limit: int = 20) -> dict[str, Any]:
    """Explain why AI is or is not doing work right now."""
    with connect() as con:
        total_cases = con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"]
        state_rows = con.execute("SELECT state, COUNT(*) c FROM cases GROUP BY state ORDER BY c DESC").fetchall()
        candidate_rows = con.execute(
            """
            SELECT c.id, c.state, c.event_type, c.claim_kind, c.priority, c.missing_json, c.quality_json,
                   e.subject, e.from_addr,
                   (SELECT COUNT(*) FROM ai_suggestions s WHERE s.case_id=c.id) ai_attempts,
                   (SELECT COUNT(*) FROM ai_suggestions s WHERE s.case_id=c.id AND s.accepted=1) ai_accepted
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.state IN ('needs_review','needs_link')
            ORDER BY
              CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
              COALESCE(c.deadline_at, e.received_at) ASC, c.id ASC
            LIMIT ?
            """,
            (int(limit),),
        ).fetchall()
        usage_rows = con.execute(
            """
            SELECT provider, model, ok, cached, COUNT(*) count,
                   SUM(prompt_chars) prompt_chars, SUM(response_chars) response_chars,
                   MAX(created_at) last_at
            FROM ai_usage
            GROUP BY provider, model, ok, cached
            ORDER BY MAX(created_at) DESC, count DESC
            LIMIT 20
            """
        ).fetchall()
        error_rows = con.execute(
            """
            SELECT id, case_id, provider, model, error, created_at
            FROM ai_usage
            WHERE ok=0 AND COALESCE(error,'') <> ''
            ORDER BY id DESC LIMIT 20
            """
        ).fetchall()
        suggestions = con.execute(
            """
            SELECT accepted, COUNT(*) count FROM ai_suggestions GROUP BY accepted ORDER BY accepted DESC
            """
        ).fetchall()
    candidates = []
    for r in candidate_rows:
        candidates.append({
            "case_id": int(r["id"]),
            "state": r["state"],
            "event_type": r["event_type"],
            "claim_kind": r["claim_kind"],
            "priority": r["priority"],
            "missing": loads(r["missing_json"], []),
            "quality": loads(r["quality_json"], []),
            "subject": r["subject"],
            "from_addr": r["from_addr"],
            "ai_attempts": int(r["ai_attempts"] or 0),
            "ai_accepted": int(r["ai_accepted"] or 0),
        })
    enabled = bool(getattr(settings, "enable_ai", False))
    reasons = []
    if not enabled:
        reasons.append("ENABLE_AI=false")
    if total_cases == 0:
        reasons.append("в базе нет писем/кейсов после reset — запусти импорт")
    if enabled and total_cases and not candidates:
        reasons.append("нет кейсов в needs_review/needs_link для AI")
    return {
        "ok": True,
        "enabled": enabled,
        "provider": settings.ai_provider,
        "model": settings.ai_model,
        "base_url": settings.ai_base_url,
        "auto_apply_validated": getattr(settings, "auto_apply_ai_validated", False),
        "total_cases": int(total_cases or 0),
        "state_counts": [dict(x) for x in state_rows],
        "candidate_count_shown": len(candidates),
        "candidates": candidates,
        "usage": [dict(x) for x in usage_rows],
        "last_errors": [dict(x) for x in error_rows],
        "suggestions": [dict(x) for x in suggestions],
        "reasons": reasons,
        "recommendation": "Если есть candidates, но AI accepted=0 — смотри last_errors и raw_excerpt в Live. Если candidates=0 после reset — сначала импортируй письма.",
    }


def _select_full_ai_ids(limit: int) -> list[int]:
    """v2.1 «Сразу ИИ»: ИИ смотрит КАЖДОЕ письмо (любой event_type), а не только возвраты —
    чтобы претензия, спрятанная скелетом в «корректировке»/связках, не прошла мимо ИИ.
    Берём не-обработанные ИИ кейсы, новые первыми; пропускаем уже AI-обработанные,
    отправленные и внутренние."""
    ids: list[int] = []
    with connect() as con:
        rows = con.execute(
            """SELECT c.id, c.payload_json FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
               WHERE c.state NOT IN ('exported','closed','delivered','ignored_internal')
               ORDER BY e.received_at DESC, c.id DESC LIMIT ?""",
            (max(1, limit) * 6,),
        ).fetchall()
    for r in rows:
        try:
            pj = loads(r["payload_json"] or "{}")
        except Exception:
            pj = {}
        if pj.get("ai_overlay") or pj.get("processing_source") == "ai":
            continue
        ids.append(int(r["id"]))
        if len(ids) >= max(1, limit):
            break
    return ids


def _autopilot_cycle(*, import_limit: int = 50, ai_limit: int = 10, deliver: bool = False, mode: str = "full_ai") -> dict[str, Any]:
    """Один цикл автопилота. v2.1 AI-only: всегда mode='full_ai' —
    импорт → ИИ по ВСЕМ свежим возвратам (паттернов нет) → Сверка → 1С."""
    apply_runtime_settings()
    started = utcnow()
    _log("autopilot", f"Цикл автопилота ({mode}): старт", details={"mode": mode, "import_limit": import_limit, "ai_limit": ai_limit, "deliver": deliver, "enable_ai": settings.enable_ai})
    result: dict[str, Any] = {"ok": True, "started_at": started, "import": {}, "classify": {}, "ai": {}, "queue": {}, "delivery": {}}
    # Один писатель на всю длину цикла: ручные кнопки (Паттерны/AI/Обучение) ждут
    # очереди в _PIPELINE_LOCK, пока конвейер пишет в БД → нет "database is locked".
    _PIPELINE_LOCK.acquire()
    _AUTOPILOT_STATE["running_cycle"] = True
    _AUTOPILOT_STATE["last_cycle_started_at"] = started
    try:
        if not runtime_control.can_run("import"):
            result["import"] = {"ok": True, "skipped": True, "reason": "import_paused"}
            _log("import", "IMAP: пропущен (на паузе)", level="info", details=result["import"])
        elif import_limit > 0 and settings.imap_username and settings.imap_password:
            if not _IMPORT_LOCK.acquire(blocking=False):
                result["import"] = {"ok": False, "skipped": True, "reason": "import_already_running"}
            else:
                try:
                    _log("import", "IMAP: импорт писем", details={"limit": import_limit, "folders": settings.folders})
                    result["import"] = import_from_imap_raw(limit=import_limit)
                    _log("import", "IMAP: импорт завершён", level="ok" if result["import"].get("ok", True) else "warn", details=result["import"])
                    global _LAST_IMPORT_RESULT
                    _LAST_IMPORT_RESULT = {**result["import"], "finished_at": utcnow(), "source": "autopilot"}
                finally:
                    _IMPORT_LOCK.release()
        else:
            result["import"] = {"ok": True, "skipped": True, "reason": "imap_not_configured_or_import_limit_zero"}
            _log("import", "IMAP: пропущен", level="warn", details=result["import"])

        # ── Классификация паттернами: свежие raw → кейсы (звено конвейера) ──
        # Без этого шага импортированные письма никогда не становились кейсами в
        # автопилоте — приходилось жать кнопку «Паттерны» вручную.
        try:
            if not runtime_control.can_run("stage2"):
                result["classify"] = {"ok": True, "skipped": True, "reason": "stage2_paused"}
                _log("classifier", "Конвейер: классификация пропущена (пауза)", level="info", details=result["classify"])
            else:
                with connect() as con:
                    classified = _classify_pending(con, only_missing=True, stop_event=_AUTOPILOT_STOP)
                result["classify"] = {"ok": True, "classified": classified}
                _log("classifier", "Конвейер: классификация паттернами", level="ok", details=result["classify"])
        except Exception as exc:
            result["classify"] = {"ok": False, "error": str(exc)}
            _log("classifier", "Конвейер: ошибка классификации", level="error", details=result["classify"])

        # Выбор кандидатов ИИ: pattern — «неуверенный хвост»; full_ai — ВСЁ подряд (новые возвраты).
        # AI на паузе или выключен → не выбираем кандидатов и не зовём модель.
        if not runtime_control.can_run("ai"):
            ai_ids = []
            ai_snapshot = {"reasons": ["ai_paused_or_disabled"], "candidates": [], "usage": []}
            _ai_purpose = "autopilot_ai_apply"
            _log("ai", "AI: пропущен (пауза/выключен)", level="info", details={"enable_ai": bool(getattr(settings, "enable_ai", False))})
        elif mode == "full_ai":
            ai_ids = _select_full_ai_ids(ai_limit)
            ai_snapshot = {"reasons": ["full_ai: все свежие возвраты без ИИ"], "candidates": [], "usage": []}
            _ai_purpose = "autopilot_full_ai"
        else:
            ai_ids = _select_ai_candidate_ids(ai_limit)
            ai_snapshot = _ai_decision_snapshot(limit=10)
            _ai_purpose = "autopilot_ai_apply"
        _log("ai", f"AI ({mode}): выбраны кандидаты", level="info" if ai_ids else "warn", details={"enabled": bool(getattr(settings, "enable_ai", False)), "selected": len(ai_ids), "case_ids": ai_ids[:20], "why": ai_snapshot.get("reasons"), "candidate_preview": ai_snapshot.get("candidates", [])[:5], "usage": ai_snapshot.get("usage", [])[:5]})
        if not ai_ids:
            _log("ai", "AI: работы нет / не выбран ни один кейс", level="warn", details=ai_snapshot)
        ai_items: list[dict[str, Any]] = []
        for cid in ai_ids:
            if _AUTOPILOT_STOP.is_set():
                break
            try:
                ai_items.append(_apply_ai_to_case_id(cid, purpose=_ai_purpose))
            except Exception as exc:
                ai_items.append({"ok": False, "case_id": cid, "error": str(exc), "applied": False})
        result["ai"] = {
            "enabled": bool(getattr(settings, "enable_ai", False)),
            "selected": len(ai_ids),
            "processed": len(ai_items),
            "applied": sum(1 for x in ai_items if x.get("applied")),
            "errors": sum(1 for x in ai_items if not x.get("ok")),
            "items": ai_items[-20:],
        }

        # Санитарная проверка перед outbox
        _run_sanity_pass()
        _log("sanity", "Санитарная проверка завершена")

        with connect() as con:
            _log("outbox", "Outbox: сверка/создание событий контроля", details={"limit": 2000, "auto_queue_control_events": getattr(settings, "auto_queue_control_events", True)})
            if not runtime_control.can_run("outbox"):
                result["queue"] = {"ok": True, "skipped": True, "reason": "outbox_paused"}
                _log("outbox", "Outbox: постановка в очередь на паузе", level="info", details=result["queue"])
            else:
                result["queue"] = queue_control_events(con, limit=2000) if getattr(settings, "auto_queue_control_events", True) else {"ok": True, "skipped": True}
            _log("outbox", "Outbox: события контроля обновлены", level="ok", details=result["queue"])
            if (deliver or getattr(settings, "auto_deliver_outbox", False)) and runtime_control.can_run("delivery"):
                _log("outbox", "Outbox: доставка старт", details={"deliver": deliver, "auto_deliver_outbox": getattr(settings, "auto_deliver_outbox", False)})
                result["delivery"] = deliver_outbox_events(con, limit=500)
                _log("outbox", "Outbox: доставка завершена", level="ok" if result["delivery"].get("ok", True) else "warn", details=result["delivery"])
            else:
                result["delivery"] = {"ok": True, "skipped": True, "reason": "auto_delivery_disabled"}
                _log("outbox", "Outbox: доставка выключена", level="info", details=result["delivery"])
        with connect() as con:
            result["stats"] = _test_run_snapshot(con)
        result["finished_at"] = utcnow()
        _log("autopilot", "Цикл автопилота: финиш", level="ok" if result.get("ok") else "warn", details={"ai": result.get("ai"), "queue": result.get("queue"), "delivery": result.get("delivery"), "stats": result.get("stats")})
        # ── Telegram уведомления ──
        try:
            from .telegram import notify_cycle_done, notify_unresolved, notify_delivery_error
            notify_cycle_done(result)
            # Неразобранные
            if deliver:
                with connect() as con:
                    unresolved_rows = con.execute(
                        """SELECT c.id, c.buyer_name, c.missing_json, e.subject
                           FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
                           WHERE c.state='needs_review' AND c.event_type='new_return'
                           ORDER BY e.received_at DESC LIMIT 30"""
                    ).fetchall()
                    unresolved = [{"case_id": r["id"], "buyer_name": r["buyer_name"], "subject": r["subject"], "missing": loads(r["missing_json"] or "[]")} for r in unresolved_rows]
                notify_unresolved(unresolved)
            # Ошибки доставки
            with connect() as con:
                err_rows = con.execute("SELECT id, case_id, last_error FROM outbox WHERE status='error' LIMIT 10").fetchall()
                if err_rows:
                    notify_delivery_error([dict(r) for r in err_rows])
        except Exception:
            pass
        return result
    except Exception as exc:
        result.update({"ok": False, "error": str(exc), "finished_at": utcnow()})
        _log("autopilot", "Цикл автопилота: ошибка", level="error", details={"error": str(exc)})
        return result
    finally:
        _AUTOPILOT_STATE["running_cycle"] = False
        _AUTOPILOT_STATE["last_cycle_finished_at"] = utcnow()
        _PIPELINE_LOCK.release()


def _autopilot_loop() -> None:
    while not _AUTOPILOT_STOP.is_set():
        cfg = dict(_AUTOPILOT_STATE.get("config") or {})
        cycle = _autopilot_cycle(
            import_limit=int(cfg.get("import_limit") or 50),
            ai_limit=int(cfg.get("ai_limit") or 10),
            deliver=bool(cfg.get("deliver")),
            mode="full_ai",  # v2.1 AI-only: режим паттернов отключён
        )
        _AUTOPILOT_STATE["last_cycle"] = cycle
        _AUTOPILOT_STATE["cycle_count"] = int(_AUTOPILOT_STATE.get("cycle_count") or 0) + 1
        _AUTOPILOT_STATE["last_error"] = cycle.get("error") if not cycle.get("ok") else None
        interval = max(30, int(cfg.get("interval_seconds") or 300))
        # Small sleeps keep Stop responsive.
        from datetime import datetime, timezone, timedelta
        _AUTOPILOT_STATE["next_cycle_at"] = (datetime.now(timezone.utc).replace(microsecond=0) + timedelta(seconds=interval)).isoformat()
        for _ in range(interval):
            if _AUTOPILOT_STOP.is_set():
                break
            time.sleep(1)
    _AUTOPILOT_STATE["enabled"] = False
    _AUTOPILOT_STATE["stopped_at"] = utcnow()
    _AUTOPILOT_STATE["next_cycle_at"] = None

def _scan_loop() -> None:
    while True:
        try:
            apply_runtime_settings()
            interval = int(settings.scan_interval_seconds or 0)
            if interval > 0 and settings.imap_username and settings.imap_password:
                _safe_import_cycle()
                time.sleep(max(15, interval))
            else:
                time.sleep(15)
        except Exception as exc:
            _log("import", "Ошибка scan_loop", level="error", details={"error": str(exc)})
            time.sleep(30)


def _start_scan_thread_once() -> None:
    global _SCAN_THREAD_STARTED
    if _SCAN_THREAD_STARTED:
        return
    th = threading.Thread(target=_scan_loop, name="readmail-auto-import", daemon=True)
    th.start()
    _SCAN_THREAD_STARTED = True


def _check_item(code: str, title: str, status: str, message: str, *, details: dict[str, Any] | None = None) -> dict[str, Any]:
    """Build a uniform launch/self-check item.

    status: ok | warn | error | info
    """
    return {"code": code, "title": title, "status": status, "message": message, "details": details or {}}


def _system_self_check(*, live: bool = False) -> dict[str, Any]:
    """Pre-flight checklist for real-mail launch.

    This endpoint is deliberately conservative: it does not require the operator to open .env
    and it does not silently change data. Live checks are opt-in because IMAP/AI/HTTP can be slow.
    """
    apply_runtime_settings()
    checks: list[dict[str, Any]] = []

    # Database and control coverage.
    try:
        with connect() as con:
            schema = con.execute("SELECT value FROM meta WHERE key='schema_version'").fetchone()
            con.execute("CREATE TABLE IF NOT EXISTS __readmail_write_probe(x INTEGER)")
            con.execute("DELETE FROM __readmail_write_probe")
            con.execute("INSERT INTO __readmail_write_probe(x) VALUES (1)")
            raw_count = con.execute("SELECT COUNT(*) c FROM raw_emails").fetchone()["c"]
            case_count = con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"]
            raw_without_case = con.execute(
                "SELECT COUNT(*) c FROM raw_emails r LEFT JOIN cases c ON c.raw_email_id=r.id WHERE c.id IS NULL"
            ).fetchone()["c"]
            relevant_cases_without_event = con.execute(
                """
                SELECT COUNT(*) c
                FROM cases c
                WHERE c.state NOT IN ('ignored_internal','closed')
                  AND NOT EXISTS (SELECT 1 FROM outbox o WHERE o.case_id=c.id)
                """
            ).fetchone()["c"]
            outbox_errors = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='error'").fetchone()["c"]
            outbox_new = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='new'").fetchone()["c"]
            checks.append(_check_item(
                "database", "База данных", "ok",
                f"SQLite доступна, schema={schema['value'] if schema else 'unknown'}, писем={raw_count}, кейсов={case_count}.",
                details={"raw_emails": raw_count, "cases": case_count, "schema_version": schema["value"] if schema else None},
            ))
            checks.append(_check_item(
                "case_coverage", "Письма → кейсы",
                "ok" if raw_without_case == 0 else "error",
                "Все импортированные письма имеют кейс." if raw_without_case == 0 else f"Есть писем без кейса: {raw_without_case}. Нажми пересчёт.",
                details={"raw_without_case": raw_without_case},
            ))
            checks.append(_check_item(
                "outbox_coverage", "Кейсы → события контроля",
                "ok" if relevant_cases_without_event == 0 else "warn",
                "У всех значимых кейсов есть событие outbox." if relevant_cases_without_event == 0 else f"Есть кейсов без события для 1С/контроля: {relevant_cases_without_event}. Нажми 'Сверить очередь'.",
                details={"relevant_cases_without_event": relevant_cases_without_event, "outbox_new": outbox_new, "outbox_errors": outbox_errors},
            ))
            checks.append(_check_item(
                "outbox_errors", "Ошибки доставки", "ok" if outbox_errors == 0 else "error",
                "Ошибок доставки нет." if outbox_errors == 0 else f"Есть ошибок доставки: {outbox_errors}. Нажми 'Повторить ошибки' или посмотри журнал.",
                details={"outbox_errors": outbox_errors, "outbox_new": outbox_new},
            ))
    except Exception as exc:
        checks.append(_check_item("database", "База данных", "error", f"Ошибка SQLite: {exc}"))

    # IMAP configuration.
    folders = settings.folders
    imap_configured = bool(settings.imap_host and settings.imap_username and settings.imap_password and folders)
    checks.append(_check_item(
        "imap_config", "Почта / IMAP", "ok" if imap_configured else "error",
        "IMAP настроен." if imap_configured else "Заполни в панели IMAP username/password/folders.",
        details={"host": settings.imap_host, "port": settings.imap_port, "username_set": bool(settings.imap_username), "password_set": bool(settings.imap_password), "folders": folders},
    ))
    if live and imap_configured:
        try:
            folders_result = list_imap_folders()
            ok = bool(folders_result.get("ok", True)) and not folders_result.get("error")
            checks.append(_check_item(
                "imap_live", "Почта / live test", "ok" if ok else "error",
                f"IMAP отвечает, папок найдено: {folders_result.get('count', len(folders_result.get('folders', [])))}." if ok else f"IMAP ошибка: {folders_result.get('error')}",
                details=folders_result,
            ))
        except Exception as exc:
            checks.append(_check_item("imap_live", "Почта / live test", "error", f"IMAP ошибка: {exc}"))

    # AI configuration.
    if settings.enable_ai:
        ai_basic_ok = bool(settings.ai_provider and settings.ai_model and (settings.ai_base_url or settings.ai_provider == "gigachat"))
        if settings.ai_provider == "gigachat":
            ai_basic_ok = ai_basic_ok and bool(getattr(settings, "gigachat_auth_key", ""))
        else:
            ai_basic_ok = ai_basic_ok and bool(settings.ai_api_key)
        checks.append(_check_item(
            "ai_config", "AI / модель", "ok" if ai_basic_ok else "error",
            "AI включён и базово настроен." if ai_basic_ok else "AI включён, но не хватает URL/model/key.",
            details={"provider": settings.ai_provider, "model": settings.ai_model, "base_url": settings.ai_base_url, "key_set": bool(settings.ai_api_key or getattr(settings, 'gigachat_auth_key', ''))},
        ))
        if live and ai_basic_ok:
            try:
                ai_result = test_ai_connection()
                checks.append(_check_item(
                    "ai_live", "AI / live test", "ok" if ai_result.get("ok") else "error",
                    "AI отвечает." if ai_result.get("ok") else f"AI ошибка: {ai_result.get('error') or ai_result}",
                    details=ai_result,
                ))
            except Exception as exc:
                checks.append(_check_item("ai_live", "AI / live test", "error", f"AI ошибка: {exc}"))
    else:
        checks.append(_check_item("ai_config", "AI / модель", "warn", "AI выключен. Для первого прогона можно, но новые/сложные клиенты будут чаще попадать в needs_review."))

    # 1C/outbox configuration.
    mode = str(settings.one_c_export_mode or "off").lower()
    if mode == "off":
        checks.append(_check_item("one_c_mode", "1С / Outbox", "warn", "Экспорт в 1С выключен. События будут видны в панели, но не будут уходить наружу."))
    else:
        if mode in {"file", "both"}:
            try:
                root = Path(settings.one_c_file_dir)
                root.mkdir(parents=True, exist_ok=True)
                probe = root / ".readmail_write_probe"
                probe.write_text("ok", encoding="utf-8")
                probe.unlink(missing_ok=True)
                checks.append(_check_item("one_c_file", "1С / file outbox", "ok", f"Папка JSON доступна: {root}", details={"dir": str(root)}))
            except Exception as exc:
                checks.append(_check_item("one_c_file", "1С / file outbox", "error", f"Не могу писать JSON в папку: {exc}", details={"dir": str(settings.one_c_file_dir)}))
        if mode in {"http", "both"}:
            url_ok = bool(str(settings.one_c_http_url or "").strip())
            checks.append(_check_item(
                "one_c_http", "1С / HTTP", "ok" if url_ok else "error",
                "HTTP endpoint задан." if url_ok else "ONE_C_HTTP_URL пустой.",
                details={"url_set": url_ok, "token_set": bool(settings.one_c_http_token)},
            ))

    errors = sum(1 for c in checks if c["status"] == "error")
    warnings = sum(1 for c in checks if c["status"] == "warn")
    ok_count = sum(1 for c in checks if c["status"] == "ok")
    ready_for_real_import = errors == 0 and bool(settings.imap_username and settings.imap_password)
    return {
        "ok": errors == 0,
        "schema": "readmail-new-self-check-v1.10",
        "generated_at": utcnow(),
        "live": live,
        "ready_for_real_import": ready_for_real_import,
        "summary": {"ok": ok_count, "warnings": warnings, "errors": errors, "total": len(checks)},
        "checks": checks,
        "next_steps": _self_check_next_steps(checks),
    }


def _self_check_next_steps(checks: list[dict[str, Any]]) -> list[str]:
    steps: list[str] = []
    by_code = {c["code"]: c for c in checks}
    if by_code.get("imap_config", {}).get("status") == "error":
        steps.append("Открой Настройки → Почта и заполни IMAP-логин, пароль приложения и клиентские папки.")
    if by_code.get("case_coverage", {}).get("status") == "error":
        steps.append("Нажми 'Пересчитать', чтобы создать кейсы для писем без карточек.")
    if by_code.get("outbox_coverage", {}).get("status") == "warn":
        steps.append("Открой 1С / Outbox и нажми 'Сверить очередь', чтобы создать события для всех значимых кейсов.")
    if by_code.get("outbox_errors", {}).get("status") == "error":
        steps.append("Открой 1С / Outbox: проверь last_error и нажми 'Повторить ошибки'.")
    if by_code.get("ai_config", {}).get("status") == "warn":
        steps.append("AI можно оставить выключенным для сухого теста; для новых клиентов включи AI / MLX в панели.")
    if by_code.get("one_c_mode", {}).get("status") == "warn":
        steps.append("Включи ONE_C_EXPORT_MODE=file для теста JSON-файлов или both/http для интеграции.")
    return steps or ["Можно запускать демо-тест или импорт реальных писем небольшим лимитом."]



def _count_rows(con: Any, sql: str, params: tuple[Any, ...] = ()) -> int:
    try:
        return int(con.execute(sql, params).fetchone()["c"])
    except Exception:
        return 0


def _test_run_snapshot(con: Any) -> dict[str, Any]:
    """Small stable snapshot before/after a real-mail test run."""
    return {
        "raw_emails": _count_rows(con, "SELECT COUNT(*) c FROM raw_emails"),
        "cases": _count_rows(con, "SELECT COUNT(*) c FROM cases"),
        "max_case_id": _count_rows(con, "SELECT COALESCE(MAX(id),0) c FROM cases"),
        "outbox": _count_rows(con, "SELECT COUNT(*) c FROM outbox"),
        "outbox_new": _count_rows(con, "SELECT COUNT(*) c FROM outbox WHERE status='new'"),
        "outbox_sent": _count_rows(con, "SELECT COUNT(*) c FROM outbox WHERE status='sent'"),
        "outbox_error": _count_rows(con, "SELECT COUNT(*) c FROM outbox WHERE status='error'"),
        "ai_usage": _count_rows(con, "SELECT COUNT(*) c FROM ai_usage"),
    }


def _bump_counter(counter: dict[str, int], key: Any) -> None:
    key = str(key or "empty")
    counter[key] = counter.get(key, 0) + 1


def _case_quality_report(con: Any, *, since_case_id: int = 0, limit: int = 500) -> dict[str, Any]:
    """Report what the first real test actually produced.

    This intentionally looks for dangerous cases, not just totals: ready rows with missing fields,
    strange part numbers, new returns without strong keys, and relevant cases missing control events.
    """
    params: list[Any] = []
    where = ""
    if since_case_id:
        where = "WHERE c.id > ?"
        params.append(int(since_case_id))
    rows = [
        row_to_dict(r)
        for r in con.execute(
            f"""
            SELECT c.*, e.subject, e.from_addr, e.direction, e.received_at, e.snippet, e.mailbox
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            {where}
            ORDER BY c.id DESC
            LIMIT ?
            """,
            params + [int(limit)],
        ).fetchall()
    ]
    state_counts: dict[str, int] = {}
    event_counts: dict[str, int] = {}
    claim_counts: dict[str, int] = {}
    priority_counts: dict[str, int] = {}
    missing_counts: dict[str, int] = {}
    quality_counts: dict[str, int] = {}
    red_flags: list[dict[str, Any]] = []
    ready = 0
    needs_review = 0
    suspicious_words = {
        "цена", "сумма", "причина", "товар", "артикул", "количество", "поставщик", "покупатель",
        "брак", "недовоз", "пересорт", "некондиция", "наименование", "описание", "комментарий",
    }
    for item in rows:
        _bump_counter(state_counts, item.get("state"))
        _bump_counter(event_counts, item.get("event_type"))
        _bump_counter(claim_counts, item.get("claim_kind"))
        _bump_counter(priority_counts, item.get("priority"))
        if item.get("ready_for_export"):
            ready += 1
        if item.get("needs_review"):
            needs_review += 1
        for m in item.get("missing") or []:
            _bump_counter(missing_counts, m)
        q_errors = []
        for q in item.get("quality") or []:
            code = q.get("code") or q.get("message") or q.get("level")
            if code:
                _bump_counter(quality_counts, code)
            if q.get("level") == "error":
                q_errors.append(q)
        fields = item.get("fields") or {}
        part = str(fields.get("part_number") or "").strip()
        direction = item.get("direction") or ((item.get("payload") or {}).get("direction"))
        if item.get("ready_for_export") and (item.get("event_type") != "new_return" or direction != "inbound_customer"):
            red_flags.append({"case_id": item.get("id"), "code": "ready_not_new_inbound", "subject": item.get("subject"), "event_type": item.get("event_type"), "direction": direction})
        if item.get("ready_for_export") and (item.get("missing") or q_errors):
            red_flags.append({"case_id": item.get("id"), "code": "ready_has_missing_or_quality_errors", "subject": item.get("subject"), "missing": item.get("missing") or [], "quality_errors": q_errors[:5]})
        if item.get("event_type") == "new_return" and not item.get("strong_key"):
            red_flags.append({"case_id": item.get("id"), "code": "new_return_without_strong_key", "subject": item.get("subject"), "from_addr": item.get("from_addr")})
        if part:
            p_l = part.lower()
            if p_l in suspicious_words or len(part) < int(settings.part_number_min_len or 3) or len(part) > int(settings.part_number_max_len or 50):
                red_flags.append({"case_id": item.get("id"), "code": "suspicious_part_number", "part_number": part, "subject": item.get("subject")})
    raw_without_case = _count_rows(con, "SELECT COUNT(*) c FROM raw_emails r LEFT JOIN cases c ON c.raw_email_id=r.id WHERE c.id IS NULL")
    cases_without_outbox = _count_rows(con, """
        SELECT COUNT(*) c
        FROM cases c
        WHERE c.state NOT IN ('ignored_internal','closed')
          AND NOT EXISTS (SELECT 1 FROM outbox o WHERE o.case_id=c.id)
    """)
    if raw_without_case:
        red_flags.append({"code": "raw_emails_without_cases", "count": raw_without_case})
    if cases_without_outbox:
        red_flags.append({"code": "cases_without_control_outbox", "count": cases_without_outbox})
    return {
        "schema": "readmail-new-test-quality-report-v1.12",
        "generated_at": utcnow(),
        "scope": {"since_case_id": since_case_id, "sampled_cases": len(rows), "limit": limit},
        "summary": {
            "sampled_cases": len(rows),
            "ready_for_export": ready,
            "needs_review": needs_review,
            "red_flags": len(red_flags),
            "raw_without_case": raw_without_case,
            "cases_without_control_outbox": cases_without_outbox,
        },
        "state_counts": state_counts,
        "event_counts": event_counts,
        "claim_counts": claim_counts,
        "priority_counts": priority_counts,
        "missing_counts": missing_counts,
        "quality_counts": quality_counts,
        "red_flags": red_flags[:100],
        "cases_sample": rows[: min(50, len(rows))],
        "recommendation": _test_run_recommendation(red_flags, ready, needs_review),
    }


def _test_run_recommendation(red_flags: list[dict[str, Any]], ready: int, needs_review: int) -> str:
    if red_flags:
        return "Не включать автоотправку в 1С: сначала разобрать red_flags и усилить правила/валидатор."
    if ready == 0 and needs_review > 0:
        return "Импорт безопасный, но всё ушло в review: подключи AI/MLX или добавь правила по клиентам/документам."
    if ready > 0 and needs_review == 0:
        return "Тест чистый: можно расширить лимит импорта, но AUTO_DELIVER_OUTBOX держать выключенным до проверки 1С."
    return "Тест пригоден для ручной выборочной проверки: посмотри ready, needs_review и outbox journal."


def _save_test_run(con: Any, *, started_at: str, status: str, params: dict[str, Any], before: dict[str, Any], after: dict[str, Any], import_result: dict[str, Any], queue_result: dict[str, Any], delivery_result: dict[str, Any] | None, summary: dict[str, Any], error: str | None = None) -> int:
    cur = con.execute(
        """
        INSERT INTO test_runs(started_at, finished_at, status, params_json, before_json, after_json, import_result_json, queue_result_json, delivery_result_json, summary_json, error)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (started_at, utcnow(), status, dumps(params), dumps(before), dumps(after), dumps(import_result), dumps(queue_result), dumps(delivery_result or {}), dumps(summary), error),
    )
    return int(cur.lastrowid)




@app.on_event("startup")
def startup() -> None:
    init_db()
    apply_runtime_settings()
    _start_scan_thread_once()
    try:
        from .telegram import start_hourly_reporter
        start_hourly_reporter()
    except Exception:
        pass
    # Баннер запуска (без секретов): Local/LAN URL, auth, workers, developer mode.
    try:
        print(server_core.startup_banner(runtime_control.get_runtime_status()), flush=True)
        if server_core.auth_required() and not server_core.admin_configured():
            print("⚠️  Auth включён, но admin не создан — создайте admin перед LAN-доступом.", flush=True)
    except Exception:
        pass


@app.get("/", response_class=HTMLResponse)
def index() -> HTMLResponse:
    """Serve index with auto-versioned static assets (cache-bust by file mtime)."""
    html = (WEB_DIR / "index.html").read_text(encoding="utf-8")
    try:
        css_v = int((WEB_DIR / "static" / "style.css").stat().st_mtime)
        js_v = int((WEB_DIR / "static" / "app.js").stat().st_mtime)
        html = re.sub(r'style\.css\?v=[^"]*', f'style.css?v={css_v}', html)
        html = re.sub(r'app\.js\?v=[^"]*', f'app.js?v={js_v}', html)
    except Exception:
        pass
    # index.html не кэшировать: иначе браузер крутит старую страницу со старой ?v= → не видит правок
    # фронта (это и был «нужен Option+Cmd+R»). Сам app.js/css кэшируются и бьются по mtime в ?v=.
    return HTMLResponse(html, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.get("/api/health")
def health() -> dict[str, Any]:
    apply_runtime_settings()
    return {
        "ok": True,
        "app": settings.app_name,
        "version": settings.app_version,
        "database": str(settings.database_path),
        "imap_folders": settings.folders,
        "discover_all_folders": settings.discover_all_folders,
        "company_domains_configured": bool(settings.company_domain_list),
        "enable_ai": settings.enable_ai,
        "ai_provider": settings.ai_provider,
        "auto_learn_unknown_buyers": settings.auto_learn_unknown_buyers,
        "auto_promote_unknown_buyer_after": settings.auto_promote_unknown_buyer_after,
        "one_c_export_mode": settings.one_c_export_mode,
        "scan_interval_seconds": settings.scan_interval_seconds,
        "last_import": _LAST_IMPORT_RESULT,
    }


# ── Локальный приёмник 1С (контрольная точка обмена, НЕ реальная внешняя 1С) ──
class _Local1CIn(BaseModel):
    payload: dict[str, Any] | None = None
    case_id: int | None = None
    profile: str = "standard"


def _local_1c_receive(body: _Local1CIn) -> dict[str, Any]:
    from . import local_1c
    payload = body.payload
    if payload is None and body.case_id is not None:
        payload = local_1c.build_payload_for_case(body.case_id, profile=body.profile)
        if not payload:
            raise HTTPException(status_code=404, detail="case_not_found")
    if not payload:
        raise HTTPException(status_code=400, detail="empty_payload")
    return local_1c.receive_payload(payload, source="api")


@app.post("/api/local-1c/receive")
def api_local_1c_receive(body: _Local1CIn) -> dict[str, Any]:
    """Локальный приёмник 1С: приём пакета в журнал обмена audit_out/local_1c_receiver.jsonl.

    НЕ вызывает реальную внешнюю 1С, НЕ меняет real outbox/статусы, НЕ делает внешний HTTP.
    Можно передать готовый `payload`, либо `case_id`(+`profile`) — тогда payload строится из кейса.
    """
    return _local_1c_receive(body)


@app.get("/api/local-1c/events")
def api_local_1c_events(limit: int = Query(default=20, ge=1, le=200),
                        include_payload: bool = Query(default=False)) -> dict[str, Any]:
    """Последние N пакетов локального приёмника 1С (read-only)."""
    from . import local_1c
    return local_1c.get_events(limit=limit, include_payload=include_payload)


@app.get("/api/local-1c/status")
def api_local_1c_status() -> dict[str, Any]:
    from . import local_1c
    return local_1c.receiver_status()


# Back-compat alias-эндпоинты (старые demo-маршруты) — без публикации в UI.
@app.post("/api/demo-1c/receive", include_in_schema=False)
def api_demo_1c_receive(body: _Local1CIn) -> dict[str, Any]:
    return _local_1c_receive(body)


@app.get("/api/demo-1c/events", include_in_schema=False)
def api_demo_1c_events(limit: int = Query(default=20, ge=1, le=200),
                       include_payload: bool = Query(default=False)) -> dict[str, Any]:
    from . import local_1c
    return local_1c.get_events(limit=limit, include_payload=include_payload)


@app.get("/api/demo-1c/status", include_in_schema=False)
def api_demo_1c_status() -> dict[str, Any]:
    from . import local_1c
    return local_1c.receiver_status()


# ── Decision trace (Visual Accounting + Safety Router, read-only) ──────
@app.get("/api/cases/{case_id}/decision")
def api_case_decision(case_id: int) -> dict[str, Any]:
    from . import visual_accounting as va
    with connect() as con:
        res = va.decision_for_case(con, case_id)
    if not res.get("ok"):
        raise HTTPException(status_code=404, detail=res.get("error", "not_found"))
    return res


@app.get("/api/raw-emails/{raw_email_id}/decision")
def api_raw_decision(raw_email_id: int) -> dict[str, Any]:
    from . import visual_accounting as va
    with connect() as con:
        res = va.decision_for_raw(con, raw_email_id)
    if not res.get("ok"):
        raise HTTPException(status_code=404, detail=res.get("error", "not_found"))
    return res


# ── Hidden Processed Mail (раздел «Обработанные / не требуют действия», read-only) ──
@app.get("/api/processed-hidden/summary")
def api_processed_hidden_summary() -> dict[str, Any]:
    from . import processed_hidden as ph
    with connect() as con:
        return ph.build_processed_hidden_summary(con)


@app.get("/api/processed-hidden/items")
def api_processed_hidden_items(group: str | None = Query(default=None),
                               subcategory: str | None = Query(default=None),
                               page: int = Query(default=1, ge=1),
                               page_size: int = Query(default=50, ge=1, le=500),
                               q: str = Query(default="")) -> dict[str, Any]:
    from . import processed_hidden as ph
    with connect() as con:
        return ph.list_processed_hidden_items(con, group=group, subcategory=subcategory,
                                              page=page, page_size=page_size, q=q)


# ── Canonical Pipeline (read-only) ─────────────────────────────────────
@app.get("/api/pipeline/accounting")
def api_pipeline_accounting() -> dict[str, Any]:
    from . import canonical_pipeline as cp
    with connect() as con:
        return cp.build_pipeline_accounting(con)


@app.get("/api/pipeline/items")
def api_pipeline_items(route: str | None = Query(default=None),
                       reason: str | None = Query(default=None),
                       page: int = Query(default=1, ge=1),
                       page_size: int = Query(default=50, ge=1, le=500),
                       q: str = Query(default="")) -> dict[str, Any]:
    from . import canonical_pipeline as cp
    with connect() as con:
        return cp.list_pipeline_items(con, route=route, reason=reason, page=page, page_size=page_size, q=q)


@app.get("/api/system/self-check")
def api_system_self_check(live: bool = Query(default=False)) -> dict[str, Any]:
    return _system_self_check(live=live)


@app.post("/api/system/self-check")
def api_system_self_check_post(live: bool = Query(default=True)) -> dict[str, Any]:
    return _system_self_check(live=live)


@app.post("/api/system/repair-control-coverage")
def api_system_repair_control_coverage(limit: int = Query(default=1000, ge=1, le=10000)) -> dict[str, Any]:
    apply_runtime_settings()
    with connect() as con:
        queued = queue_control_events(con, limit=limit)
        dashboard = outbox_dashboard(con, limit=50)
    return {"ok": True, "queued": queued, "dashboard": dashboard, "self_check": _system_self_check(live=False)}


@app.post("/api/system/smoke-test")
def api_system_smoke_test() -> dict[str, Any]:
    """Create demo data, queue control events, deliver file outbox, and return a readiness report."""
    apply_runtime_settings()
    demo = generate_demo_data(queue=True)
    with connect() as con:
        queued = queue_control_events(con, limit=1000)
        delivery = deliver_outbox_events(con, limit=200, channel="file") if str(settings.one_c_export_mode or "file") in {"file", "both"} else {"ok": True, "skipped": True, "reason": "file_outbox_disabled"}
        control = control_dashboard(con, limit=50)
        outbox = outbox_dashboard(con, limit=50)
    return {"ok": True, "demo": demo, "queued": queued, "delivery": delivery, "control": control, "outbox": outbox, "self_check": _system_self_check(live=False)}


@app.get("/api/system/traffic-stats")
def api_system_traffic_stats() -> dict[str, Any]:
    """Email/API/outbox volume counters for top-bar badges and settings."""
    apply_runtime_settings()
    total_emails = 0
    total_cases = 0
    mail_raw_bytes = 0
    mail_body_bytes = 0
    attachment_db_bytes = 0
    attachment_disk_bytes = 0
    ai_prompt_bytes = 0
    ai_response_bytes = 0
    ai_requests = 0
    outbox_payload_bytes = 0
    outbox_file_bytes = 0
    db_file_bytes = 0
    outbox_new = 0
    outbox_sent = 0
    outbox_error = 0

    def _dir_size(path: Path) -> int:
        total = 0
        try:
            if not path.exists():
                return 0
            for item in path.rglob("*"):
                try:
                    if item.is_file():
                        total += item.stat().st_size
                except Exception:
                    pass
        except Exception:
            return 0
        return total

    try:
        with connect() as con:
            try:
                total_emails = con.execute("SELECT COUNT(*) c FROM raw_emails").fetchone()["c"]
                total_cases = con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"]
            except Exception:
                pass
            try:
                row = con.execute(
                    """
                    SELECT
                      COALESCE(SUM(raw_size),0) raw_bytes,
                      COALESCE(SUM(LENGTH(CAST(COALESCE(body_text,'') AS BLOB))),0)
                    + COALESCE(SUM(LENGTH(CAST(COALESCE(body_html,'') AS BLOB))),0)
                    + COALESCE(SUM(LENGTH(CAST(COALESCE(snippet,'') AS BLOB))),0) body_bytes
                    FROM raw_emails
                    """
                ).fetchone()
                mail_raw_bytes = int(row["raw_bytes"] or 0)
                mail_body_bytes = int(row["body_bytes"] or 0)
            except Exception:
                pass
            try:
                attachment_db_bytes = int(con.execute("SELECT COALESCE(SUM(size_bytes),0) s FROM attachments").fetchone()["s"] or 0)
            except Exception:
                pass
            try:
                row = con.execute(
                    "SELECT COUNT(*) c, COALESCE(SUM(prompt_chars),0) p, COALESCE(SUM(response_chars),0) r FROM ai_usage"
                ).fetchone()
                ai_requests = int(row["c"] or 0)
                ai_prompt_bytes = int(row["p"] or 0)
                ai_response_bytes = int(row["r"] or 0)
            except Exception:
                pass
            try:
                outbox_payload_bytes = int(con.execute(
                    """
                    SELECT COALESCE(SUM(LENGTH(CAST(COALESCE(payload_json,'') AS BLOB))),0)
                         + COALESCE(SUM(LENGTH(CAST(COALESCE(delivery_response_json,'') AS BLOB))),0) s
                    FROM outbox
                    """
                ).fetchone()["s"] or 0)
            except Exception:
                pass
            try:
                outbox_new = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='new'").fetchone()["c"]
            except Exception:
                pass
            try:
                outbox_sent = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='sent'").fetchone()["c"]
            except Exception:
                pass
            try:
                outbox_error = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='error'").fetchone()["c"]
            except Exception:
                pass
    except Exception:
        pass
    attachment_disk_bytes = _dir_size(BASE_DIR.parent / "data" / "attachments")
    outbox_file_bytes = _dir_size(Path(str(getattr(settings, "one_c_file_dir", "") or "/app/data/outbox_1c")).relative_to("/app") if str(getattr(settings, "one_c_file_dir", "") or "").startswith("/app/") else BASE_DIR.parent / "data" / "outbox_1c")
    try:
        db_file_bytes = (BASE_DIR.parent / "data" / "readmail.sqlite3").stat().st_size
    except Exception:
        db_file_bytes = 0
    mail_transfer_bytes = mail_raw_bytes + attachment_db_bytes
    ai_bytes = ai_prompt_bytes + ai_response_bytes
    one_c_bytes = outbox_payload_bytes + outbox_file_bytes
    traffic_total_bytes = mail_transfer_bytes + ai_bytes + one_c_bytes
    storage_total_bytes = db_file_bytes + attachment_disk_bytes + outbox_file_bytes

    def _mb(n: int) -> float:
        return round((int(n or 0) / (1024 * 1024)), 2)

    return {
        "total_emails": total_emails,
        "total_cases": total_cases,
        "outbox_new": outbox_new,
        "outbox_sent": outbox_sent,
        "outbox_error": outbox_error,
        "mail": {
            "raw_bytes": mail_raw_bytes,
            "body_db_bytes": mail_body_bytes,
            "attachments_db_bytes": attachment_db_bytes,
            "attachments_disk_bytes": attachment_disk_bytes,
            "transfer_bytes": mail_transfer_bytes,
            "transfer_mb": _mb(mail_transfer_bytes),
        },
        "ai": {
            "requests": ai_requests,
            "prompt_bytes": ai_prompt_bytes,
            "response_bytes": ai_response_bytes,
            "total_bytes": ai_bytes,
            "total_mb": _mb(ai_bytes),
        },
        "one_c": {
            "payload_bytes": outbox_payload_bytes,
            "files_bytes": outbox_file_bytes,
            "total_bytes": one_c_bytes,
            "total_mb": _mb(one_c_bytes),
        },
        "storage": {
            "db_file_bytes": db_file_bytes,
            "attachments_disk_bytes": attachment_disk_bytes,
            "outbox_file_bytes": outbox_file_bytes,
            "total_bytes": storage_total_bytes,
            "total_mb": _mb(storage_total_bytes),
        },
        "total": {
            "traffic_bytes": traffic_total_bytes,
            "traffic_mb": _mb(traffic_total_bytes),
            "storage_bytes": storage_total_bytes,
            "storage_mb": _mb(storage_total_bytes),
        },
    }


@app.post("/api/test-run/start")
def api_test_run_start(
    limit: int = Query(default=30, ge=1, le=500),
    folder: str | None = None,
    search: str | None = None,
    deliver: bool = Query(default=False),
) -> dict[str, Any]:
    """Safe first real-mail run: import a small batch, queue control events, do not deliver by default, then report quality."""
    apply_runtime_settings()
    if not _IMPORT_LOCK.acquire(blocking=False):
        return {"ok": False, "skipped": True, "reason": "import_already_running"}
    started_at = utcnow()
    params = {"limit": limit, "folder": folder, "search": search, "deliver": deliver}
    try:
        with connect() as con:
            before = _test_run_snapshot(con)
        folders = [folder] if folder else None
        import_result = import_from_imap_raw(limit=limit, folders=folders, search=search)
        with connect() as con:
            queue_result = queue_control_events(con, limit=1000)
            delivery_result = deliver_outbox_events(con, limit=500) if deliver else {"ok": True, "skipped": True, "reason": "test_run_delivery_disabled"}
            after = _test_run_snapshot(con)
            report = _case_quality_report(con, since_case_id=int(before.get("max_case_id") or 0), limit=500)
            status = "error" if not import_result.get("ok", True) else ("ok" if not report["red_flags"] else "warn")
            run_id = _save_test_run(
                con,
                started_at=started_at,
                status=status,
                params=params,
                before=before,
                after=after,
                import_result=import_result,
                queue_result=queue_result,
                delivery_result=delivery_result,
                summary=report,
            )
        global _LAST_IMPORT_RESULT
        _LAST_IMPORT_RESULT = {**import_result, "test_run_id": run_id, "finished_at": utcnow()}
        return {"ok": True, "test_run_id": run_id, "status": status, "params": params, "before": before, "after": after, "import": import_result, "queue": queue_result, "delivery": delivery_result, "report": report}
    except Exception as exc:
        with connect() as con:
            before = before if 'before' in locals() else _test_run_snapshot(con)
            after = _test_run_snapshot(con)
            run_id = _save_test_run(
                con,
                started_at=started_at,
                status="error",
                params=params,
                before=before,
                after=after,
                import_result=locals().get("import_result", {}),
                queue_result=locals().get("queue_result", {}),
                delivery_result=locals().get("delivery_result", {}),
                summary={},
                error=str(exc),
            )
        return {"ok": False, "test_run_id": run_id, "error": str(exc)}
    finally:
        _IMPORT_LOCK.release()


@app.get("/api/test-run/report")
def api_test_run_report(since_case_id: int = Query(default=0, ge=0), limit: int = Query(default=500, ge=50, le=2000)) -> dict[str, Any]:
    apply_runtime_settings()
    with connect() as con:
        return _case_quality_report(con, since_case_id=since_case_id, limit=limit)


@app.get("/api/test-run/list")
def api_test_run_list(limit: int = Query(default=20, ge=1, le=100)) -> dict[str, Any]:
    with connect() as con:
        rows = [row_to_dict(r) for r in con.execute(
            """
            SELECT id, started_at, finished_at, status, params_json, before_json, after_json,
                   import_result_json, queue_result_json, delivery_result_json, summary_json, error
            FROM test_runs
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()]
    return {"count": len(rows), "items": rows}


@app.get("/api/test-run/latest")
def api_test_run_latest() -> dict[str, Any]:
    with connect() as con:
        row = con.execute(
            """
            SELECT id, started_at, finished_at, status, params_json, before_json, after_json,
                   import_result_json, queue_result_json, delivery_result_json, summary_json, error
            FROM test_runs
            ORDER BY id DESC
            LIMIT 1
            """
        ).fetchone()
    if not row:
        return {"ok": True, "empty": True, "message": "test_run_not_started"}
    return {"ok": True, "item": row_to_dict(row)}



@app.get("/api/settings")
def api_settings() -> dict[str, Any]:
    return get_settings_payload(mask_secrets=True)


@app.post("/api/settings")
def api_update_settings(update: SettingsUpdate) -> dict[str, Any]:
    result = update_settings_from_panel(update.values)
    return {**result, "settings": get_settings_payload(mask_secrets=True)}


@app.post("/api/telegram/test")
def api_telegram_test() -> dict[str, Any]:
    """Send a test message to verify Telegram bot configuration."""
    apply_runtime_settings()
    from .telegram import test_connection
    return test_connection()


@app.post("/api/telegram/daily-report")
def api_telegram_daily_report() -> dict[str, Any]:
    """Отправить суточный отчёт сейчас (превью/тест)."""
    apply_runtime_settings()
    from .telegram import send_daily_report
    with connect() as con:
        send_daily_report(con)
    return {"ok": True, "sent": True}


@app.post("/api/telegram/hourly-report")
def api_telegram_hourly_report() -> dict[str, Any]:
    """Отправить часовой отчёт сейчас (превью/тест)."""
    apply_runtime_settings()
    from .telegram import send_hourly_report
    with connect() as con:
        send_hourly_report(con)
    return {"ok": True, "sent": True}


@app.get("/api/telegram/status")
def api_telegram_status() -> dict[str, Any]:
    apply_runtime_settings()
    return {
        "configured": bool(settings.tg_bot_token and settings.tg_chat_ids),
        "token_set": bool(settings.tg_bot_token),
        "chat_ids": [c.strip() for c in (settings.tg_chat_ids or "").split(",") if c.strip()],
        "whitelist_enabled": settings.tg_whitelist_enabled,
        "notify_on_cycle": settings.tg_notify_on_cycle,
        "notify_unresolved": settings.tg_notify_unresolved,
        "notify_errors": settings.tg_notify_errors,
        "notify_ready": settings.tg_notify_ready,
        "unresolved_min": settings.tg_unresolved_min,
        "report_interval_minutes": settings.tg_report_interval_minutes,
        "report_include_reasons": settings.tg_report_include_reasons,
    }


@app.get("/api/review/cases")
def api_review_cases(
    source: str = "all",   # all | pattern | ai
    buyer: str = "",
    folder: str = "all",
    q: str = "",           # поиск по теме/номеру документа
    missing: str = "",     # фильтр «пустые ячейки»: csv из document_number,document_date,part_number,quantity,claim_kind
    kind: str = "",        # фильтр по причине (claim_kind): defect|quality_refusal|shortage|...
    page: int = 1,
    limit: int = 50,
) -> dict[str, Any]:
    """Полный реестр кейсов для сверки с безопасной разбивкой по папкам."""
    # SQL-условие «ячейка пуста» для каждого важного поля (для фильтра по пробелам).
    # Поля кроме claim_kind лежат в fields_json → json_extract; claim_kind — отдельная колонка.
    EMPTY_FIELD_SQL = {
        "document_number": "(json_extract(c.fields_json,'$.document_number') IS NULL OR json_extract(c.fields_json,'$.document_number')='')",
        "document_date":   "(json_extract(c.fields_json,'$.document_date') IS NULL OR json_extract(c.fields_json,'$.document_date')='')",
        "part_number":     "(json_extract(c.fields_json,'$.part_number') IS NULL OR json_extract(c.fields_json,'$.part_number')='')",
        "quantity":        "(json_extract(c.fields_json,'$.quantity') IS NULL OR json_extract(c.fields_json,'$.quantity')='')",
        "claim_kind":      "(c.claim_kind IS NULL OR c.claim_kind='')",
    }
    # Папки явно по НАЗНАЧЕНИЮ (порядок = порядок чипов в UI).
    FOLDER_SQL = {
        "ready_1c": "c.event_type IN ('new_return','pre_delivery_refusal') AND c.state='ready_to_1c'",
        "manual": "c.state='needs_review'",
        "needs_link": "c.state='needs_link'",
        "corrections": "c.event_type='correction_request' AND c.state!='needs_link'",
        "ready_to_ship": "c.event_type='ready_to_ship' AND c.state!='needs_link'",
        "followups": "c.event_type='followup_dialog' AND c.state!='needs_link'",
        "reminders": "c.event_type='followup_reminder' AND c.state!='needs_link'",
        "supplier_decisions": "c.event_type='supplier_decision' AND c.state!='needs_link'",
        "marking": "c.event_type='marking_request' AND c.state!='needs_link'",
        "problem_notice": "c.event_type='problem_notice' OR c.state='problem_notice'",
        "information": "c.state='ignored_info_only' OR c.event_type='info_only'",
        "unknown": "c.event_type='unknown' OR c.state IS NULL OR c.state='unknown'",
    }
    FOLDER_NAMES = {
        "ready_1c": "✅ Готово в 1С",
        "manual": "✋ Ручной разбор",
        "needs_link": "🔗 Связки: ждут привязки",
        "corrections": "📝 Корректировки / ЭДО",
        "ready_to_ship": "📦 Готово к выдаче",
        "followups": "💬 Диалоги (продолжения)",
        "reminders": "⏰ Напоминания",
        "supplier_decisions": "✔️ Решения поставщика",
        "marking": "🏷 Маркировка / ТНВЭД",
        "problem_notice": "⚠️ Уведомления о проблеме",
        "information": "🗑 Прайсы / отчёты / мусор",
        "unknown": "❓ Неизвестные",
    }

    def review_folder(case: dict[str, Any]) -> tuple[str, str]:
        event_type = str(case.get("event_type") or "")
        state = str(case.get("state") or "")
        if state == "needs_review":
            key = "manual"  # ручной разбор — всё, что требует глаз оператора
        elif event_type in {"new_return", "pre_delivery_refusal"} and state == "ready_to_1c":
            key = "ready_1c"
        elif state == "needs_link":
            key = "needs_link"
        elif event_type == "correction_request":
            key = "corrections"
        elif event_type == "ready_to_ship":
            key = "ready_to_ship"
        elif event_type == "followup_dialog":
            key = "followups"
        elif event_type == "followup_reminder":
            key = "reminders"
        elif event_type == "supplier_decision":
            key = "supplier_decisions"
        elif event_type == "marking_request":
            key = "marking"
        elif event_type == "problem_notice" or state == "problem_notice":
            key = "problem_notice"
        elif state == "ignored_info_only" or event_type == "info_only":
            key = "information"
        else:
            key = "unknown"
        return key, FOLDER_NAMES[key]

    with connect() as con:
        conditions: list[str] = []
        params: list[Any] = []
        if folder in FOLDER_SQL:
            conditions.append(f"({FOLDER_SQL[folder]})")
        # Фильтр пустых ячеек: показываем кейсы, где пусто ХОТЯ БЫ одно из выбранных полей (OR).
        miss_keys = [k.strip() for k in missing.split(",") if k.strip() in EMPTY_FIELD_SQL]
        if miss_keys:
            conditions.append("(" + " OR ".join(EMPTY_FIELD_SQL[k] for k in miss_keys) + ")")
        if source == "pattern":
            conditions.append("(c.payload_json NOT LIKE '%\"processing_source\":\"ai\"%' AND c.payload_json NOT LIKE '%\"ai_overlay\"%')")
        elif source == "ai":
            conditions.append("(c.payload_json LIKE '%\"processing_source\":\"ai\"%' OR c.payload_json LIKE '%\"ai_overlay\"%')")
        if buyer:
            conditions.append("c.buyer_code = ?")
            params.append(buyer)
        if kind:
            if kind == "__none__":
                conditions.append("(c.claim_kind IS NULL OR c.claim_kind='')")
            else:
                conditions.append("c.claim_kind = ?")
                params.append(kind)
        if q:
            conditions.append("(e.subject LIKE ? OR c.fields_json LIKE ?)")
            params.extend([f"%{q}%", f"%{q}%"])
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        total = con.execute(
            f"SELECT COUNT(*) c FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id {where}",
            params,
        ).fetchone()["c"]
        offset = (page - 1) * limit
        rows = con.execute(
            f"""SELECT c.id, c.buyer_code, c.buyer_name, c.event_type, c.claim_kind,
                c.state, c.priority, c.confidence, c.ready_for_export,
                c.fields_json, c.missing_json, c.quality_json, c.payload_json,
                e.subject, e.from_addr, e.received_at, e.snippet, c.raw_email_id
                FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
                {where}
                ORDER BY c.ready_for_export DESC,
                  CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
                  e.received_at DESC
                LIMIT ? OFFSET ?""",
            params + [limit, offset],
        ).fetchall()
        cases = []
        for r in rows:
            d = row_to_dict(r) or {}
            payload = d.get("payload") or {}
            source_method = payload.get("processing_source") or ("ai" if payload.get("ai_overlay") else "pattern")
            folder_key, folder_name = review_folder(d)
            can_export = (
                d.get("event_type") in {"new_return", "pre_delivery_refusal"}
                and d.get("state") in {"ready_to_1c", "needs_review"}
            )
            cases.append({
                "id": d.get("id"),
                "raw_email_id": d.get("raw_email_id"),
                "buyer_code": d.get("buyer_code"),
                "buyer_name": d.get("buyer_name"),
                "event_type": d.get("event_type"),
                "claim_kind": d.get("claim_kind"),
                "state": d.get("state"),
                "priority": d.get("priority"),
                "confidence": d.get("confidence"),
                "ready_for_export": bool(d.get("ready_for_export")),
                "source": source_method,
                "processing_mode": payload.get("processing_mode") or "auto",
                "manual_gate": bool(payload.get("manual_gate")),
                "defect_doc_flag": payload.get("defect_doc_flag"),
                "evidence_gate": payload.get("evidence_gate") or {},
                "fields": d.get("fields") or {},
                "multi_item_count": int(payload.get("multi_item_count") or 0),
                "missing": d.get("missing") or [],
                "quality": d.get("quality") or [],
                "subject": d.get("subject"),
                "from_addr": d.get("from_addr"),
                "received_at": d.get("received_at"),
                "snippet": d.get("snippet"),
                "folder_key": folder_key,
                "folder_name": folder_name,
                "can_export": can_export,
            })
        buyers = [dict(r) for r in con.execute(
            "SELECT DISTINCT buyer_code code, buyer_name name FROM cases WHERE buyer_code IS NOT NULL ORDER BY buyer_name"
        ).fetchall()]
        # Счётчики текущей папки, клиента и источника, но без фильтра missing.
        base_cond: list[str] = []
        base_params: list[Any] = []
        if folder in FOLDER_SQL:
            base_cond.append(f"({FOLDER_SQL[folder]})")
        if buyer:
            base_cond.append("c.buyer_code = ?"); base_params.append(buyer)
        if source == "pattern":
            base_cond.append("(c.payload_json NOT LIKE '%\"processing_source\":\"ai\"%' AND c.payload_json NOT LIKE '%\"ai_overlay\"%')")
        elif source == "ai":
            base_cond.append("(c.payload_json LIKE '%\"processing_source\":\"ai\"%' OR c.payload_json LIKE '%\"ai_overlay\"%')")
        base_where = ("WHERE " + " AND ".join(base_cond)) if base_cond else "WHERE 1=1"
        empty_counts = {}
        for k, sql in EMPTY_FIELD_SQL.items():
            empty_counts[k] = con.execute(
                f"SELECT COUNT(*) c FROM cases c {base_where} AND {sql}", base_params
            ).fetchone()["c"]
        # Причины (claim_kind) со счётчиками — для сортировки/фильтра Сверки
        kind_counts = [
            {"kind": (r["k"] or "__none__"), "count": r["c"]}
            for r in con.execute(
                f"SELECT c.claim_kind k, COUNT(*) c FROM cases c {base_where} GROUP BY c.claim_kind ORDER BY c DESC",
                base_params,
            ).fetchall()
        ]
        folder_counts = {"all": con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"]}
        for key, sql in FOLDER_SQL.items():
            folder_counts[key] = con.execute(
                f"SELECT COUNT(*) c FROM cases c WHERE {sql}"
            ).fetchone()["c"]
    shown_count = len(cases)
    return {
        "ok": True,
        "total": total,
        "total_count": total,
        "shown_count": shown_count,
        "page": page,
        "page_size": limit,
        "has_more": offset + shown_count < total,
        "cases": cases,
        "buyers": buyers,
        "empty_counts": empty_counts,
        "kind_counts": kind_counts,
        "folder_counts": folder_counts,
        "folder_names": FOLDER_NAMES,
    }


@app.post("/api/review/approve/{case_id}")
def api_review_approve(case_id: int) -> dict[str, Any]:
    """Подтвердить кейс из Сверки → queue в outbox."""
    case = mark_ready(case_id)
    if case.get("state") != "ready_to_1c" or not case.get("ready_for_export"):
        return {
            "ok": False,
            "error": "evidence_gate_failed",
            "message": "Evidence не подтверждён, кейс оставлен на сверке.",
            "case_id": case_id,
            "evidence_gate": case.get("evidence_gate") or {},
            "case": case,
        }
    return {"ok": True, "case_id": case_id, "case": case}


@app.post("/api/review/approve-all")
def api_review_approve_all() -> dict[str, Any]:
    """Подтвердить все ready_to_1c кейсы из Сверки."""
    with connect() as con:
        rows = con.execute("SELECT id FROM cases WHERE state='ready_to_1c' AND ready_for_export=1").fetchall()
        queued = 0
        blocked = 0
        for r in rows:
            res = queue_case_event(con, int(r["id"]))
            queued += int(res.get("queued") or 0)
            if res.get("reason") in {"evidence_gate_failed", "weak_followup_link"}:
                blocked += 1
    return {"ok": True, "checked": len(rows), "approved": queued, "queued": queued, "blocked": blocked}


@app.get("/api/automation/status")
def api_automation_status() -> dict[str, Any]:
    apply_runtime_settings()
    return {
        "scan_interval_seconds": settings.scan_interval_seconds,
        "auto_import_enabled": bool(settings.scan_interval_seconds and settings.imap_username and settings.imap_password),
        "last_import": _LAST_IMPORT_RESULT,
    }




@app.get("/api/autopilot/status")
def api_autopilot_status() -> dict[str, Any]:
    apply_runtime_settings()
    return {"ok": True, "state": dict(_AUTOPILOT_STATE), "settings": {"enable_ai": settings.enable_ai, "auto_deliver_outbox": settings.auto_deliver_outbox}}


@app.get("/api/live/events")
def api_live_events(
    limit: int = Query(default=200, ge=1, le=1000),
    since_id: int = Query(default=0, ge=0),
    stage: str | None = Query(default=None),
) -> dict[str, Any]:
    with connect() as con:
        items = list_process_events(con, limit=limit, since_id=since_id, stage=stage)
        dash = process_event_dashboard(con, limit=min(limit, 200))
    return {"ok": True, "count": len(items), "latest_id": dash.get("latest_id", 0), "counts": dash.get("counts", []), "items": items}


@app.post("/api/live/events/clear")
def api_live_events_clear(keep_last: int = Query(default=0, ge=0, le=10000)) -> dict[str, Any]:
    with connect() as con:
        return clear_process_events(con, keep_last=keep_last)


@app.post("/api/live/events/note")
def api_live_events_note(message: str = Query(default="Операторская отметка"), stage: str = Query(default="operator")) -> dict[str, Any]:
    with connect() as con:
        eid = record_process_event(con, stage=stage, message=message, level="info", details={"source": "operator"})
    return {"ok": True, "id": eid}




@app.get("/api/ai/status")
def api_ai_status(limit: int = Query(default=20, ge=1, le=200)) -> dict[str, Any]:
    """Operator-facing AI health: is it actually doing anything and what happened recently."""
    apply_runtime_settings()
    with connect() as con:
        usage_counts = [dict(r) for r in con.execute(
            """
            SELECT provider, model, ok, cached, COUNT(*) count,
                   COALESCE(SUM(prompt_chars),0) prompt_chars,
                   COALESCE(SUM(response_chars),0) response_chars
            FROM ai_usage
            GROUP BY provider, model, ok, cached
            ORDER BY MAX(id) DESC
            """
        ).fetchall()]
        recent_usage = [row_to_dict(r) for r in con.execute(
            """
            SELECT u.id, u.case_id, u.provider, u.model, u.ok, u.cached, u.prompt_chars, u.response_chars,
                   u.error, u.created_at, e.subject
            FROM ai_usage u
            LEFT JOIN cases c ON c.id=u.case_id
            LEFT JOIN raw_emails e ON e.id=c.raw_email_id
            ORDER BY u.id DESC
            LIMIT ?
            """, (int(limit),)
        ).fetchall()]
        suggestions = [row_to_dict(r) for r in con.execute(
            """
            SELECT s.id, s.case_id, s.model, s.accepted, s.created_at, e.subject,
                   substr(s.response_json,1,1200) response_excerpt
            FROM ai_suggestions s
            LEFT JOIN cases c ON c.id=s.case_id
            LEFT JOIN raw_emails e ON e.id=c.raw_email_id
            ORDER BY s.id DESC
            LIMIT ?
            """, (int(limit),)
        ).fetchall()]
        events = list_process_events(con, limit=int(limit), stage="ai")
    total = sum(int(x.get("count") or 0) for x in usage_counts)
    ok = sum(int(x.get("count") or 0) for x in usage_counts if int(x.get("ok") or 0) == 1)
    failed = total - ok
    return {
        "ok": True,
        "schema": "readmail-ai-status-v1.16",
        "enabled": bool(settings.enable_ai),
        "provider": settings.ai_provider,
        "model": settings.ai_model,
        "base_url": settings.ai_base_url,
        "summary": {"total_calls": total, "ok_calls": ok, "failed_calls": failed, "suggestions": len(suggestions)},
        "usage_counts": usage_counts,
        "recent_usage": recent_usage,
        "recent_suggestions": suggestions,
        "recent_events": events,
    }


@app.get("/api/ai/decision-report")
def api_ai_decision_report(limit: int = Query(default=20, ge=1, le=200)) -> dict[str, Any]:
    apply_runtime_settings()
    return _ai_decision_snapshot(limit=limit)


@app.post("/api/autopilot/start")
def api_autopilot_start(
    interval_seconds: int = Query(default=300, ge=30, le=86400),
    import_limit: int = Query(default=50, ge=0, le=1000),
    ai_limit: int = Query(default=10, ge=0, le=100),
    deliver: bool = Query(default=False),
    mode: str = Query(default="pattern"),
) -> dict[str, Any]:
    global _AUTOPILOT_THREAD
    mode = "full_ai" if mode == "full_ai" else "pattern"
    # full_ai по умолчанию авто-доставляет в 1С (Сверка вручную → 1С авто).
    if mode == "full_ai":
        deliver = True
    with _AUTOPILOT_LOCK:
        if _AUTOPILOT_THREAD and _AUTOPILOT_THREAD.is_alive():
            return {"ok": True, "already_running": True, "state": dict(_AUTOPILOT_STATE)}
        _AUTOPILOT_STOP.clear()
        _AUTOPILOT_STATE.update({
            "enabled": True,
            "mode": mode,
            "started_at": utcnow(),
            "stopped_at": None,
            "last_error": None,
            "next_cycle_at": None,
            "config": {"interval_seconds": interval_seconds, "import_limit": import_limit, "ai_limit": ai_limit, "deliver": deliver, "mode": mode},
        })
        _AUTOPILOT_THREAD = threading.Thread(target=_autopilot_loop, name="readmail-panel-autopilot", daemon=True)
        _AUTOPILOT_THREAD.start()
    _log("autopilot", f"Автопилот ({mode}) запущен оператором", level="ok", details={"mode": mode, "interval_seconds": interval_seconds, "import_limit": import_limit, "ai_limit": ai_limit, "deliver": deliver})
    return {"ok": True, "started": True, "mode": mode, "state": dict(_AUTOPILOT_STATE)}


@app.post("/api/autopilot/stop")
def api_autopilot_stop() -> dict[str, Any]:
    _AUTOPILOT_STOP.set()
    _AUTOPILOT_STATE["enabled"] = False
    _AUTOPILOT_STATE["stopped_at"] = utcnow()
    _AUTOPILOT_STATE["next_cycle_at"] = None
    _log("autopilot", "Автопилот поставлен на паузу", level="warn")
    return {"ok": True, "stopping": True, "state": dict(_AUTOPILOT_STATE)}


@app.post("/api/autopilot/cycle")
def api_autopilot_cycle(
    import_limit: int = Query(default=50, ge=0, le=1000),
    ai_limit: int = Query(default=10, ge=0, le=100),
    deliver: bool = Query(default=False),
) -> dict[str, Any]:
    if not _AUTOPILOT_LOCK.acquire(blocking=False):
        return {"ok": False, "skipped": True, "reason": "autopilot_cycle_already_running"}
    try:
        result = _autopilot_cycle(import_limit=import_limit, ai_limit=ai_limit, deliver=deliver)
        _AUTOPILOT_STATE["last_cycle"] = result
        _AUTOPILOT_STATE["cycle_count"] = int(_AUTOPILOT_STATE.get("cycle_count") or 0) + 1
        _AUTOPILOT_STATE["last_error"] = result.get("error") if not result.get("ok") else None
        return result
    finally:
        _AUTOPILOT_LOCK.release()


def _processing_report(limit: int = 300, include_body_excerpt: bool = False) -> dict[str, Any]:
    with connect() as con:
        state_counts = [dict(r) for r in con.execute("SELECT state, COUNT(*) count FROM cases GROUP BY state ORDER BY count DESC").fetchall()]
        event_counts = [dict(r) for r in con.execute("SELECT event_type, COUNT(*) count FROM cases GROUP BY event_type ORDER BY count DESC").fetchall()]
        kind_counts = [dict(r) for r in con.execute("SELECT COALESCE(claim_kind,'unknown') claim_kind, COUNT(*) count FROM cases GROUP BY COALESCE(claim_kind,'unknown') ORDER BY count DESC").fetchall()]
        folder_counts = [dict(r) for r in con.execute("SELECT mailbox, COUNT(*) count FROM raw_emails GROUP BY mailbox ORDER BY count DESC LIMIT 50").fetchall()]
        day_counts = [dict(r) for r in con.execute("SELECT substr(COALESCE(received_at, imported_at),1,10) day, COUNT(*) count FROM raw_emails GROUP BY substr(COALESCE(received_at, imported_at),1,10) ORDER BY day DESC LIMIT 30").fetchall()]
        attachment_summary = con.execute("SELECT COUNT(*) count, COALESCE(SUM(size_bytes),0) size_bytes FROM attachments").fetchone()
        attachment_types = [dict(r) for r in con.execute("SELECT COALESCE(content_type,'unknown') content_type, COUNT(*) count, COALESCE(SUM(size_bytes),0) size_bytes FROM attachments GROUP BY COALESCE(content_type,'unknown') ORDER BY count DESC LIMIT 30").fetchall()]
        attachment_samples = [dict(r) for r in con.execute("SELECT a.filename, a.content_type, a.size_bytes, r.subject, r.from_addr, r.received_at, r.mailbox FROM attachments a JOIN raw_emails r ON r.id=a.raw_email_id ORDER BY a.id DESC LIMIT 200").fetchall()]
        ai_counts = [dict(r) for r in con.execute("SELECT provider, model, ok, cached, COUNT(*) count, COALESCE(SUM(prompt_chars),0) prompt_chars, COALESCE(SUM(response_chars),0) response_chars FROM ai_usage GROUP BY provider, model, ok, cached ORDER BY count DESC").fetchall()]
        ai_suggestions = [dict(r) for r in con.execute("SELECT accepted, COUNT(*) count FROM ai_suggestions GROUP BY accepted ORDER BY accepted DESC").fetchall()]
        outbox_counts = [dict(r) for r in con.execute("SELECT status, event_type, channel, COUNT(*) count FROM outbox GROUP BY status, event_type, channel ORDER BY count DESC").fetchall()]
        rows = con.execute(
            """
            SELECT c.*, e.mailbox, e.uid, e.message_id, e.subject, e.from_addr, e.received_at, e.snippet, e.folder_seen_json,
                   e.direction, e.quote_markers
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            ORDER BY
              CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
              c.ready_for_export DESC,
              c.id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
        items: list[dict[str, Any]] = []
        for row in rows:
            d = row_to_dict(row) or {}
            payload = d.get("payload") or {}
            fields = d.get("fields") or {}
            export = d.get("export") or {}
            attachments = [dict(a) for a in con.execute("SELECT filename, content_type, size_bytes FROM attachments WHERE raw_email_id=? ORDER BY id", (d.get("raw_email_id"),)).fetchall()]
            outbox = [row_to_dict(o) for o in con.execute("SELECT id, event_type, status, channel, attempt_count, last_error, file_path, created_at, sent_at FROM outbox WHERE case_id=? ORDER BY id", (d.get("id"),)).fetchall()]
            ai_rows = [row_to_dict(a) for a in con.execute("SELECT id, model, accepted, created_at, response_json FROM ai_suggestions WHERE case_id=? ORDER BY id DESC LIMIT 5", (d.get("id"),)).fetchall()]
            ai_short = []
            for a in ai_rows:
                resp = a.get("response") or {}
                ai_short.append({
                    "id": a.get("id"), "model": a.get("model"), "accepted": bool(a.get("accepted")), "created_at": a.get("created_at"),
                    "ok": resp.get("ok"), "cached": resp.get("cached"),
                    "suggested_event_type": (resp.get("response") or {}).get("event_type") if isinstance(resp.get("response"), dict) else None,
                    "suggested_claim_kind": (resp.get("response") or {}).get("claim_kind") if isinstance(resp.get("response"), dict) else None,
                })
            automation = {
                "classifier": payload.get("classifier"),
                "direction": payload.get("direction"),
                "reasons": payload.get("reasons") or {},
                "ai_overlay": payload.get("ai_overlay") or {},
                "control": payload.get("control") or {},
                "evidence": payload.get("evidence") or {},
            }
            items.append({
                "case_id": d.get("id"),
                "raw_email_id": d.get("raw_email_id"),
                "source": {"mailbox": d.get("mailbox"), "folders_seen": d.get("folder_seen") or [], "uid": d.get("uid"), "message_id": d.get("message_id"), "received_at": d.get("received_at"), "from": d.get("from_addr"), "subject": d.get("subject"), "quote_markers": d.get("quote_markers")},
                "result": {"state": d.get("state"), "event_type": d.get("event_type"), "claim_kind": d.get("claim_kind"), "priority": d.get("priority"), "deadline_at": d.get("deadline_at"), "ready_for_export": bool(d.get("ready_for_export")), "confidence": d.get("confidence")},
                "fields": fields,
                "missing": d.get("missing") or [],
                "quality": d.get("quality") or [],
                "attachments": attachments,
                "automation": automation,
                "ai": ai_short,
                "outbox": outbox,
                "export_json_ready": bool(export and d.get("ready_for_export")),
                "export_preview": export if d.get("ready_for_export") else {},
                "snippet": d.get("snippet") if include_body_excerpt else None,
            })
    return {
        "schema": "readmail-processing-report-v1.14",
        "generated_at": utcnow(),
        "summary": {
            "raw_emails": sum(x.get("count", 0) for x in folder_counts),
            "cases_in_sample": len(items),
            "attachments": {"count": int(attachment_summary["count"] if attachment_summary else 0), "size_bytes": int(attachment_summary["size_bytes"] if attachment_summary else 0)},
            "state_counts": state_counts,
            "event_counts": event_counts,
            "claim_kind_counts": kind_counts,
            "folder_counts": folder_counts,
            "received_by_day": day_counts,
            "attachment_types": attachment_types,
            "ai_usage": ai_counts,
            "ai_suggestions": ai_suggestions,
            "outbox_counts": outbox_counts,
        },
        "attachment_samples": attachment_samples,
        "cases": items,
        "how_to_read": [
            "automation.classifier/reasons = что сделали правила и паттерны",
            "ai = что предлагала/применяла модель",
            "fields/missing/quality = что извлечено и что заблокировано validator",
            "outbox = что уже подготовлено/ушло/не ушло для 1С",
            "export_preview появляется только для ready_for_export=true",
        ],
    }


@app.get("/api/reports/processing")
def api_processing_report(limit: int = Query(default=300, ge=1, le=2000), include_body_excerpt: bool = Query(default=False)) -> dict[str, Any]:
    return _processing_report(limit=limit, include_body_excerpt=include_body_excerpt)


@app.get("/api/reports/processing/download")
def api_processing_report_download(limit: int = Query(default=1000, ge=1, le=5000)) -> JSONResponse:
    report = _processing_report(limit=limit, include_body_excerpt=False)
    filename = f"readmail_processing_report_{utcnow().replace(':','').replace('+','Z')}.json"
    return JSONResponse(report, headers={"Content-Disposition": f"attachment; filename={filename}"})


@app.get("/api/reports/quality")
def api_quality_report() -> dict[str, Any]:
    path = settings.database_path.parent / "quality_report.json"
    if not path.exists():
        return {
            "generated_at": utcnow(),
            "total": 0,
            "accepted": 0,
            "needs_ai_repair": 0,
            "needs_human_review": 0,
            "critical_error": 0,
            "statuses": {},
            "field_errors": {},
        }
    return loads(path.read_text(encoding="utf-8"), {}) or {}


@app.get("/api/reports/quality/errors")
def api_quality_errors(limit: int = Query(default=200, ge=1, le=2000)) -> dict[str, Any]:
    path = settings.database_path.parent / "quality_errors.jsonl"
    if not path.exists():
        return {"ok": True, "count": 0, "items": []}
    items: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines()[-int(limit):]:
        item = loads(line, None)
        if isinstance(item, dict):
            items.append(item)
    return {"ok": True, "count": len(items), "items": items}



@app.post("/api/system/reset-processing-data")
def api_reset_processing_data(
    keep_learning: bool = Query(default=True),
    keep_settings: bool = Query(default=True),
    keep_process_events: bool = Query(default=False),
    confirm: str = Query(default=""),
) -> dict[str, Any]:
    if confirm != "RESET":
        return {"ok": False, "error": "confirmation_required", "hint": "Call with confirm=RESET. Settings are kept by default."}
    with connect() as con:
        result = reset_processing_data(con, keep_settings=keep_settings, keep_learning=keep_learning, keep_process_events=keep_process_events)
        record_process_event(con, stage="operator", level="warn", message="Тестовые данные обнулены", details=result)
    return result


@app.post("/api/system/purge-junk-attachments")
def api_purge_junk_attachments(confirm: str = Query(default=""), dry_run: bool = Query(default=True)) -> dict[str, Any]:
    """«Не хранить помойку»: удалить вложения info_only/служебных писем (прайс-листы ПИТСТОП и пр.)
    с диска и из БД. Само письмо остаётся для аудита. dry_run=True — только показать объём."""
    with connect() as con:
        # Чистим вложения ТОЛЬКО у писем, где ВСЕ кейсы служебные (info_only/ignored_info_only).
        # Раньше удаляло по любому info_only-кейсу на письме → сносило вложения претензии-соседа
        # на том же письме (потеря документов брака). Теперь защищаем письма с деловым кейсом.
        rows = con.execute(
            """SELECT a.id, a.file_path, COALESCE(a.size_bytes,0) sz FROM attachments a
               WHERE a.raw_email_id IN (
                 SELECT c.raw_email_id FROM cases c
                 WHERE c.event_type='info_only' OR c.state='ignored_info_only'
               )
               AND a.raw_email_id NOT IN (
                 SELECT c2.raw_email_id FROM cases c2
                 WHERE NOT (c2.event_type='info_only' OR c2.state='ignored_info_only')
               )"""
        ).fetchall()
    n = len(rows)
    freed_mb = round(sum(r["sz"] for r in rows) / 1024 / 1024, 1)
    if dry_run or confirm != "PURGE":
        return {"ok": True, "dry_run": True, "count": n, "freed_mb": freed_mb,
                "hint": "confirm=PURGE&dry_run=false для удаления"}
    files_removed = 0
    for r in rows:
        fp = r["file_path"]
        if fp:
            try:
                p = Path(fp)
                if p.exists():
                    p.unlink()
                    files_removed += 1
            except Exception:
                pass
    with connect() as con:
        ids = [int(r["id"]) for r in rows]
        for i in range(0, len(ids), 500):
            chunk = ids[i:i + 500]
            con.execute("DELETE FROM attachments WHERE id IN (%s)" % ",".join("?" * len(chunk)), chunk)
        con.commit()
        record_process_event(con, stage="operator", level="info",
                             message=f"Помойка очищена: {n} вложений, {freed_mb} МБ", details={"count": n, "freed_mb": freed_mb})
    return {"ok": True, "deleted": n, "files_removed": files_removed, "freed_mb": freed_mb}


def _period_date10(value: object) -> str | None:
    s = str(value or "").strip()
    return s[:10] if len(s) >= 10 else None  # 'YYYY-MM-DDTHH:MM' → 'YYYY-MM-DD'


@app.post("/api/import/purge-outside-period")
def api_purge_outside_period(confirm: str = Query(default=""), dry_run: bool = Query(default=True)) -> dict[str, Any]:
    """Удалить письма ВНЕ периода загрузки (по received_at), каскадом cases/outbox/вложения.
    Период — из настроек IMAP_DATE_FROM/TO (галочки). dry_run=True только считает."""
    apply_runtime_settings()
    from_d = _period_date10(getattr(settings, "imap_date_from", "")) if getattr(settings, "imap_date_from_enabled", False) else None
    to_d = _period_date10(getattr(settings, "imap_date_to", "")) if getattr(settings, "imap_date_to_enabled", False) else None
    if not (from_d or to_d):
        return {"ok": False, "error": "Период не задан — включите дату ОТ и/или ДО в настройках почты"}
    where: list[str] = []
    params: list[str] = []
    if from_d:
        where.append("substr(COALESCE(received_at,''),1,10) < ?"); params.append(from_d)
    if to_d:
        where.append("substr(COALESCE(received_at,''),1,10) > ?"); params.append(to_d)
    wsql = " OR ".join(where)
    with connect() as con:
        rids = [int(r["id"]) for r in con.execute(f"SELECT id FROM raw_emails WHERE {wsql}", params).fetchall()]
        total = con.execute("SELECT COUNT(*) FROM raw_emails").fetchone()[0]
        if dry_run or confirm != "PURGE":
            return {"ok": True, "dry_run": True, "outside_period": len(rids), "total": total,
                    "period": {"from": from_d, "to": to_d}, "hint": "confirm=PURGE&dry_run=false для удаления"}
        files_removed = 0
        for i in range(0, len(rids), 500):
            chunk = rids[i:i + 500]
            ph = ",".join("?" * len(chunk))
            case_ids = [int(r["id"]) for r in con.execute(f"SELECT id FROM cases WHERE raw_email_id IN ({ph})", chunk).fetchall()]
            for r in con.execute(f"SELECT file_path FROM attachments WHERE raw_email_id IN ({ph})", chunk).fetchall():
                fp = r["file_path"]
                if fp:
                    try:
                        p = Path(fp)
                        if p.exists():
                            p.unlink(); files_removed += 1
                    except Exception:
                        pass
            if case_ids:
                cph = ",".join("?" * len(case_ids))
                con.execute(f"DELETE FROM outbox WHERE case_id IN ({cph})", case_ids)
            con.execute(f"DELETE FROM attachments WHERE raw_email_id IN ({ph})", chunk)
            con.execute(f"DELETE FROM cases WHERE raw_email_id IN ({ph})", chunk)
            con.execute(f"DELETE FROM raw_emails WHERE id IN ({ph})", chunk)
        remaining = con.execute("SELECT COUNT(*) FROM raw_emails").fetchone()[0]
        record_process_event(con, stage="operator", level="info",
                             message=f"Удалены письма вне периода: {len(rids)}",
                             details={"deleted": len(rids), "period": {"from": from_d, "to": to_d}})
        con.commit()
    return {"ok": True, "deleted": len(rids), "files_removed": files_removed, "remaining": remaining,
            "period": {"from": from_d, "to": to_d}}


@app.post("/api/system/reset-and-import")
def api_reset_and_import(confirm: str = Query(...), limit: int = 50) -> dict[str, Any]:
    if confirm != "RESET_IMPORT":
        raise HTTPException(status_code=400, detail="confirm must be RESET_IMPORT")
    with connect() as con:
        reset = reset_processing_data(con, keep_settings=True, keep_learning=True, keep_process_events=False)
    _log("operator", "Тестовые данные обнулены перед новым импортом", level="warn", details=reset)
    imp = _safe_import_cycle()  # v2.1: полный цикл фетч+классиф (учитывает фильтр дат)
    _log("import", "Импорт после reset завершён", level="ok" if imp.get("ok", True) else "warn", details=imp)
    return {"ok": True, "reset": reset, "import": imp}


@app.get("/api/stats")
def stats() -> dict[str, Any]:
    with connect() as con:
        total_emails = con.execute("SELECT COUNT(*) c FROM raw_emails").fetchone()["c"]
        total_cases = con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"]
        states = [dict(r) for r in con.execute("SELECT state, COUNT(*) count FROM cases GROUP BY state ORDER BY count DESC")]
        priorities = [dict(r) for r in con.execute("SELECT priority, COUNT(*) count FROM cases GROUP BY priority ORDER BY count DESC")]
        overdue = con.execute(
            "SELECT COUNT(*) c FROM cases WHERE deadline_at IS NOT NULL AND deadline_at <= datetime('now') AND state NOT IN ('exported','closed')"
        ).fetchone()["c"]
        ready = con.execute("SELECT COUNT(*) c FROM cases WHERE ready_for_export=1").fetchone()["c"]
        blocked_evidence = con.execute("SELECT COUNT(*) c FROM cases WHERE state='needs_review' AND (missing_json LIKE '%photo_evidence%' OR missing_json LIKE '%service_document%')").fetchone()["c"]
        outbox_new = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='new'").fetchone()["c"]
        outbox_errors = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='error'").fetchone()["c"]
        outbox_sent = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='sent'").fetchone()["c"]
        outbox_retry_due = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='error' AND (next_attempt_at IS NULL OR next_attempt_at<=datetime('now'))").fetchone()["c"]
    return {"emails": total_emails, "cases": total_cases, "states": states, "priorities": priorities, "overdue": overdue, "ready": ready, "blocked_evidence": blocked_evidence, "outbox_new": outbox_new, "outbox_errors": outbox_errors, "outbox_sent": outbox_sent, "outbox_retry_due": outbox_retry_due}


@app.get("/api/stats/bucket-accounting")
def api_bucket_accounting(view: str = Query(default="legacy"), items: bool = Query(default=False)) -> dict[str, Any]:
    """Read-only accounting. view=legacy — детальный (UI views overlap); view=visual — 12 buckets, sum=total."""
    if view == "visual":
        from app import visual_accounting as va
        with connect() as con:
            return va.build_visual_accounting(con, include_items=items)
    from app.bucket_accounting import build_bucket_accounting
    with connect() as con:
        return build_bucket_accounting(con, include_items=False)


@app.get("/api/stats/folder-accounting")
def api_folder_accounting(items: bool = Query(default=False)) -> dict[str, Any]:
    """Read-only operator folder tree where folder counts sum to total raw."""
    from app.folder_accounting import build_folder_accounting
    with connect() as con:
        return build_folder_accounting(con, include_items=items)


@app.get("/api/radar/today")
def api_radar_today(limit: int = Query(default=80, ge=10, le=300)) -> dict[str, Any]:
    """One-screen operator radar: urgent, ready, blocked, and follow-ups."""
    with connect() as con:
        def rows(where: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
            return [row_to_dict(r) for r in con.execute(
                f"""
                SELECT c.id, c.buyer_code, c.buyer_name, c.event_type, c.claim_kind, c.status,
                       c.priority, c.confidence, c.deadline_at, c.ready_for_export, c.needs_review,
                       c.state, c.missing_json, c.quality_json, e.direction, e.subject, e.from_addr,
                       e.received_at, e.snippet
                FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
                WHERE {where}
                ORDER BY
                    CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
                    COALESCE(c.deadline_at, e.received_at) ASC,
                    c.id DESC
                LIMIT ?
                """,
                params + (limit,),
            ).fetchall()]
        urgent = rows("c.deadline_at IS NOT NULL AND c.state NOT IN ('closed','exported') AND (c.priority IN ('critical','high') OR c.deadline_at <= datetime('now', '+1 day'))")
        ready = rows("c.ready_for_export=1 AND c.state='ready_to_1c'")
        blocked = rows("c.state='needs_review' AND (c.missing_json LIKE '%photo_evidence%' OR c.missing_json LIKE '%service_document%' OR c.missing_json LIKE '%strong_key%' OR c.missing_json LIKE '%goods_or_document%')")
        followups = rows("c.state IN ('needs_link','linked_event') OR c.event_type IN ('followup_reminder','supplier_decision')")
    return {
        "schema": "readmail-new-radar-v1",
        "generated_at": utcnow(),
        "counts": {"urgent": len(urgent), "ready": len(ready), "blocked": len(blocked), "followups": len(followups)},
        "urgent": urgent,
        "ready": ready,
        "blocked": blocked,
        "followups": followups,
    }







def _mail_domain(addr: str | None) -> str:
    text = (addr or "").lower()
    import re
    m = re.search(r"[a-z0-9._%+-]+@([a-z0-9.-]+\.[a-z]{2,})", text)
    return m.group(1) if m else ""


def _folder_role(folder: str, configured: bool, excluded: bool) -> str:
    name = (folder or "").lower()
    if excluded:
        return "excluded_noise"
    if configured:
        return "customer_intake"
    if any(x in name for x in ["возв", "претенз", "return", "claim", "claim"]):
        return "possible_customer_intake"
    if any(x in name for x in ["sent", "отправ", "исход"]):
        return "outbound_context"
    return "unknown_do_not_import_until_checked"


@app.get("/api/setup/folder-map")
def api_setup_folder_map(live: bool = Query(default=False)) -> dict[str, Any]:
    """Wizard data for the first safe rollout: which folders are configured and which are risky.

    This does not import mail. Live mode only asks IMAP for selectable folder names.
    """
    apply_runtime_settings()
    configured = settings.folders or ["INBOX"]
    available: list[str] = []
    live_result: dict[str, Any] | None = None
    live_items: list[dict[str, Any]] = []
    if live:
        live_result = list_imap_folders()
        if live_result.get("ok"):
            live_items = list(live_result.get("items") or [])
            available = [str(x.get("raw_name") or x.get("name") or "") for x in live_items]
        else:
            available = []
    else:
        available = configured
    import re
    exclude_re = re.compile(settings.imap_exclude_folders_regex) if settings.imap_exclude_folders_regex else None
    configured_set = {f.strip().lower() for f in configured}
    discover_all = settings.discover_all_folders
    items = []
    live_by_raw = {str(x.get("raw_name") or x.get("name") or ""): x for x in live_items}
    for folder in sorted(set(available + configured), key=lambda f: (0 if str(f).upper()=="INBOX" else 1, decode_imap_utf7(str(f)).lower())):
        display = str((live_by_raw.get(folder) or {}).get("display_name") or decode_imap_utf7(folder))
        excluded = bool(exclude_re.search(folder) or exclude_re.search(display)) if exclude_re else False
        is_configured = discover_all or folder.strip().lower() in configured_set
        role = _folder_role(display, is_configured, excluded)
        risk = "low" if role == "customer_intake" else "medium" if role.startswith("possible") else "high" if role == "unknown_do_not_import_until_checked" else "info"
        items.append({"folder": folder, "raw_name": folder, "display_name": display, "configured": is_configured, "excluded": excluded, "role": role, "risk": risk})
    warnings = []
    if discover_all:
        warnings.append("IMAP_FOLDERS=* включён: для первого реального теста безопаснее указать только клиентские папки явно.")
    if not settings.imap_username or not settings.imap_password:
        warnings.append("IMAP логин/пароль не заданы — live-карта папок будет недоступна.")
    if not settings.configured_folders_are_customer:
        warnings.append("CONFIGURED_FOLDERS_ARE_CUSTOMER=false: новые письма чаще будут уходить в external_unknown/needs_review.")
    return {
        "schema": "readmail-new-folder-map-v1.12",
        "generated_at": utcnow(),
        "configured_folders": configured,
        "discover_all_folders": discover_all,
        "configured_folders_are_customer": settings.configured_folders_are_customer,
        "exclude_regex": settings.imap_exclude_folders_regex,
        "live": live,
        "live_result": live_result,
        "warnings": warnings,
        "items": items,
        "recommendation": "Для первого запуска: IMAP_FOLDERS=явные клиентские папки, IMAP_LIMIT=30, IMAP_TOTAL_LIMIT=100, AUTO_DELIVER_OUTBOX=false.",
    }


@app.post("/api/setup/folders/save")
def api_setup_save_folders(update: FolderSelectionUpdate) -> dict[str, Any]:
    """Save selected IMAP folders from the live folder picker.

    Folder values are raw IMAP names. Display names are only for the UI.
    """
    selected: list[str] = []
    seen: set[str] = set()
    for f in update.folders or []:
        val = str(f or "").strip()
        if not val or val in seen:
            continue
        selected.append(val)
        seen.add(val)
    if not selected:
        raise HTTPException(status_code=400, detail="Выбери хотя бы одну папку")
    result = update_settings_from_panel({
        "IMAP_FOLDERS": ",".join(selected),
        "CONFIGURED_FOLDERS_ARE_CUSTOMER": True,
    })
    return {
        "ok": True,
        "saved_folders": selected,
        "display_folders": [decode_imap_utf7(x) for x in selected],
        "settings": result,
    }


@app.get("/api/setup/intake-sources")
def api_setup_intake_sources(limit: int = Query(default=50, ge=10, le=300)) -> dict[str, Any]:
    """What the imported data looks like by folders/senders/domains.

    Helps catch the case where a wrong folder is imported and internal/outbound mail pollutes the queue.
    """
    apply_runtime_settings()
    with connect() as con:
        folder_rows = [dict(r) for r in con.execute(
            "SELECT mailbox, COUNT(*) count FROM raw_emails GROUP BY mailbox ORDER BY count DESC LIMIT ?", (limit,)
        ).fetchall()]
        sender_rows = [dict(r) for r in con.execute(
            "SELECT from_addr, COUNT(*) count FROM raw_emails GROUP BY from_addr ORDER BY count DESC LIMIT ?", (limit,)
        ).fetchall()]
        domain_counts: dict[str, int] = {}
        for r in con.execute("SELECT from_addr FROM raw_emails WHERE from_addr IS NOT NULL"):
            d = _mail_domain(r["from_addr"])
            if d:
                domain_counts[d] = domain_counts.get(d, 0) + 1
        learned = [dict(r) for r in con.execute(
            "SELECT identity_type, identity_value, buyer_code, buyer_name, source, confidence, seen_count FROM buyer_identities ORDER BY seen_count DESC, identity_value LIMIT ?",
            (limit,),
        ).fetchall()]
        unknown = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.state, c.event_type, c.claim_kind, c.missing_json, c.quality_json,
                   e.mailbox, e.subject, e.from_addr, e.received_at, e.snippet
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE (c.buyer_code IS NULL OR c.buyer_code='') AND c.state NOT IN ('closed','ignored_internal')
            ORDER BY c.id DESC LIMIT ?
            """,
            (limit,),
        ).fetchall()]
    domain_items = [{"domain": k, "count": v, "company_domain": k in set(settings.company_domain_list)} for k, v in sorted(domain_counts.items(), key=lambda kv: (-kv[1], kv[0]))[:limit]]
    risky_domains = [d for d in domain_items if d["company_domain"]]
    return {
        "schema": "readmail-new-intake-sources-v1.12",
        "generated_at": utcnow(),
        "folders": folder_rows,
        "senders": sender_rows,
        "domains": domain_items,
        "learned_identities": learned,
        "unknown_buyer_samples": unknown,
        "warnings": ["В импорте есть домены нашей компании — проверь папки/COMPANY_DOMAINS." if risky_domains else ""],
    }


@app.get("/api/diagnostics/quality")
def api_diagnostics_quality(limit: int = Query(default=30, ge=5, le=200)) -> dict[str, Any]:
    """Post-test-run diagnostic: exact reasons why mail is blocked or unsafe.

    This is intentionally opinionated: it highlights issues that would have caused the old project to lie.
    """
    apply_runtime_settings()
    suspicious_words = {"сумма","цена","причина","товар","товара","количество","поставщик","покупатель","артикул","бренд","наименование","описание"}
    with connect() as con:
        ready_bad = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.state, c.event_type, c.claim_kind, c.ready_for_export, c.missing_json, c.quality_json,
                   c.fields_json, e.direction, e.mailbox, e.subject, e.from_addr, e.received_at, e.snippet
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.ready_for_export=1 AND (
                c.event_type<>'new_return' OR e.direction<>'inbound_customer' OR c.missing_json<>'[]' OR c.quality_json<>'[]'
            )
            ORDER BY c.id DESC LIMIT ?
            """, (limit,)
        ).fetchall()]
        missing_evidence = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.state, c.event_type, c.claim_kind, c.missing_json, c.quality_json,
                   c.fields_json, e.mailbox, e.subject, e.from_addr, e.received_at, e.snippet
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.state='needs_review' AND (c.missing_json LIKE '%photo_evidence%' OR c.missing_json LIKE '%service_document%')
            ORDER BY c.id DESC LIMIT ?
            """, (limit,)
        ).fetchall()]
        unknown_buyers = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.state, c.event_type, c.claim_kind, c.missing_json, c.fields_json,
                   e.mailbox, e.subject, e.from_addr, e.received_at, e.snippet
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE (c.buyer_code IS NULL OR c.buyer_code='') AND c.state NOT IN ('closed','ignored_internal')
            ORDER BY c.id DESC LIMIT ?
            """, (limit,)
        ).fetchall()]
        without_outbox = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.state, c.event_type, c.claim_kind, e.mailbox, e.subject, e.from_addr, e.received_at
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.state NOT IN ('ignored_internal','closed') AND NOT EXISTS (SELECT 1 FROM outbox o WHERE o.case_id=c.id)
            ORDER BY c.id DESC LIMIT ?
            """, (limit,)
        ).fetchall()]
        outbox_errors = [row_to_dict(r) for r in con.execute(
            "SELECT id, case_id, event_type, channel, attempt_count, last_error, next_attempt_at, created_at FROM outbox WHERE status='error' ORDER BY id DESC LIMIT ?",
            (limit,),
        ).fetchall()]
        all_parts = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.fields_json, e.subject, e.from_addr FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.fields_json LIKE '%part_number%' ORDER BY c.id DESC LIMIT 1000
            """
        ).fetchall()]
    suspicious_parts = []
    for row in all_parts:
        fields = row.get("fields") or {}
        parts = fields.get("part_number") or fields.get("parts") or []
        if isinstance(parts, str):
            parts = [parts]
        for p in parts if isinstance(parts, list) else []:
            part = str(p).strip().lower()
            if not part:
                continue
            if part in suspicious_words or len(part) < int(settings.part_number_min_len or 3) or len(part) > int(settings.part_number_max_len or 50):
                suspicious_parts.append({"case_id": row.get("id"), "part_number": p, "subject": row.get("subject"), "from_addr": row.get("from_addr")})
                break
        if len(suspicious_parts) >= limit:
            break
    buckets = {
        "ready_bad": len(ready_bad),
        "missing_evidence": len(missing_evidence),
        "unknown_buyers": len(unknown_buyers),
        "cases_without_outbox": len(without_outbox),
        "outbox_errors": len(outbox_errors),
        "suspicious_parts": len(suspicious_parts),
    }
    release_blockers = [k for k, v in buckets.items() if v and k in {"ready_bad", "cases_without_outbox", "outbox_errors"}]
    recommendation = "Можно расширять тестовый лимит" if not release_blockers else "Сначала исправить блокеры: " + ", ".join(release_blockers)
    return {
        "schema": "readmail-new-quality-diagnostics-v1.12",
        "generated_at": utcnow(),
        "buckets": buckets,
        "release_blockers": release_blockers,
        "recommendation": recommendation,
        "samples": {
            "ready_bad": ready_bad,
            "missing_evidence": missing_evidence,
            "unknown_buyers": unknown_buyers,
            "cases_without_outbox": without_outbox,
            "outbox_errors": outbox_errors,
            "suspicious_parts": suspicious_parts,
        },
    }

@app.get("/api/control/dashboard")
def api_control_dashboard(limit: int = Query(default=100, ge=10, le=500)) -> dict[str, Any]:
    apply_runtime_settings()
    with connect() as con:
        return control_dashboard(con, limit=limit)


@app.get("/api/import/folders")
def api_import_folders() -> dict[str, Any]:
    apply_runtime_settings()
    return list_imap_folders()


@app.get("/api/import/server-counts")
def api_import_server_counts() -> dict[str, Any]:
    """Щиток серверной почты: писем на сервере vs в базе по папкам (для окна настроек)."""
    # Во время активного импорта IMAP-соединение занято — не дёргаем сервер.
    if _IMPORT_BG_STATE.get("running") or _IMPORT_LOCK.locked():
        return {"ok": False, "error": "Идёт импорт — сверка с сервером после завершения", "folders": []}
    from .imap_importer import server_counts
    r = server_counts()
    if r.get("ok"):
        _SERVER_COUNT_CACHE.update({"total": r["total_server"], "db": r["total_db"], "gap": r["total_gap"], "at": utcnow(), "ts": time.time()})
    return r


# Кэш счётчика сервера для зелёной плашки «Импорт». Плашка читает кэш мгновенно,
# а IMAP опрашивается в фоне не чаще раза в 5 минут (и не во время импорта).
_SERVER_COUNT_CACHE: dict[str, Any] = {"total": None, "db": None, "gap": None, "at": None, "ts": 0.0, "refreshing": False}


def _refresh_server_count_bg() -> None:
    if _SERVER_COUNT_CACHE.get("refreshing"):
        return
    if _IMPORT_BG_STATE.get("running") or _IMPORT_LOCK.locked():
        return
    _SERVER_COUNT_CACHE["refreshing"] = True

    def _run() -> None:
        try:
            from .imap_importer import server_counts
            r = server_counts()
            if r.get("ok"):
                _SERVER_COUNT_CACHE.update({"total": r["total_server"], "db": r["total_db"], "gap": r["total_gap"], "at": utcnow(), "ts": time.time()})
        except Exception:
            pass
        finally:
            _SERVER_COUNT_CACHE["refreshing"] = False

    threading.Thread(target=_run, name="readmail-server-count", daemon=True).start()


@app.get("/api/import/server-total")
def api_import_server_total() -> dict[str, Any]:
    """Лёгкий: общий счётчик писем на сервере из кэша (для плашки). IMAP не дёргает напрямую —
    обновляет кэш в фоне раз в 5 мин. Ориентир по объёму."""
    stale = (time.time() - float(_SERVER_COUNT_CACHE.get("ts") or 0)) > 300
    if stale:
        _refresh_server_count_bg()
    # «Застряло» = реально не скачанные (failed/мёртвые), БЕЗ дублей. Дешёвый запрос к БД.
    stuck = 0
    try:
        with connect() as con:
            stuck = con.execute(
                "SELECT COUNT(*) FROM import_uid_failures WHERE status='failed' OR attempts>=5"
            ).fetchone()[0]
    except Exception:
        pass
    return {
        "ok": True,
        "server_total": _SERVER_COUNT_CACHE.get("total"),
        "local_raw_total": _SERVER_COUNT_CACHE.get("db"),
        "count_gap_estimate": _SERVER_COUNT_CACHE.get("gap"),
        "failed_or_stuck": stuck,
        # Backward-compatible aliases for older panels.
        "total": _SERVER_COUNT_CACHE.get("total"),
        "db": _SERVER_COUNT_CACHE.get("db"),
        "gap": _SERVER_COUNT_CACHE.get("gap"),
        "stuck": stuck,
        "at": _SERVER_COUNT_CACHE.get("at"),
        "stale": stale and _SERVER_COUNT_CACHE.get("total") is None,
    }


@app.get("/api/import/quarantine")
def api_import_quarantine(
    status: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=2000),
) -> dict[str, Any]:
    """READ-ONLY список UID-сбоев импорта (карантин/ретрай). Для блока «Сверка почты».

    Не подключается к IMAP, не качает письма, ничего не меняет. Targeted backfill выполняется
    отдельным CLI `scripts/backfill_missing_imap_uids.py` (на панели — только подсказка)."""
    cols = ("account", "mailbox", "uid", "uidvalidity", "message_id", "stage", "error_type",
            "error_message", "attempts", "status", "recoverable", "first_seen_at",
            "last_seen_at", "next_retry_at")
    query = f"SELECT {', '.join(cols)} FROM import_uid_failures"
    params: list[Any] = []
    if status:
        query += " WHERE status=?"
        params.append(status)
    query += " ORDER BY last_seen_at DESC LIMIT ?"
    params.append(int(limit))
    items: list[dict[str, Any]] = []
    try:
        with connect() as con:
            items = [dict(row) for row in con.execute(query, params).fetchall()]
    except Exception as exc:
        return {"ok": False, "error": str(exc), "items": [], "read_only": True}
    return {"ok": True, "total": len(items), "items": items, "read_only": True}


@app.get("/api/import/quarantine/summary")
def api_import_quarantine_summary() -> dict[str, Any]:
    """READ-ONLY сводка карантина: по статусам + сколько готово к ретраю сейчас."""
    by_status: dict[str, int] = {}
    quarantined = retry_due = recoverable = 0
    try:
        with connect() as con:
            for row in con.execute("SELECT status, COUNT(*) AS n FROM import_uid_failures GROUP BY status"):
                by_status[str(row["status"])] = int(row["n"])
            quarantined = int(con.execute(
                "SELECT COUNT(*) FROM import_uid_failures WHERE status='quarantined'").fetchone()[0])
            recoverable = int(con.execute(
                "SELECT COUNT(*) FROM import_uid_failures WHERE recoverable=1 AND status<>'resolved'").fetchone()[0])
            retry_due = int(con.execute(
                "SELECT COUNT(*) FROM import_uid_failures WHERE status<>'resolved' "
                "AND (next_retry_at IS NULL OR next_retry_at<=?)", (utcnow(),)).fetchone()[0])
    except Exception as exc:
        return {"ok": False, "error": str(exc), "read_only": True}
    return {
        "ok": True,
        "by_status": by_status,
        "quarantined": quarantined,
        "recoverable_pending": recoverable,
        "retry_due_now": retry_due,
        "hint": "Targeted backfill: python3 scripts/backfill_missing_imap_uids.py "
                "--from-missing audit_out/imap_reconcile_missing_server_uids.jsonl --apply",
        "read_only": True,
    }


# ── Runtime control: pause/resume воркеров ────────────────────────────

@app.get("/api/runtime/status")
def api_runtime_status() -> dict[str, Any]:
    """Runtime status 2.0: сервер/auth/LAN, воркеры/паузы, import window, uidvalidity, outbox,
    AI cost (день/месяц), reconcile/quarantine. Read-only, БЕЗ секретов."""
    import datetime as _dt
    today = _dt.datetime.now(_dt.timezone.utc).date().isoformat()
    month = today[:7]
    # AI cost (из ledger, не из БД)
    ai_cost_today = ai_cost_month = 0.0
    try:
        agg = ai_cost_ledger.aggregate(by="day")
        ai_cost_today = float((agg["groups"].get(today) or {}).get("total_cost") or 0.0)
        ai_cost_month = round(sum(float(v.get("total_cost") or 0.0)
                                  for k, v in agg["groups"].items() if str(k).startswith(month)), 6)
    except Exception:
        pass
    # outbox + reconcile + quarantine (read-only)
    outbox = {"new": 0, "error": 0, "sent": 0}
    quarantine = 0
    cases_total = 0
    try:
        with connect() as con:
            for r in con.execute("SELECT status, COUNT(*) n FROM outbox GROUP BY status"):
                outbox[str(r["status"])] = int(r["n"])
            quarantine = int(con.execute(
                "SELECT COUNT(*) FROM import_uid_failures WHERE status='quarantined'").fetchone()[0])
            cases_total = int(con.execute("SELECT COUNT(*) FROM cases").fetchone()[0])
    except Exception:
        pass
    recon = {}
    try:
        p = Path(__file__).resolve().parent.parent / "audit_out" / "imap_reconcile_summary.json"
        if p.exists():
            d = json.loads(p.read_text(encoding="utf-8"))
            recon = {"server_total": d.get("server_total"), "local_raw": d.get("local_raw_total"),
                     "missing_local": d.get("missing_local_total"), "fetch_failed": d.get("fetch_failed_total"),
                     "checked_at": d.get("checked_at")}
    except Exception:
        pass
    return {
        "ok": True,
        "server": server_core.public_status(),
        "auth": {"required": server_core.auth_required(), "enforced": server_core.auth_required(),
                 "bootstrap_required": auth_mod.bootstrap_required()},
        "runtime": runtime_control.get_runtime_status(),
        "import_window": {
            "enabled": bool(getattr(settings, "import_window_enabled", False)),
            "import_from_datetime": getattr(settings, "import_from_datetime", "") or None,
            "skip_before_start": bool(getattr(settings, "skip_before_start", True)),
        },
        "uidvalidity_in_identity": True,
        "concurrency": {
            "static_workers": getattr(settings, "static_workers", None),
            "stage2_workers": getattr(settings, "stage2_workers", None),
            "ai_text_workers": getattr(settings, "ai_text_workers", None),
            "ai_vision_workers": getattr(settings, "ai_vision_workers", None),
            "outbox_workers": getattr(settings, "outbox_workers", None),
            "max_parallel_cases": getattr(settings, "max_parallel_cases", None),
        },
        "ai": {"enabled": bool(getattr(settings, "enable_ai", False)),
               "cost_today": ai_cost_today, "cost_month": ai_cost_month},
        "outbox": outbox,
        "delivery_enabled": bool(getattr(settings, "auto_deliver_outbox", False)),
        "cases_total": cases_total,
        "quarantine": quarantine,
        "reconcile": recon,
        "autopilot": {
            "enabled": bool(_AUTOPILOT_STATE.get("enabled")),
            "running_cycle": bool(_AUTOPILOT_STATE.get("running_cycle")),
            "cycle_count": _AUTOPILOT_STATE.get("cycle_count"),
            "next_cycle_at": _AUTOPILOT_STATE.get("next_cycle_at"),
            "last_error": _AUTOPILOT_STATE.get("last_error"),
        },
    }


@app.get("/api/dashboard/overview")
def api_dashboard_overview() -> dict[str, Any]:
    """Агрегирующий снимок для экрана «Пульт». Read-only, без секретов, не падает без snapshot'ов."""
    return dashboard_mod.build_overview()


@app.post("/api/runtime/pause")
def api_runtime_pause_all() -> dict[str, Any]:
    return runtime_control.pause("all")


@app.post("/api/runtime/resume")
def api_runtime_resume_all() -> dict[str, Any]:
    return runtime_control.resume("all")


@app.post("/api/runtime/pause/{worker}")
def api_runtime_pause_worker(worker: str) -> dict[str, Any]:
    return runtime_control.pause(worker)


@app.post("/api/runtime/resume/{worker}")
def api_runtime_resume_worker(worker: str) -> dict[str, Any]:
    return runtime_control.resume(worker)


# ── UI mode / developer visibility ────────────────────────────────────

@app.get("/api/ui/mode")
def api_ui_mode(role: str = Query(default="operator")) -> dict[str, Any]:
    return {"ok": True, **server_core.ui_mode(role)}


# ── AI cost metrics ───────────────────────────────────────────────────

@app.get("/api/metrics/ai-cost")
def api_metrics_ai_cost(by: str = Query(default="day")) -> dict[str, Any]:
    return {"ok": True, **ai_cost_ledger.aggregate(by=by)}


@app.get("/api/metrics/ai-cost/by-provider")
def api_metrics_ai_cost_provider() -> dict[str, Any]:
    return {"ok": True, **ai_cost_ledger.aggregate(by="provider")}


@app.get("/api/metrics/ai-cost/by-mode")
def api_metrics_ai_cost_mode() -> dict[str, Any]:
    return {"ok": True, **ai_cost_ledger.aggregate(by="mode")}


@app.get("/api/metrics/ai-cost/by-supplier")
def api_metrics_ai_cost_supplier() -> dict[str, Any]:
    return {"ok": True, **ai_cost_ledger.aggregate(by="supplier")}


@app.get("/api/metrics/ai-cost/by-claim-kind")
def api_metrics_ai_cost_claim_kind() -> dict[str, Any]:
    return {"ok": True, **ai_cost_ledger.aggregate(by="claim_kind")}


# ── Processing metrics (skeleton) ─────────────────────────────────────

@app.get("/api/metrics/processing")
def api_metrics_processing() -> dict[str, Any]:
    """Сводка обработки по состояниям/типам. Read-only."""
    out: dict[str, Any] = {"ok": True, "by_state": {}, "by_event_type": {}, "totals": {}}
    try:
        with connect() as con:
            out["by_state"] = {str(r["state"]): int(r["n"]) for r in con.execute(
                "SELECT state, COUNT(*) n FROM cases GROUP BY state")}
            out["by_event_type"] = {str(r["event_type"]): int(r["n"]) for r in con.execute(
                "SELECT event_type, COUNT(*) n FROM cases GROUP BY event_type")}
            total = int(con.execute("SELECT COUNT(*) FROM cases").fetchone()[0])
            ready = int(con.execute("SELECT COUNT(*) FROM cases WHERE state='ready_to_1c'").fetchone()[0])
            review = int(con.execute("SELECT COUNT(*) FROM cases WHERE state='needs_review'").fetchone()[0])
            out["totals"] = {"total_cases": total, "ready_to_1c": ready, "needs_review": review,
                             "auto_export_percent": round(ready * 100 / total, 2) if total else 0}
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return out


@app.get("/api/metrics/suppliers")
def api_metrics_suppliers() -> dict[str, Any]:
    out: dict[str, Any] = {"ok": True, "suppliers": []}
    try:
        with connect() as con:
            rows = con.execute(
                """
                SELECT buyer_code,
                       COUNT(*) total,
                       SUM(CASE WHEN state='ready_to_1c' THEN 1 ELSE 0 END) ready,
                       SUM(CASE WHEN state='needs_review' THEN 1 ELSE 0 END) review
                FROM cases GROUP BY buyer_code ORDER BY total DESC
                """
            ).fetchall()
            for r in rows:
                total = int(r["total"] or 0)
                out["suppliers"].append({
                    "buyer_code": r["buyer_code"] or "unknown",
                    "total": total,
                    "ready_to_1c": int(r["ready"] or 0),
                    "needs_review": int(r["review"] or 0),
                    "auto_export_percent": round(int(r["ready"] or 0) * 100 / total, 2) if total else 0,
                })
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return out


@app.get("/api/metrics/defect")
def api_metrics_defect() -> dict[str, Any]:
    out: dict[str, Any] = {"ok": True}
    try:
        with connect() as con:
            out["defect_total"] = int(con.execute(
                "SELECT COUNT(*) FROM cases WHERE claim_kind='defect'").fetchone()[0])
            out["defect_needs_review"] = int(con.execute(
                "SELECT COUNT(*) FROM cases WHERE claim_kind='defect' AND state='needs_review'").fetchone()[0])
    except Exception as exc:
        return {"ok": False, "error": str(exc)}
    return out


@app.post("/api/import/imap")
def api_import_imap(limit: int | None = Query(default=None), folder: str | None = Query(default=None), search: str | None = Query(default=None)) -> dict[str, Any]:
    apply_runtime_settings()
    _log("import", "Ручной IMAP импорт: старт", details={"limit": limit, "folder": folder, "search": search})
    if folder is None and limit is None and search is None:
        return _safe_import_cycle()
    if not _IMPORT_LOCK.acquire(blocking=False):
        return {"ok": False, "skipped": True, "reason": "import_already_running"}
    try:
        folders = [folder] if folder else None
        result = import_from_imap_raw(limit=limit, folders=folders, search=search)
        _log("import", "Ручной IMAP импорт: завершён", level="ok" if result.get("ok", True) else "warn", details=result)
        queued = _auto_queue_ready_if_enabled()
        if queued:
            result["auto_outbox"] = queued
        global _LAST_IMPORT_RESULT
        _LAST_IMPORT_RESULT = {**result, "finished_at": utcnow()}
        return _LAST_IMPORT_RESULT
    except Exception as exc:
        _log("import", "Ручной IMAP импорт: ошибка", level="error", details={"error": str(exc)})
        raise
    finally:
        _IMPORT_LOCK.release()


@app.post("/api/import/fetch-only")
def api_import_fetch_only(
    limit: int | None = Query(default=None),
    folder: str | None = Query(default=None),
    search: str | None = Query(default=None),
) -> dict[str, Any]:
    """Загрузить письма из IMAP без классификации (паттерны запускаются отдельно).

    Параметры:
        limit: лимит на папку (по умолчанию из IMAP_LIMIT)
        folder: конкретная папка (по умолчанию все настроенные)
        search: IMAP search query (по умолчанию ALL)

    Это Stage 1: IMAP → raw_emails.
    """
    apply_runtime_settings()
    if not _IMPORT_LOCK.acquire(blocking=False):
        return {"ok": False, "skipped": True, "reason": "import_already_running"}
    try:
        folders = [folder] if folder else None
        result = import_from_imap_raw(limit=limit, folders=folders, search=search)
        global _LAST_IMPORT_RESULT
        _LAST_IMPORT_RESULT = {**result, "finished_at": utcnow(), "classify": False}
        return _LAST_IMPORT_RESULT
    finally:
        _IMPORT_LOCK.release()


@app.post("/api/import/fetch-only-raw")
def api_import_raw_v2(
    limit: int | None = Query(default=None),
    folder: str | None = Query(default=None),
    search: str | None = Query(default=None),
    batch_size: int | None = Query(default=None),
    total_limit: int | None = Query(default=None),
) -> dict[str, Any]:
    """[v2] Импорт писем из IMAP — только raw_emails, без классификации.

    Самая безопасная операция: письма забираются с сервера, парсятся,
    сохраняются в raw_emails со статусом 'imported'.
    Классификация, AI и паттерны не запускаются.

    Параметры:
        limit: лимит на папку
        folder: конкретная папка (иначе все настроенные)
        search: IMAP search query
        batch_size: сколько UID в одном batch запросе
        total_limit: общий лимит за весь импорт
    """
    apply_runtime_settings()
    if not _IMPORT_LOCK.acquire(blocking=False):
        return {"ok": False, "skipped": True, "reason": "import_already_running"}
    try:
        if batch_size is not None:
            old = getattr(settings, "imap_batch_size", 5)
            try:
                settings.imap_batch_size = max(1, min(int(batch_size), 50))
            except Exception:
                pass
        folders = [folder] if folder else None
        effective_limit = limit if limit is not None else settings.imap_limit
        result = import_from_imap_raw(limit=effective_limit, folders=folders, search=search)
        global _LAST_IMPORT_RESULT
        _LAST_IMPORT_RESULT = {**result, "finished_at": utcnow(), "classify": False}
        return _LAST_IMPORT_RESULT
    finally:
        _IMPORT_LOCK.release()


@app.get("/api/v2/import/status")
def api_import_status_v2() -> dict[str, Any]:
    """[v2] Статус текущего/последнего импорта.

    Возвращает live heartbeat с диагностикой possible_hang.
    """
    apply_runtime_settings()
    with connect() as con:
        job = get_import_job_status(con)
    if not job:
        return {
            "ok": True,
            "running": False,
            "job_id": None,
            "stage": None,
            "message": "No import job found",
        }
    return {
        "ok": True,
        "running": job.get("status") == "running",
        "job_id": job.get("job_id"),
        "stage": job.get("current_stage"),
        "folder": job.get("current_folder"),
        "display_folder": job.get("current_display_folder"),
        "uid": job.get("current_uid"),
        "processed": job.get("processed_count") or 0,
        "imported": job.get("imported_count") or 0,
        "imported_this_run": job.get("imported_count") or 0,
        "skipped": job.get("skipped_count") or 0,
        "failed": job.get("failed_count") or 0,
        "errors_count": job.get("error_count") or 0,
        "started_at": job.get("started_at"),
        "finished_at": job.get("finished_at"),
        "last_heartbeat_at": job.get("last_heartbeat_at"),
        "possible_hang": job.get("possible_hang", False),
        "mode": job.get("mode"),
        "status": job.get("status"),
    }


@app.get("/api/v2/import/errors")
def api_import_errors_v2(
    job_id: str | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=1000),
) -> dict[str, Any]:
    """[v2] Ошибки импорта."""
    apply_runtime_settings()
    with connect() as con:
        errors = get_import_errors(con, job_id=job_id, limit=limit)
    return {"ok": True, "count": len(errors), "items": errors}


@app.get("/api/v2/import/quarantine")
def api_import_quarantine_v2(
    limit: int = Query(default=100, ge=1, le=1000),
) -> dict[str, Any]:
    """[v2] UID в карантине (повторно падающие письма)."""
    apply_runtime_settings()
    with connect() as con:
        items = get_quarantined_uids(con, limit=limit)
    return {"ok": True, "count": len(items), "items": items}


@app.get("/api/v2/import/health")
def api_import_health_v2() -> dict[str, Any]:
    """Compact import health for the operator panel."""
    apply_runtime_settings()
    with connect() as con:
        live_import = bool(_IMPORT_LOCK.locked() or _IMPORT_BG_STATE.get("running"))
        if not live_import:
            con.execute(
                "UPDATE import_jobs SET status='abandoned', finished_at=?, last_heartbeat_at=? WHERE status='running'",
                (utcnow(), utcnow()),
            )
        latest = con.execute("SELECT * FROM import_jobs ORDER BY id DESC LIMIT 1").fetchone()
        quarantine_count = con.execute(
            "SELECT COUNT(*) c FROM import_uid_failures WHERE status='quarantined'"
        ).fetchone()["c"]
        oversized_count = con.execute(
            """
            SELECT COUNT(*) c
            FROM import_uid_failures
            WHERE stage='oversized' AND status IN ('failed','quarantined')
            """
        ).fetchone()["c"]
        retry_pending_count = con.execute(
            "SELECT COUNT(*) c FROM import_uid_failures WHERE status='retry_pending'"
        ).fetchone()["c"]
        recent_error = con.execute(
            """
            SELECT * FROM import_errors
            WHERE job_id=(SELECT job_id FROM import_jobs ORDER BY id DESC LIMIT 1)
            ORDER BY id DESC LIMIT 1
            """
        ).fetchone()

    job = row_to_dict(latest) if latest else {}
    recent = row_to_dict(recent_error) if recent_error else {}
    status = str(job.get("status") or "idle")
    heartbeat = job.get("last_heartbeat_at")
    heartbeat_age_seconds = None
    possible_hang = False
    if heartbeat:
        try:
            last = datetime.fromisoformat(str(heartbeat).replace("Z", "+00:00"))
            if last.tzinfo is None:
                last = last.replace(tzinfo=timezone.utc)
            heartbeat_age_seconds = int((datetime.now(timezone.utc) - last).total_seconds())
            possible_hang = status == "running" and heartbeat_age_seconds > 120
        except Exception:
            heartbeat_age_seconds = None

    level = "ok"
    messages: list[str] = []
    if status == "error" or possible_hang or (status == "running" and not bool(_IMPORT_LOCK.locked() or _IMPORT_BG_STATE.get("running"))):
        level = "error"
        messages.append("импорт завис или был брошен")
    elif status == "abandoned":
        level = "warn"
        messages.append("последний импорт был прерван")
    if quarantine_count:
        if level == "ok":
            level = "warn"
        messages.append(f"карантин UID: {quarantine_count}")
    if oversized_count:
        if level == "ok":
            level = "warn"
        messages.append(f"oversized: {oversized_count}")
    if retry_pending_count:
        messages.append(f"на повтор: {retry_pending_count}")
    if not messages:
        messages.append("импорт без явных проблем")

    return {
        "ok": True,
        "level": level,
        "message": "; ".join(messages),
        "limits": {
            "imap_max_raw_email_mb": int(getattr(settings, "imap_max_raw_email_mb", 25) or 25),
            "import_max_attachment_mb": int(getattr(settings, "import_max_attachment_mb", 10) or 10),
            "imap_batch_size": int(getattr(settings, "imap_batch_size", 20) or 20),
            "imap_limit": int(getattr(settings, "imap_limit", 200) or 200),
            "imap_total_limit": int(getattr(settings, "imap_total_limit", 2000) or 2000),
        },
        "counts": {
            "quarantine": quarantine_count,
            "oversized": oversized_count,
            "retry_pending": retry_pending_count,
        },
        "last_job": {
            "job_id": job.get("job_id"),
            "status": status,
            "stage": job.get("current_stage"),
            "folder": job.get("current_display_folder") or job.get("current_folder"),
            "processed": job.get("processed_count") or 0,
            "imported": job.get("imported_count") or 0,
            "errors": job.get("error_count") or 0,
            "started_at": job.get("started_at"),
            "finished_at": job.get("finished_at"),
            "last_heartbeat_at": heartbeat,
            "heartbeat_age_seconds": heartbeat_age_seconds,
            "possible_hang": possible_hang,
        },
        "recent_error": {
            "type": recent.get("error_type"),
            "stage": recent.get("stage"),
            "mailbox": recent.get("display_folder") or recent.get("mailbox"),
            "uid": recent.get("uid"),
            "message": recent.get("error_message"),
            "created_at": recent.get("created_at"),
        } if recent else None,
    }


@app.post("/api/v2/import/quarantine/retry")
def api_import_quarantine_retry_v2(
    mailbox: str = Query(...),
    uid: str = Query(...),
    stage: str = Query(default="fetch_single"),
) -> dict[str, Any]:
    """[v2] Снять UID с карантина для повторной попытки."""
    apply_runtime_settings()
    with connect() as con:
        retry_quarantined_uid(con, mailbox=mailbox, uid=uid, stage=stage)
    return {"ok": True, "mailbox": mailbox, "uid": uid, "stage": stage, "status": "retry_pending"}


@app.post("/api/v2/mail/process-imported")
def api_process_imported_v2(
    limit: int = Query(default=500, ge=1, le=5000),
    supplier: str | None = Query(default=None),
    dry_run: bool = Query(default=False),
) -> dict[str, Any]:
    """[v2] Обработать импортированные письма: паттерны → кейсы.

    Stage 2: raw_emails (status='imported') → parser → pattern matcher → cases.

    Параметры:
        limit: максимум писем за раз
        supplier: фильтр по поставщику (from_addr содержит)
        dry_run: только подсчитать, не создавать кейсы
    """
    apply_runtime_settings()
    _log("process", "Обработка импортированных писем: старт",
         details={"limit": limit, "supplier": supplier, "dry_run": dry_run})
    try:
        result = process_imported_emails(limit=limit, supplier=supplier, dry_run=dry_run, manual_review_gate=True)
        _log("process", "Обработка импортированных писем: завершена",
             level="ok" if result.get("ok") else "error", details=result)
        return result
    except Exception as exc:
        _log("process", "Обработка импортированных писем: ошибка",
             level="error", details={"error": str(exc)})
        return {"ok": False, "error": str(exc)}


@app.post("/api/import/eml")
def api_import_eml(path: str = "/app/data/eml_inbox") -> dict[str, Any]:
    return import_from_eml_dir(path)


_PATTERNS_STATE: dict[str, Any] = {"running": False, "total": 0, "processed": 0, "started_at": None, "finished_at": None, "error": None}


def _load_attachments_with_text(con: Any, raw_email_id: int) -> list[dict[str, Any]]:
    """Вложения из БД + текст из Excel/PDF (с диска) для детекции документов брака.

    Excel/PDF читается из file_path (≤ лимит) → extracted_text, чтобы classify_defect_documents
    видел СОДЕРЖИМОЕ (акт ТОРГ-2/заключение), а не только имя файла. Read-only по файлам.
    """
    from .email_parser import _extract_xlsx_text, _extract_pdf_text
    MAX_READ = 3 * 1024 * 1024
    out: list[dict[str, Any]] = []
    for a in con.execute(
            "SELECT filename, content_type, size_bytes, file_path FROM attachments WHERE raw_email_id=?",
            (raw_email_id,)):
        d = dict(a)
        d["extracted_text"] = ""
        low = str(d.get("filename") or "").lower()
        fp = d.get("file_path")
        if fp and int(d.get("size_bytes") or 0) < MAX_READ and low.endswith((".xls", ".xlsx", ".pdf")):
            try:
                from pathlib import Path as _P
                p = _P(fp)
                if p.exists():
                    data = p.read_bytes()
                    if low.endswith((".xls", ".xlsx")):
                        d["extracted_text"] = _extract_xlsx_text(data)[:8000]
                    elif low.endswith(".pdf"):
                        d["extracted_text"] = _extract_pdf_text(data)[:8000]
            except Exception:
                pass
        out.append(d)
    return out


def _classify_pending(
    con: Any,
    *,
    only_missing: bool = True,
    stop_event: Any = None,
    state: dict[str, Any] | None = None,
    buyer_rules: Any = None,
    learned: Any = None,
    existing_cases: list[dict[str, Any]] | None = None,
    manual_review_gate: bool = False,
) -> int:
    """Классифицирует raw-письма паттернами в кейсы. Переиспользуется ручной кнопкой
    «Паттерны» и конвейером автопилота. Вызывающий держит _PIPELINE_LOCK (один писатель).

    only_missing=True — только raw без кейса (инкрементально, для конвейера 24/7).
    Возвращает число обработанных писем.
    """
    if buyer_rules is None:
        buyer_rules = load_buyer_rules()
    if learned is None:
        learned = load_buyer_identities(con)
    if existing_cases is None:
        existing_cases = []
    if only_missing:
        rows = con.execute(
            """
            SELECT r.*
            FROM raw_emails r
            LEFT JOIN cases c ON c.raw_email_id=r.id
            WHERE c.id IS NULL
              AND COALESCE(r.status, 'imported') <> 'duplicate'
            ORDER BY r.id
            """
        ).fetchall()
    else:
        rows = con.execute(
            "SELECT * FROM raw_emails WHERE COALESCE(status, 'imported') <> 'duplicate' ORDER BY id"
        ).fetchall()
    if state is not None:
        state["total"] = len(rows)
    count = 0
    for row in rows:
        if stop_event is not None and stop_event.is_set():
            break
        email_data = row_to_dict(row) or {}
        if state is not None:
            state["current_raw_id"] = int(row["id"])
            state["current_subject"] = str(row["subject"] or "")[:160]
            state["current_stage"] = "attachments"
        email_data["attachments"] = _load_attachments_with_text(con, int(row["id"]))
        # Кэш видимого текста: если в raw_emails ещё нет — расклеиваем один раз и
        # пишем обратно, чтобы следующие reprocess не парсили 25КБ HTML заново.
        _stored_vt = email_data.get("visible_text")
        if not _stored_vt:
            if state is not None:
                state["current_stage"] = "select_visible_text"
            _vt = select_visible_text(email_data.get("body_text"), email_data.get("body_html"))
            email_data["visible_text"] = _vt
            try:
                if state is not None:
                    state["current_stage"] = "cache_visible_text"
                con.execute("UPDATE raw_emails SET visible_text=? WHERE id=?", (_vt, row["id"]))
            except Exception:
                pass
        # Полный перепрогон НЕ трогает кейсы, добитые ИИ (иначе паттерны затрут ИИ-поля).
        if not only_missing:
            _ex = con.execute("SELECT payload_json FROM cases WHERE raw_email_id=? AND COALESCE(item_index,0)=0", (int(row["id"]),)).fetchone()
            if _ex:
                try:
                    _pj = loads(_ex["payload_json"] or "{}")
                except Exception:
                    _pj = {}
                if _pj.get("ai_overlay") or _pj.get("processing_source") == "ai":
                    count += 1
                    if state is not None:
                        state["processed"] = count
                    continue
        try:
            if state is not None:
                state["current_stage"] = "classify_email"
            case_data = classify_email(email_data, buyer_rules, learned_identities=learned, existing_cases=existing_cases)
            if manual_review_gate:
                case_data = force_operator_review(case_data)
            if state is not None:
                state["current_stage"] = "save_case"
            case_id = save_case(con, int(row["id"]), case_data)
            case_data["export"]["case_id"] = case_id
            if state is not None:
                state["current_stage"] = "save_export_case_id"
            con.execute("UPDATE cases SET export_json=?, updated_at=? WHERE id=?", (dumps(case_data.get("export") or {}), utcnow(), case_id))
            et = case_data.get("event_type", "")
            if et in ("followup_reminder", "followup_dialog", "supplier_decision", "correction_request", "marking_request", "ready_to_ship"):
                if state is not None:
                    state["current_stage"] = "auto_link"
                _auto_link_followup(con, case_id, case_data)
            _link_problem_notice(con, case_id, case_data)
            # ── Один кейс на письмо (multi-case split ОТКЛЮЧЁН для скорости/стабильности) ──
            # Позиции таблицы доступны в export.items для 1С; multi_item_count в payload —
            # quality_gate его учитывает и не флагает ложно. Чистим старые сиблинги.
            con.execute("DELETE FROM cases WHERE raw_email_id=? AND item_index>=1", (int(row["id"]),))
            count += 1
            if state is not None:
                state["processed"] = count
                state["current_stage"] = "commit" if count % 25 == 0 else "next"
            if count % 25 == 0:
                con.commit()
        except Exception as exc:
            try:
                record_process_event(
                    con,
                    stage="classifier",
                    level="warn",
                    message=f"Паттерны: ошибка письма {row['id']}: {exc}",
                    raw_email_id=int(row["id"]),
                )
            except Exception:
                pass
            if state is not None:
                state["current_stage"] = f"error:{exc}"
    con.commit()
    if state is not None:
        state["current_stage"] = "done"
    return count


def _patterns_worker(only_missing: bool) -> None:
    """Фоновая классификация писем паттернами. only_missing=True — только raw без кейса."""
    _PATTERNS_STATE.update({"running": True, "total": 0, "processed": 0, "started_at": utcnow(), "finished_at": None, "error": None})
    _PATTERNS_STOP.clear()
    try:
        # WAL + busy_timeout=60000 уже сериализуют записи без глобального замка,
        # поэтому ручная кнопка «Паттерны» работает сразу, не дожидаясь других задач.
        apply_runtime_settings()
        # Полный перепрогон (not only_missing) перезаписывает все кейсы → старый лог качества
        # неактуален и копится (был 43МБ → O(N²) в сводке). Обрезаем перед прогоном.
        if not only_missing:
            _data_dir = Path(settings.database_path).parent
            for _qf in ("quality_checks.jsonl", "quality_errors.jsonl", "review_queue.jsonl"):
                try:
                    (_data_dir / _qf).unlink()
                except FileNotFoundError:
                    pass
                except Exception:
                    pass
        with connect() as con:
            count = _classify_pending(
                con,
                only_missing=only_missing,
                stop_event=_PATTERNS_STOP,
                state=_PATTERNS_STATE,
                manual_review_gate=True,
            )
        # Полную сводку качества считаем ОДИН раз после батча (в per-case пути она троттлится).
        try:
            from .quality_gate import flush_quality_summary
            flush_quality_summary()
        except Exception:
            pass
        _log("classifier", "Пересчёт завершён", level="ok", details={"reprocessed": count})
    except Exception as exc:
        _PATTERNS_STATE["error"] = str(exc)
    finally:
        _PATTERNS_STATE["running"] = False
        _PATTERNS_STATE["finished_at"] = utcnow()


@app.post("/api/reprocess")
def reprocess(only_missing: bool = True) -> dict[str, Any]:
    """Классификация паттернами в ФОНЕ. only_missing=True — только новые (raw без кейса).

    Прогресс — через /api/patterns/progress.
    """
    if _PATTERNS_STATE.get("running"):
        return {"ok": True, "already_running": True, "state": dict(_PATTERNS_STATE)}
    th = threading.Thread(target=_patterns_worker, args=(only_missing,), name="readmail-patterns", daemon=True)
    th.start()
    return {"ok": True, "started": True, "background": True}


@app.get("/api/patterns/progress")
def api_patterns_progress() -> dict[str, Any]:
    return {"ok": True, "state": dict(_PATTERNS_STATE)}


@app.post("/api/patterns/stop")
def api_patterns_stop() -> dict[str, Any]:
    """Остановка обработки паттернами."""
    _PATTERNS_STOP.set()
    return {"ok": True, "message": "Остановка паттернов..."}


# Удалён: старый v1 endpoint /api/cases (дублировал api_cases_v2 с другим форматом ответа)


@app.get("/api/cases/search")
def api_cases_search(
    q: str = Query(..., min_length=1),
    by: str = Query(default="auto"),
    limit: int = Query(default=50, ge=1, le=200),
) -> dict[str, Any]:
    """Унифицированный поиск/трассировка кейса по case_id / raw_email_id / outbox_id /
    document_number / part_number / message_id / export_id. Read-only."""
    with connect() as con:
        result = search_cases(con, q, by=by, limit=limit)
    result["read_only"] = True
    return result


@app.get("/api/search")
def api_global_search(
    q: str = Query(default=""),
    scope: str = Query(default="all"),
    limit: int = Query(default=20, ge=1, le=100),
) -> dict[str, Any]:
    """Единый поиск по письмам/кейсам/outbox/клиентам. Типизированные результаты с
    подсказкой open_tab/open_params. Read-only, без секретов. Пустой q не падает."""
    if scope not in ("all", "emails", "cases", "outbox", "clients"):
        scope = "all"
    with connect() as con:
        result = search_mod.unified_search(con, q, scope=scope, limit=limit)
    result["read_only"] = True
    return result


@app.get("/api/search/trace/{entity_type}/{entity_id}")
def api_search_trace(entity_type: str, entity_id: int) -> dict[str, Any]:
    """Трассировка цепочки: письмо → кейс(ы) → outbox(ы) → попытки доставки. Read-only."""
    with connect() as con:
        result = search_mod.trace(con, entity_type, entity_id)
    result["read_only"] = True
    return result


@app.get("/api/cases/{case_id}")
def get_case(case_id: int) -> dict[str, Any]:
    with connect() as con:
        row = con.execute(
            """
            SELECT c.*, e.mailbox, e.uid, e.message_id, e.in_reply_to, e.references_json, e.subject, e.from_addr,
                   e.to_addr, e.cc_addr, e.direction, e.folder_seen_json, e.received_at, e.body_text, e.body_html, e.visible_text, e.snippet
            FROM cases c JOIN raw_emails e ON e.id = c.raw_email_id WHERE c.id=?
            """,
            (case_id,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Case not found")
        data = row_to_dict(row) or {}
        data["evidence_gate"] = (data.get("payload") or {}).get("evidence_gate") or {}
        data["attachments"] = [dict(a) for a in con.execute("SELECT filename, content_type, size_bytes FROM attachments WHERE raw_email_id=?", (data["raw_email_id"],))]
        if data.get("thread_key"):
            data["thread"] = [
                row_to_dict(r)
                for r in con.execute(
                    """
                    SELECT c.id, c.state, c.event_type, c.claim_kind, c.priority, e.subject, e.from_addr, e.received_at, e.snippet
                    FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
                    WHERE c.thread_key=? AND c.id<>?
                    ORDER BY e.received_at ASC, c.id ASC
                    LIMIT 100
                    """,
                    (data["thread_key"], case_id),
                )
            ]
    return data


@app.patch("/api/cases/{case_id}")
def update_case(case_id: int, upd: ManualUpdate) -> dict[str, Any]:
    with connect() as con:
        row = con.execute(
            """SELECT c.*, e.subject, e.from_addr, e.received_at, e.body_text, e.snippet
               FROM cases c LEFT JOIN raw_emails e ON e.id=c.raw_email_id
               WHERE c.id=?""",
            (case_id,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Case not found")
        current = row_to_dict(row) or {}
        fields = dict(current.get("fields") or {})
        if upd.fields:
            fields.update({k: v for k, v in upd.fields.items() if v not in (None, "")})
        state = upd.state if upd.state is not None else current.get("state")
        ready = upd.ready_for_export if upd.ready_for_export is not None else bool(current.get("ready_for_export"))
        needs_review = upd.needs_review if upd.needs_review is not None else not ready
        case_data = {
            **current,
            "buyer_code": upd.buyer_code if upd.buyer_code is not None else current.get("buyer_code"),
            "buyer_name": upd.buyer_name if upd.buyer_name is not None else current.get("buyer_name"),
            "event_type": upd.event_type if upd.event_type is not None else current.get("event_type"),
            "claim_kind": upd.claim_kind if upd.claim_kind is not None else current.get("claim_kind"),
            "status": upd.status if upd.status is not None else current.get("status"),
            "priority": upd.priority if upd.priority is not None else current.get("priority"),
            "deadline_at": upd.deadline_at if upd.deadline_at is not None else current.get("deadline_at"),
            "state": state,
            "ready_for_export": bool(ready),
            "needs_review": bool(needs_review),
            "fields": fields,
        }
        case_data["export"] = _case_export_from_row({**case_data, "id": case_id}, fields)
        save_case(
            con,
            int(current["raw_email_id"]),
            case_data,
            item_index=int(current.get("item_index") or 0),
        )
        updated_case = con.execute(
            "SELECT state, ready_for_export FROM cases WHERE id=?",
            (case_id,),
        ).fetchone()
        if updated_case and updated_case["state"] == "ready_to_1c" and int(updated_case["ready_for_export"] or 0):
            queue_case_event(con, case_id)
    return get_case(case_id)


@app.post("/api/cases/{case_id}/confirm")
def confirm_case(case_id: int, upd: ManualUpdate | None = None) -> dict[str, Any]:
    """Confirm operator-corrected fields and teach buyer identity from From-address/domain.

    This is deliberately separate from plain PATCH: only confirmation writes trusted learning rows.
    """
    if upd is not None:
        update_case(case_id, upd)
    with connect() as con:
        result = learn_buyer_from_case(con, case_id, source="human_confirmed")
        record_learning_event(
            con,
            kind="case_confirmed",
            source="human_confirmed",
            case_id=case_id,
            confidence=1.0,
            payload={"note": "operator confirmed case fields; eligible for learning", "buyer_learning": result},
        )
        if getattr(settings, "auto_queue_control_events", True):
            queue_case_event(con, case_id, event_type="case_confirmed")
    data = get_case(case_id)
    data["learning"] = result
    return data


@app.post("/api/cases/{case_id}/mark_ready")
def mark_ready(case_id: int) -> dict[str, Any]:
    with connect() as con:
        row = con.execute("SELECT * FROM cases WHERE id=?", (case_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Case not found")
        data = row_to_dict(row) or {}
        fields = data.get("fields") or {}
        payload = data.get("payload") or {}
        direction = payload.get("direction") or "unknown"
        strong_key = data.get("strong_key") or make_strong_key(data.get("buyer_code"), fields)
        missing, quality = quality_check(
            data.get("event_type") or "unknown",
            data.get("claim_kind"),
            fields,
            strong_key,
            direction,
            data.get("buyer_code"),
            (payload.get("evidence") or {}),
        )
        hard_errors = [q for q in quality if q.get("level") == "error"]
        if data.get("event_type") != "new_return" or direction != "inbound_customer" or missing or hard_errors:
            raise HTTPException(
                400,
                {
                    "error": "case_not_ready_for_export",
                    "event_type": data.get("event_type"),
                    "direction": direction,
                    "missing": missing,
                    "quality": quality,
                },
            )
    return update_case(case_id, ManualUpdate(state="ready_to_1c", ready_for_export=True, needs_review=False))


@app.post("/api/cases/{case_id}/close")
def close_case(case_id: int) -> dict[str, Any]:
    return update_case(case_id, ManualUpdate(state="closed", ready_for_export=False, needs_review=False))


@app.get("/api/export/ready")
def export_ready(limit: int = Query(default=500, ge=1, le=5000)) -> dict[str, Any]:
    with connect() as con:
        rows = con.execute(
            "SELECT id, export_json FROM cases WHERE ready_for_export=1 AND state='ready_to_1c' ORDER BY id LIMIT ?",
            (limit,),
        ).fetchall()
    items = []
    for r in rows:
        payload = loads(r["export_json"], {}) or {}
        payload["case_id"] = r["id"]
        items.append(payload)
    return {"schema_version": "readmail-new-export-v1", "count": len(items), "items": items}


@app.post("/api/export/{case_id}/queue")
def queue_export(case_id: int) -> dict[str, Any]:
    with connect() as con:
        result = queue_case_export(con, case_id)
    if not result.get("ok"):
        raise HTTPException(400, result)
    return result


@app.post("/api/export/queue-ready")
def queue_ready_export(limit: int = Query(default=500, ge=1, le=5000)) -> dict[str, Any]:
    with connect() as con:
        return queue_ready_cases(con, limit=limit)


@app.post("/api/export/queue-control")
def queue_control_export(limit: int = Query(default=1000, ge=1, le=10000)) -> dict[str, Any]:
    apply_runtime_settings()
    with connect() as con:
        return queue_control_events(con, limit=limit)


@app.post("/api/export/deliver")
def deliver_export(channel: str | None = None, limit: int = Query(default=200, ge=1, le=2000)) -> dict[str, Any]:
    apply_runtime_settings()
    with connect() as con:
        return deliver_outbox_events(con, limit=limit, channel=channel)


@app.get("/api/config/buyers")
def buyers() -> dict[str, Any]:
    rules = load_buyer_rules()
    return {"count": len(rules), "items": [{"code": r.code, "name": r.name, "domains": r.domains, "senders": r.senders[:10]} for r in rules]}


@app.get("/api/learning/buyers")
def learned_buyers() -> dict[str, Any]:
    with connect() as con:
        rows = [dict(r) for r in con.execute(
            """
            SELECT identity_type, identity_value, buyer_code, buyer_name, source, confidence,
                   seen_count, confirmed_count, rejected_count, updated_at
            FROM buyer_identities
            ORDER BY confirmed_count DESC, seen_count DESC, identity_value
            """
        ).fetchall()]
    return {"count": len(rows), "items": rows}


@app.get("/api/learning/events")
def learning_events(limit: int = Query(default=100, ge=1, le=500)) -> dict[str, Any]:
    with connect() as con:
        rows = [row_to_dict(r) for r in con.execute(
            """
            SELECT id, case_id, raw_email_id, kind, source, confidence, payload_json, created_at
            FROM learning_events
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()]
    return {"count": len(rows), "items": rows}


# v2.1 AI-only: эндпоинты обучения паттернов удалены (learning/dashboard|profiles|pattern-candidates|autolearn).


@app.post("/api/cases/{case_id}/ai_suggest")
def api_case_ai_suggest(case_id: int) -> dict[str, Any]:
    _log("ai", "Ручная AI-подсказка: старт", case_id=case_id)
    with connect() as con:
        row = con.execute(
            """
            SELECT c.*, e.mailbox, e.uid, e.message_id, e.in_reply_to, e.references_json, e.subject, e.from_addr,
                   e.to_addr, e.cc_addr, e.direction, e.folder_seen_json, e.received_at, e.body_text, e.body_html, e.visible_text, e.snippet
            FROM cases c JOIN raw_emails e ON e.id = c.raw_email_id WHERE c.id=?
            """,
            (case_id,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Case not found")
        data = row_to_dict(row) or {}
        email_data = {
            "mailbox": data.get("mailbox"),
            "uid": data.get("uid"),
            "message_id": data.get("message_id"),
            "subject": data.get("subject"),
            "from_addr": data.get("from_addr"),
            "to_addr": data.get("to_addr"),
            "cc_addr": data.get("cc_addr"),
            "received_at": data.get("received_at"),
            "body_text": data.get("body_text"),
            "body_html": data.get("body_html"),
            "snippet": data.get("snippet"),
        }
        case_data = {
            "buyer_code": data.get("buyer_code"),
            "buyer_name": data.get("buyer_name"),
            "event_type": data.get("event_type"),
            "claim_kind": data.get("claim_kind"),
            "fields": data.get("fields") or {},
            "missing": data.get("missing") or [],
        }
        suggestion = run_ai_suggestion(email_data, case_data, con=con, case_id=case_id, purpose="manual_ai_suggest")
        con.execute(
            "INSERT INTO ai_suggestions(case_id, model, prompt_hash, response_json, accepted, created_at) VALUES (?, ?, ?, ?, 0, ?)",
            (case_id, suggestion.get("model"), suggestion.get("prompt_hash"), dumps(suggestion), utcnow()),
        )
    _log("ai", "Ручная AI-подсказка: ответ сохранён", level="ok" if suggestion.get("ok") else "warn", case_id=case_id, details={"ok": suggestion.get("ok"), "model": suggestion.get("model"), "error": suggestion.get("error")})
    return suggestion


@app.post("/api/cases/{case_id}/ai_apply")
def api_case_ai_apply(case_id: int) -> dict[str, Any]:
    learning_result: dict[str, Any] = {"skipped": "not_started"}
    with connect() as con:
        row = con.execute(
            """
            SELECT c.*, e.mailbox, e.uid, e.message_id, e.in_reply_to, e.references_json, e.subject, e.from_addr,
                   e.to_addr, e.cc_addr, e.direction, e.folder_seen_json, e.received_at, e.body_text, e.body_html, e.snippet, e.quote_markers
            FROM cases c JOIN raw_emails e ON e.id = c.raw_email_id WHERE c.id=?
            """,
            (case_id,),
        ).fetchone()
        if not row:
            raise HTTPException(404, "Case not found")
        data = row_to_dict(row) or {}
        email_data = {
            "mailbox": data.get("mailbox"), "uid": data.get("uid"), "message_id": data.get("message_id"),
            "in_reply_to": data.get("in_reply_to"), "references": data.get("references") or [],
            "subject": data.get("subject"), "from_addr": data.get("from_addr"), "to_addr": data.get("to_addr"),
            "cc_addr": data.get("cc_addr"), "received_at": data.get("received_at"),
            "body_text": data.get("body_text"), "body_html": data.get("body_html"), "snippet": data.get("snippet"),
            "quote_markers": data.get("quote_markers"),
        }
        case_data = {
            "buyer_code": data.get("buyer_code"), "buyer_name": data.get("buyer_name"),
            "event_type": data.get("event_type"), "claim_kind": data.get("claim_kind"), "status": data.get("status"),
            "priority": data.get("priority"), "confidence": data.get("confidence"), "deadline_at": data.get("deadline_at"),
            "thread_key": data.get("thread_key"), "strong_key": data.get("strong_key"), "weak_key": data.get("weak_key"),
            "is_followup": bool(data.get("is_followup")), "ready_for_export": bool(data.get("ready_for_export")),
            "needs_review": bool(data.get("needs_review")), "state": data.get("state"),
            "fields": data.get("fields") or {}, "missing": data.get("missing") or [], "quality": data.get("quality") or [],
            "payload": data.get("payload") or {}, "export": data.get("export") or {},
        }
        suggestion = run_ai_suggestion(email_data, case_data, con=con, case_id=case_id, purpose="manual_ai_apply")
        con.execute(
            "INSERT INTO ai_suggestions(case_id, model, prompt_hash, response_json, accepted, created_at) VALUES (?, ?, ?, ?, 0, ?)",
            (case_id, suggestion.get("model"), suggestion.get("prompt_hash"), dumps(suggestion), utcnow()),
        )
        if not suggestion.get("ok") or not isinstance(suggestion.get("response"), dict):
            return {"ok": False, "suggestion": suggestion, "applied": False}
        updated = force_operator_review(apply_ai_overlay(email_data, case_data, suggestion["response"]))
        updated_id = save_case(con, int(data["raw_email_id"]), updated)
        updated["export"]["case_id"] = updated_id
        con.execute(
            "UPDATE cases SET export_json=?, updated_at=? WHERE id=?",
            (dumps(updated.get("export") or {}), utcnow(), updated_id),
        )
        con.execute("UPDATE ai_suggestions SET accepted=1 WHERE case_id=? AND prompt_hash=?", (case_id, suggestion.get("prompt_hash")))
        _link_problem_notice(con, updated_id, updated)
        con.commit()
    return {"ok": True, "applied": True, "suggestion": suggestion, "learning": learning_result, "case": get_case(case_id)}


class RouteCaseReq(BaseModel):
    destination: str | None = None   # куда переместить (опц.)
    claim_kind: str | None = None    # сменить причину/статус (опц.)
    reason: str | None = None        # комментарий оператора


_ROUTE_KINDS = {"defect", "nonconforming", "number_replacement", "wrong_item", "shortage",
                "overdelivery", "incomplete_set", "correction_request", "marking_request", "quality_refusal"}


@app.post("/api/cases/{case_id}/route")
def api_case_route(case_id: int, req: RouteCaseReq) -> dict[str, Any]:
    """Ручной перенос кейса И/ИЛИ смена причины (claim_kind) оператором со статусом/причиной."""
    DEST = {
        "ready_1c":       {"state": "ready_to_1c", "ready_for_export": 1, "needs_review": 0},
        "junk":           {"state": "ignored_info_only", "event_type": "info_only", "ready_for_export": 0, "needs_review": 0},
        "needs_link":     {"state": "needs_link", "ready_for_export": 0, "needs_review": 0},
        "problem_notice": {"state": "problem_notice", "event_type": "problem_notice", "ready_for_export": 0, "needs_review": 0},
        "manual":         {"state": "needs_review", "ready_for_export": 0, "needs_review": 1},
    }
    spec = DEST.get(req.destination) if req.destination else None
    if req.destination and not spec:
        return {"ok": False, "error": f"неизвестное назначение: {req.destination}"}
    new_kind = (req.claim_kind or "").strip() or None
    if new_kind and new_kind not in _ROUTE_KINDS:
        return {"ok": False, "error": f"неизвестная причина: {new_kind}"}
    if not spec and not new_kind:
        return {"ok": False, "error": "не задано ни назначение, ни причина"}
    reason = (req.reason or "").strip() or None
    with connect() as con:
        row = con.execute("SELECT id, payload_json FROM cases WHERE id=?", (case_id,)).fetchone()
        if not row:
            return {"ok": False, "error": "кейс не найден"}
        try:
            payload = json.loads(row["payload_json"] or "{}")
        except Exception:
            payload = {}
        payload["operator_route"] = {"to": req.destination, "claim_kind": new_kind, "reason": reason, "at": utcnow()}
        payload["processing_mode"] = "manual"
        sets = ["payload_json=?", "updated_at=?"]
        vals: list[Any] = [json.dumps(payload, ensure_ascii=False), utcnow()]
        cols: dict[str, Any] = dict(spec or {})
        if new_kind:
            cols["claim_kind"] = new_kind
            cols["status"] = new_kind
        for col, v in cols.items():
            sets.append(f"{col}=?")
            vals.append(v)
        vals.append(case_id)
        con.execute(f"UPDATE cases SET {', '.join(sets)} WHERE id=?", vals)
        if req.destination == "ready_1c":
            try:
                queue_case_event(con, case_id)
            except Exception:
                pass
        _parts = []
        if req.destination:
            _parts.append(f"→ {req.destination}")
        if new_kind:
            _parts.append(f"статус={new_kind}")
        record_process_event(con, stage="operator", level="info",
            message=f"Оператор #{case_id}: " + ", ".join(_parts) + (f" ({reason})" if reason else ""),
            case_id=case_id)
        con.commit()
    return {"ok": True, "case_id": case_id, "destination": req.destination, "claim_kind": new_kind,
            "state": (spec or {}).get("state")}


@app.post("/api/ai/test")
def api_ai_test() -> dict[str, Any]:
    apply_runtime_settings()
    _log("ai", "AI test: старт", details={"provider": settings.ai_provider, "model": settings.ai_model, "base_url": settings.ai_base_url})
    result = test_ai_connection()
    _log("ai", "AI test: завершён", level="ok" if result.get("ok") else "error", details={"ok": result.get("ok"), "error": result.get("error"), "provider": result.get("provider"), "model": result.get("model")})
    return result


class AiTestText(BaseModel):
    prompt: str
    system: str | None = None


@app.post("/api/ai/test-text")
def api_ai_test_text(req: AiTestText) -> dict[str, Any]:
    """Мини-чат для проверки ТЕКСТОВОЙ модели: вопрос → ответ (свободный текст)."""
    from .ai_client import _request_chat, _response_content, _usage_tokens, trace_freeform_ai_call
    apply_runtime_settings()
    if not (req.prompt or "").strip():
        return {"ok": False, "error": "Пустой запрос"}
    messages = []
    if (req.system or "").strip():
        messages.append({"role": "system", "content": req.system.strip()})
    messages.append({"role": "user", "content": req.prompt.strip()})
    # Свободный чат: временно снимаем форс JSON (иначе модель отвечает «[ ]» вместо текста).
    _orig_fmt = settings.ai_response_format
    try:
        settings.ai_response_format = "text"  # type: ignore[attr-defined]
        raw, provider, model, _url = _request_chat(messages)
        response = _response_content(raw)
        ptok, ctok = _usage_tokens(raw, len(req.prompt), len(response or ""))
        trace_freeform_ai_call(
            req.prompt,
            provider=provider,
            model=model,
            response_text=response,
            usage={"prompt_tokens": ptok, "completion_tokens": ctok},
        )
        return {"ok": True, "provider": provider, "model": model, "response": response}
    except Exception as exc:
        trace_freeform_ai_call(
            req.prompt,
            provider=settings.ai_provider,
            model=settings.ai_model,
            error=str(exc),
        )
        return {"ok": False, "error": str(exc), "model": settings.ai_model, "provider": settings.ai_provider}
    finally:
        settings.ai_response_format = _orig_fmt  # type: ignore[attr-defined]


@app.post("/api/ai/test-vision")
async def api_ai_test_vision(prompt: str = Form(""), file: UploadFile = File(...)) -> dict[str, Any]:
    """Проверка ВИЗУАЛЬНОЙ модели: загруженный файл + вопрос → ответ модели."""
    from .ai_client import run_vision_extraction
    apply_runtime_settings()
    try:
        data = await file.read()
    except Exception as exc:
        return {"ok": False, "error": f"Не прочитать файл: {exc}"}
    if not data:
        return {"ok": False, "error": "Пустой файл"}
    ct = file.content_type or "image/jpeg"
    res = run_vision_extraction(data, content_type=ct, prompt_text=((prompt or "").strip() or None))
    if res.get("skipped"):
        skip = res["skipped"]
        msg = ("Vision выключен — включите «Vision AI» (AI_VISION_ENABLED) в настройках."
               if skip == "ai_vision_enabled=false" else f"Vision недоступен: {skip}")
        return {"ok": False, "error": msg, "skipped": skip, "model": settings.ai_vision_model}
    if res.get("error"):
        return {"ok": False, "error": res["error"], "model": settings.ai_vision_model}
    # Для свободного вопроса возвращаем сырой текст ответа (raw_excerpt), а не распарсенный JSON.
    answer = res.get("raw_excerpt") or res.get("response") or ""
    return {"ok": True, "provider": res.get("provider"), "model": res.get("model"),
            "response": answer, "parsed": res.get("response"), "usage": res.get("usage")}


@app.post("/api/ai/full-batch")
def api_ai_full_batch(limit: int = Query(default=10, ge=1, le=100)) -> dict[str, Any]:
    """Ручной полный ИИ-прогон (режим обучения) по N свежим возвратам — в ФОНЕ.
    Прогресс — /api/ai/batch-progress, живой лог — /api/ai/live-log, стоп — /api/ai/stop-batch."""
    apply_runtime_settings()
    if not settings.enable_ai:
        return {"ok": False, "error": "AI выключен в настройках (ENABLE_AI=false)"}
    if _AI_BATCH_STATE.get("running"):
        return {"ok": True, "already_running": True, "state": dict(_AI_BATCH_STATE)}
    _AI_BATCH_STOP.clear()
    th = threading.Thread(target=_ai_full_batch_worker, args=(int(limit),),
                          name="readmail-ai-full-batch", daemon=True)
    th.start()
    return {"ok": True, "started": True, "background": True, "limit": int(limit)}


@app.post("/api/ai/probe")
def api_ai_probe() -> dict[str, Any]:
    apply_runtime_settings()
    return probe_ai_server()


@app.get("/api/ai/models")
def api_ai_models() -> dict[str, Any]:
    apply_runtime_settings()
    return list_ai_models()


@app.post("/api/ai/models")
def api_ai_models_post() -> dict[str, Any]:
    apply_runtime_settings()
    return list_ai_models()


@app.get("/api/ai/config")
def api_ai_config() -> dict[str, Any]:
    apply_runtime_settings()
    return {
        "enable_ai": settings.enable_ai,
        "provider": settings.ai_provider,
        "base_url": settings.ai_base_url,
        "model": settings.ai_model,
        "response_format": settings.ai_response_format,
        "endpoint_mode": getattr(settings, "ai_endpoint_mode", "auto"),
        "endpoint_path": getattr(settings, "ai_endpoint_path", ""),
        "cache_enabled": settings.ai_cache_enabled,
        "conserve_tokens": settings.ai_conserve_tokens,
        "max_chars": settings.ai_max_chars,
        "max_output_tokens": settings.ai_max_output_tokens,
        "auto_ai_first_unknown_customer": settings.auto_ai_first_unknown_customer,
        "auto_apply_ai_on_first_unknown_customer": settings.auto_apply_ai_on_first_unknown_customer,
    }


@app.get("/api/ai/usage")
def api_ai_usage(limit: int = Query(default=100, ge=1, le=1000)) -> dict[str, Any]:
    with connect() as con:
        rows = [
            row_to_dict(r)
            for r in con.execute(
                """
                SELECT id, case_id, provider, model, prompt_hash, prompt_chars, response_chars, cached, ok, error, created_at
                FROM ai_usage ORDER BY id DESC LIMIT ?
                """,
                (limit,),
            ).fetchall()
        ]
        total = con.execute("SELECT COUNT(*) c FROM ai_usage").fetchone()["c"]
        cached = con.execute("SELECT COUNT(*) c FROM ai_usage WHERE cached=1").fetchone()["c"]
        chars = con.execute("SELECT COALESCE(SUM(prompt_chars),0) p, COALESCE(SUM(response_chars),0) r FROM ai_usage").fetchone()
    return {"total": total, "cached": cached, "prompt_chars": chars["p"], "response_chars": chars["r"], "items": rows}


@app.get("/api/export/outbox")
def export_outbox(limit: int = Query(default=100, ge=1, le=1000), status: str | None = None,
                  page: int = Query(default=1, ge=1)) -> dict[str, Any]:
    filters = []
    params: list[Any] = []
    if status and status != "all":
        filters.append("status=?")
        params.append(status)
    where = ("WHERE " + " AND ".join(filters)) if filters else ""
    offset = (page - 1) * limit
    with connect() as con:
        total = con.execute(f"SELECT COUNT(*) c FROM outbox {where}", params).fetchone()["c"]
        rows = [row_to_dict(r) for r in con.execute(
            f"""
            SELECT id, case_id, payload_json, status, event_type, channel, business_priority,
                   attempt_count, last_attempt_at, next_attempt_at, last_error, file_path,
                   delivery_response_json, created_at, sent_at, resolved_at, resolution_note
            FROM outbox
            {where}
            ORDER BY id DESC
            LIMIT ? OFFSET ?
            """,
            params + [limit, offset],
        ).fetchall()]
    pages = max(1, (total + limit - 1) // limit)
    return {"total": total, "count": len(rows), "items": rows,
            "page": page, "page_size": limit, "pages": pages,
            "has_more": offset + len(rows) < total}


@app.get("/api/export/journal")
def export_journal(limit: int = Query(default=100, ge=1, le=1000)) -> dict[str, Any]:
    with connect() as con:
        return outbox_dashboard(con, limit=limit)


@app.post("/api/export/retry-errors")
def export_retry_errors(limit: int = Query(default=500, ge=1, le=5000)) -> dict[str, Any]:
    with connect() as con:
        reset = reset_outbox_errors(con, limit=limit)
        delivery = deliver_outbox_events(con, limit=limit)
        return {"ok": True, "reset": reset, "delivery": delivery}


@app.post("/api/export/reconcile")
def export_reconcile(limit: int = Query(default=1000, ge=1, le=10000)) -> dict[str, Any]:
    apply_runtime_settings()
    with connect() as con:
        return reconcile_outbox_events(con, limit=limit)



@app.get("/api/project/analytics")
def project_analytics(live: bool = Query(default=False)) -> dict[str, Any]:
    """Full local project analytics for release/debug.

    This does not export secrets.  It checks source size, database coverage, outbox,
    AI configuration and common business safety invariants.
    """
    apply_runtime_settings()
    init_db()
    root = BASE_DIR.parent
    source_files = []
    for path in root.rglob("*"):
        if path.is_file() and path.suffix in {".py", ".js", ".html", ".css", ".yml", ".yaml", ".md"}:
            try:
                rel = str(path.relative_to(root))
                if "__pycache__" in rel or "/.venv/" in rel:
                    continue
                text = path.read_text(errors="ignore")
                source_files.append({"path": rel, "lines": text.count("\n") + 1, "bytes": path.stat().st_size})
            except Exception:
                pass
    largest = sorted(source_files, key=lambda x: x["lines"], reverse=True)[:12]
    issues: list[dict[str, Any]] = []
    with connect() as con:
        counts = {
            "raw_emails": con.execute("SELECT COUNT(*) c FROM raw_emails").fetchone()["c"],
            "cases": con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"],
            "outbox": con.execute("SELECT COUNT(*) c FROM outbox").fetchone()["c"],
            "ai_usage": con.execute("SELECT COUNT(*) c FROM ai_usage").fetchone()["c"],
        }
        raw_without_case = con.execute("SELECT COUNT(*) c FROM raw_emails r LEFT JOIN cases c ON c.raw_email_id=r.id WHERE c.id IS NULL").fetchone()["c"]
        case_without_event = con.execute("""
            SELECT COUNT(*) c FROM cases c
            WHERE c.state NOT IN ('ignored_internal','closed')
              AND NOT EXISTS (SELECT 1 FROM outbox o WHERE o.case_id=c.id)
        """).fetchone()["c"]
        ready_bad = con.execute("""
            SELECT COUNT(*) c FROM cases
            WHERE ready_for_export=1 AND (needs_review=1 OR state!='ready_to_1c' OR event_type!='new_return')
        """).fetchone()["c"]
        outbox_errors = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='error'").fetchone()["c"]
        outbox_new = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='new'").fetchone()["c"]
        suspicious = [row_to_dict(r) for r in con.execute("""
            SELECT c.id, c.state, c.event_type, c.claim_kind, c.fields_json, e.subject, e.from_addr
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE lower(c.fields_json) GLOB '*сумма*' OR lower(c.fields_json) GLOB '*цена*' OR lower(c.fields_json) GLOB '*причина*'
            LIMIT 20
        """).fetchall()]
        if raw_without_case:
            issues.append({"severity":"error","code":"raw_without_case","message":f"Писем без кейса: {raw_without_case}. Нужен пересчёт."})
        if case_without_event:
            issues.append({"severity":"warn","code":"case_without_outbox_event","message":f"Кейсов без события контроля: {case_without_event}. Нажми 'Починить события контроля'."})
        if ready_bad:
            issues.append({"severity":"error","code":"unsafe_ready","message":f"Небезопасных ready-кейсов: {ready_bad}. Нужна диагностика качества."})
        if outbox_errors:
            issues.append({"severity":"error","code":"outbox_errors","message":f"Ошибок доставки outbox: {outbox_errors}."})
        if settings.enable_ai and not settings.ai_provider:
            issues.append({"severity":"error","code":"ai_provider_empty","message":"AI включён, но провайдер пустой."})
        if settings.enable_ai and settings.ai_provider == "openai_compatible" and "127.0.0.1" in settings.ai_base_url:
            issues.append({"severity":"warn","code":"docker_localhost","message":"В Docker 127.0.0.1 указывает на контейнер. Используй host.docker.internal."})
        if settings.enable_ai and settings.ai_provider == "gigachat" and not (settings.gigachat_auth_key or (settings.ai_api_key and settings.ai_api_key not in {'local','none',''})):
            issues.append({"severity":"warn","code":"gigachat_no_key","message":"GigaChat включён, но не задан Authorization key или access token."})
    ai_live = None
    if live and settings.enable_ai:
        ai_live = test_ai_connection()
    status = "ok"
    if any(i["severity"] == "error" for i in issues):
        status = "error"
    elif any(i["severity"] == "warn" for i in issues):
        status = "warn"
    return {
        "schema": "readmail-new-project-analytics-v1.14",
        "generated_at": utcnow(),
        "status": status,
        "version": settings.app_version,
        "source": {
            "files": len(source_files),
            "lines": sum(x["lines"] for x in source_files),
            "bytes": sum(x["bytes"] for x in source_files),
            "largest": largest,
        },
        "database": {
            **counts,
            "raw_without_case": raw_without_case,
            "cases_without_control_event": case_without_event,
            "ready_bad": ready_bad,
            "outbox_errors": outbox_errors,
            "outbox_new": outbox_new,
            "suspicious_field_samples": suspicious,
        },
        "ai": {
            "enabled": settings.enable_ai,
            "provider": settings.ai_provider,
            "base_url": settings.ai_base_url,
            "model": settings.ai_model,
            "endpoint_mode": getattr(settings, "ai_endpoint_mode", "auto"),
            "endpoint_path": getattr(settings, "ai_endpoint_path", ""),
            "gigachat_base_url": settings.gigachat_base_url,
            "gigachat_key_configured": bool(settings.gigachat_auth_key),
            "live_test": ai_live,
        },
        "settings_safety": {
            "folders": settings.folders,
            "configured_folders_are_customer": settings.configured_folders_are_customer,
            "strict_evidence_validation": settings.strict_evidence_validation,
            "one_c_export_mode": settings.one_c_export_mode,
            "auto_deliver_outbox": settings.auto_deliver_outbox,
            "scan_interval_seconds": settings.scan_interval_seconds,
        },
        "issues": issues,
        "recommendation": "Можно расширять тест" if status == "ok" else "Сначала исправить issues и повторить диагностику",
    }


# ── Экспорт метаданных писем для AI / бэкапа ──────────────────────────────

@app.get("/api/export/emails-metadata")
def export_emails_metadata(limit: int = Query(default=0, ge=0, le=100000)) -> dict[str, Any]:
    """Export all emails with attachment metadata (no actual attachments).
    Returns a structured JSON array suitable for feeding to an AI model."""
    apply_runtime_settings()
    items: list[dict[str, Any]] = []
    with connect() as con:
        sql = """
            SELECT e.id, e.mailbox, e.uid, e.direction, e.message_id, e.in_reply_to,
                   e.subject, e.from_addr, e.to_addr, e.cc_addr, e.received_at,
                   e.body_text, e.body_html, e.snippet, e.imported_at, e.updated_at
            FROM raw_emails e
            ORDER BY e.received_at DESC, e.id DESC
        """
        if limit > 0:
            sql += " LIMIT ?"
            rows = con.execute(sql, (limit,)).fetchall()
        else:
            rows = con.execute(sql).fetchall()

        for row in rows:
            d = dict(row)
            # Fetch attachments for this email
            att_rows = con.execute(
                "SELECT id, filename, content_type, size_bytes FROM attachments WHERE raw_email_id = ? ORDER BY id",
                (d["id"],),
            ).fetchall()
            d["attachments"] = [
                {
                    "id": a["id"],
                    "filename": a["filename"],
                    "content_type": a["content_type"],
                    "size_bytes": a["size_bytes"],
                }
                for a in att_rows
            ]
            d["has_attachments"] = len(d["attachments"]) > 0
            d["total_attachments_size"] = sum((a["size_bytes"] or 0) for a in d["attachments"])
            items.append(d)

    return {
        "ok": True,
        "schema": "readmail-emails-metadata-v1",
        "generated_at": utcnow(),
        "total_emails": len(items),
        "total_with_attachments": sum(1 for i in items if i["has_attachments"]),
        "emails": items,
    }


@app.get("/api/export/emails-metadata/download")
def download_emails_metadata(limit: int = Query(default=0, ge=0, le=100000)) -> FileResponse:
    """Download all email metadata as a pretty-printed JSON file."""
    data = export_emails_metadata(limit=limit)
    tmp = Path(f"/tmp/readmail_emails_metadata_{int(time.time())}.json")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return FileResponse(
        path=str(tmp),
        media_type="application/json",
        filename="readmail_emails_metadata.json",
    )


@app.get("/api/export/compare-json/download")
def download_compare_json() -> FileResponse:
    """Download paired JSONL files: original emails and processed cases with matching IDs."""
    apply_runtime_settings()
    export_root = Path(__file__).resolve().parent.parent / "exports"
    stamp = time.strftime("%Y%m%d_%H%M%S")
    outdir = export_root / f"mail_compare_{stamp}"
    outdir.mkdir(parents=True, exist_ok=True)

    original_path = outdir / "original_emails.jsonl"
    processed_path = outdir / "processed_cases.jsonl"
    pattern_path = outdir / "etalon_pattern.jsonl"
    full_trace_path = outdir / "etalon_pattern_ai_1c.jsonl"
    folders_report_path = outdir / "folders_report.json"
    paired_path = outdir / "paired_index.json"
    manifest_path = outdir / "manifest.json"
    zip_path = export_root / f"mail_compare_{stamp}.zip"

    def dump_line(fp: Any, obj: dict[str, Any]) -> None:
        fp.write(json.dumps(obj, ensure_ascii=False, sort_keys=True) + "\n")

    # Корзина письма по состоянию кейса — для группировки в отчёте.
    def _bucket(state: str | None) -> str:
        s = str(state or "")
        if s == "ready_to_1c":
            return "Сверка (готово)"
        if s == "needs_review":
            return "Сверка (на проверке)"
        if s in ("needs_link", "linked_event"):
            return "Диалоги"
        if s.startswith("ignored_"):
            return "Не по теме"
        return "Прочее"

    # Распарсить текст из вложения (Excel/PDF) для отчёта — сами файлы не кладём.
    def _attachment_text(file_path: str | None, filename: str, ctype: str) -> str:
        if not file_path:
            return ""
        try:
            p = Path(file_path)
            if not p.exists() or p.stat().st_size > 3 * 1024 * 1024:
                return ""
            data = p.read_bytes()
            low = (filename or "").lower()
            from .email_parser import _extract_xlsx_text, _extract_pdf_text
            if low.endswith((".xlsx", ".xls")) or "spreadsheet" in ctype or "excel" in ctype:
                return _extract_xlsx_text(data)[:6000]
            if low.endswith(".pdf") or "pdf" in ctype:
                return _extract_pdf_text(data)[:6000]
        except Exception:
            return ""
        return ""

    with connect() as con:
        attachments: dict[int, list[dict[str, Any]]] = {}
        for r in con.execute("SELECT raw_email_id, id, filename, content_type, size_bytes, file_path FROM attachments ORDER BY raw_email_id, id"):
            d = dict(r)
            d["extracted_text"] = _attachment_text(d.get("file_path"), d.get("filename") or "", str(d.get("content_type") or "").lower())
            attachments.setdefault(int(r["raw_email_id"]), []).append(d)

        outboxes: dict[int, list[dict[str, Any]]] = {}
        for r in con.execute("SELECT * FROM outbox ORDER BY case_id, id"):
            d = dict(r)
            d["payload_json"] = loads(d.get("payload_json"), {}) or {}
            d["delivery_response_json"] = loads(d.get("delivery_response_json"), {}) or {}
            outboxes.setdefault(int(r["case_id"]), []).append(d)

        ai_suggestions: dict[int, list[dict[str, Any]]] = {}
        for r in con.execute("SELECT id, case_id, model, prompt_hash, response_json, accepted, created_at FROM ai_suggestions ORDER BY case_id, id"):
            d = dict(r)
            d["response_json"] = loads(d.get("response_json"), {}) or {}
            ai_suggestions.setdefault(int(r["case_id"]), []).append(d)

        cases_by_raw: dict[int, list[dict[str, Any]]] = {}
        for r in con.execute("SELECT * FROM cases ORDER BY raw_email_id, id"):
            d = dict(r)
            d["fields_json"] = loads(d.get("fields_json"), {}) or {}
            d["missing_json"] = loads(d.get("missing_json"), []) or []
            d["quality_json"] = loads(d.get("quality_json"), []) or []
            d["payload_json"] = loads(d.get("payload_json"), {}) or {}
            d["export_json"] = loads(d.get("export_json"), {}) or {}
            d["outbox"] = outboxes.get(int(d["id"]), [])
            d["ai_suggestions"] = ai_suggestions.get(int(d["id"]), [])
            cases_by_raw.setdefault(int(d["raw_email_id"]), []).append(d)

        emails = con.execute("SELECT * FROM raw_emails ORDER BY id").fetchall()
        pairs: list[dict[str, Any]] = []
        folders_report: dict[str, dict[str, Any]] = {}
        original_count = 0
        processed_count = 0

        def _email_folders(email: dict[str, Any]) -> list[str]:
            seen = loads(email.get("folder_seen_json"), []) or []
            if not seen and email.get("mailbox"):
                seen = [email.get("mailbox")]
            return [str(x) for x in seen if x]

        def _record_folder(email: dict[str, Any], cases: list[dict[str, Any]]) -> None:
            target_folders = _email_folders(email) or ["(без папки)"]
            for folder_name in target_folders:
                item = folders_report.setdefault(folder_name, {
                    "folder": folder_name,
                    "raw_emails": 0,
                    "with_cases": 0,
                    "without_cases": 0,
                    "cases": 0,
                    "buckets": {},
                    "processing_sources": {"pattern": 0, "ai": 0, "unknown": 0},
                    "outbox": {"new": 0, "sent": 0, "error": 0, "other": 0},
                })
                item["raw_emails"] += 1
                if cases:
                    item["with_cases"] += 1
                else:
                    item["without_cases"] += 1
                for case in cases:
                    item["cases"] += 1
                    bucket = _bucket(case.get("state"))
                    item["buckets"][bucket] = item["buckets"].get(bucket, 0) + 1
                    payload = case.get("payload_json") or {}
                    source = payload.get("processing_source") or ("ai" if payload.get("ai_overlay") else "pattern")
                    if source not in item["processing_sources"]:
                        source = "unknown"
                    item["processing_sources"][source] += 1
                    for ob in case.get("outbox") or []:
                        status = str(ob.get("status") or "other")
                        if status not in item["outbox"]:
                            status = "other"
                        item["outbox"][status] += 1

        def _original_obj(email: dict[str, Any], raw_id: int, export_id: str) -> dict[str, Any]:
            return {
                "etalon_id": export_id,
                "export_id": export_id,
                "raw_email_id": raw_id,
                "kind": "original_email",
                "mailbox": email.get("mailbox"),
                "uid": email.get("uid"),
                "folder_seen": _email_folders(email),
                "canonical_key": email.get("canonical_key"),
                "duplicate_of_raw_email_id": email.get("duplicate_of_raw_email_id"),
                "direction": email.get("direction"),
                "message_id": email.get("message_id"),
                "in_reply_to": email.get("in_reply_to"),
                "references": loads(email.get("references_json"), []) or [],
                "subject": email.get("subject"),
                "from_addr": email.get("from_addr"),
                "to_addr": email.get("to_addr"),
                "cc_addr": email.get("cc_addr"),
                "received_at": email.get("received_at"),
                "snippet": email.get("snippet"),
                "visible_text": email.get("visible_text") or email.get("body_text") or email.get("snippet"),
                "body_text": email.get("body_text"),
                "body_html": email.get("body_html"),
                "raw_hash": email.get("raw_hash"),
                "raw_path": email.get("raw_path"),
                "quote_markers": email.get("quote_markers"),
                "archived_body": email.get("archived_body"),
                "raw_size": email.get("raw_size"),
                "status": email.get("status"),
                "imported_at": email.get("imported_at"),
                "updated_at": email.get("updated_at"),
                "attachments": attachments.get(raw_id, []),
            }

        def _accepted_ai_fields(case: dict[str, Any]) -> dict[str, Any]:
            fields: dict[str, Any] = {}
            for sug in case.get("ai_suggestions") or []:
                if not sug.get("accepted"):
                    continue
                response = sug.get("response_json") or {}
                if isinstance(response.get("response"), dict):
                    response = response.get("response") or {}
                ai_fields = response.get("fields") if isinstance(response.get("fields"), dict) else {}
                for k, v in ai_fields.items():
                    if v not in (None, "", [], {}):
                        fields[k] = v
            return fields

        def _norm_field_value(value: Any) -> str:
            return str(value or "").strip().lower().replace(" ", "").replace(",", ".")

        def _field_sources(case: dict[str, Any], slim: dict[str, Any]) -> dict[str, str]:
            payload = case.get("payload_json") or {}
            ai_fields = _accepted_ai_fields(case)
            sources: dict[str, str] = {}
            for key, value in slim.items():
                if key in ai_fields and _norm_field_value(ai_fields.get(key)) == _norm_field_value(value):
                    sources[key] = "ai"
                elif payload.get("ai_overlay") and key in ai_fields:
                    sources[key] = "ai_normalized"
                else:
                    sources[key] = "pattern"
            return sources

        def _case_slim(case: dict[str, Any], raw_id: int, export_id: str, idx: int = 1, total: int = 1) -> dict[str, Any]:
            f = case.get("fields_json") or {}
            payload = case.get("payload_json") or {}
            slim = {k: f.get(k) for k in (
                "document_number", "claim_number", "return_number", "client_request_number",
                "document_date", "part_number", "brand", "product_name", "quantity", "comment",
            ) if f.get(k)}
            source = payload.get("processing_source") or ("ai" if payload.get("ai_overlay") else "pattern")
            stage_export_id = export_id if total == 1 else f"{export_id}#{idx}"
            field_sources = _field_sources(case, slim)
            return {
                "etalon_id": export_id,
                "export_id": stage_export_id,
                "parent_export_id": export_id,
                "raw_email_id": raw_id,
                "kind": "processed_case",
                "has_case": True,
                "case_id": case.get("id"),
                "bucket": _bucket(case.get("state")),
                "processing": {
                    "source": source,
                    "mode": payload.get("processing_mode") or "auto",
                    "manual_gate": bool(payload.get("manual_gate")),
                    "ai_overlay": bool(payload.get("ai_overlay")),
                    "pattern_before_ai_available": not bool(payload.get("ai_overlay")),
                },
                "stage_ids": {
                    "original": export_id,
                    "pattern": f"{stage_export_id}:pattern",
                    "ai": f"{stage_export_id}:ai",
                    "final": f"{stage_export_id}:final",
                    "one_c": f"{stage_export_id}:1c",
                },
                "buyer_code": case.get("buyer_code"),
                "buyer_name": case.get("buyer_name"),
                "event_type": case.get("event_type"),
                "claim_kind": case.get("claim_kind"),
                "state": case.get("state"),
                "ready_for_export": case.get("ready_for_export"),
                "fields": slim,
                "field_sources": field_sources,
                "missing": case.get("missing_json") or [],
                "quality": case.get("quality_json") or [],
                "defect_documents": payload.get("defect_documents"),
                "attachments": [
                    {"filename": a.get("filename"), "content_type": a.get("content_type"),
                     "size_bytes": a.get("size_bytes"), "extracted_text": a.get("extracted_text") or ""}
                    for a in attachments.get(raw_id, [])
                ],
            }

        with (
            original_path.open("w", encoding="utf-8") as fo,
            processed_path.open("w", encoding="utf-8") as fp,
            pattern_path.open("w", encoding="utf-8") as fpat,
            full_trace_path.open("w", encoding="utf-8") as ffull,
        ):
            for row in emails:
                email = dict(row)
                raw_id = int(email["id"])
                export_id = f"ID{raw_id}"
                original = _original_obj(email, raw_id, export_id)
                dump_line(fo, original)
                original_count += 1

                cases = cases_by_raw.get(raw_id, [])
                _record_folder(email, cases)
                pairs.append({"export_id": export_id, "raw_email_id": raw_id, "case_ids": [c["id"] for c in cases]})
                if not cases:
                    stage_ids = {"original": export_id, "pattern": f"{export_id}:pattern", "ai": f"{export_id}:ai", "final": f"{export_id}:final", "one_c": f"{export_id}:1c"}
                    dump_line(fp, {"etalon_id": export_id, "export_id": export_id, "raw_email_id": raw_id, "kind": "processed_case", "has_case": False, "case_id": None, "stage_ids": stage_ids})
                    dump_line(fpat, {"etalon_id": export_id, "export_id": export_id, "raw_email_id": raw_id, "stage_ids": stage_ids, "original": original, "pattern_result": None})
                    dump_line(ffull, {"etalon_id": export_id, "export_id": export_id, "raw_email_id": raw_id, "stage_ids": stage_ids, "original": original, "pattern_result": None, "ai_suggestions": [], "one_c": []})
                    processed_count += 1
                    continue
                for idx, case in enumerate(cases, 1):
                    case_obj = _case_slim(case, raw_id, export_id, idx=idx, total=len(cases))
                    dump_line(fp, case_obj)
                    dump_line(fpat, {
                        "etalon_id": export_id,
                        "export_id": case_obj["export_id"],
                        "parent_export_id": export_id,
                        "raw_email_id": raw_id,
                        "stage_ids": case_obj["stage_ids"],
                        "original": original,
                        "pattern_result": case_obj,
                    })
                    dump_line(ffull, {
                        "etalon_id": export_id,
                        "export_id": case_obj["export_id"],
                        "parent_export_id": export_id,
                        "raw_email_id": raw_id,
                        "stage_ids": case_obj["stage_ids"],
                        "original": original,
                        "pattern_result": case_obj,
                        "ai_suggestions": case.get("ai_suggestions") or [],
                        "final_case": case_obj,
                        "one_c": case.get("outbox") or [],
                        "export_json": case.get("export_json") or {},
                    })
                    processed_count += 1

    bucket_counts: dict[str, int] = {}
    for _cs in cases_by_raw.values():
        for _c in _cs:
            b = _bucket(_c.get("state"))
            bucket_counts[b] = bucket_counts.get(b, 0) + 1
    manifest = {
        "created_at": utcnow(),
        "format": "jsonl: one JSON object per line",
        "files": {
            "original_emails.jsonl": "Эталон: исходные письма + текст вложений, ID вида ID1.",
            "processed_cases.jsonl": "Итоговая обработка кейсов с источником pattern/ai.",
            "etalon_pattern.jsonl": "Эталон + результат паттернов/текущего кейса.",
            "etalon_pattern_ai_1c.jsonl": "Эталон + паттерны + AI-подсказки + итоговый JSON + outbox/1C.",
            "folders_report.json": "Сводка по всем папкам, корзинам, источникам обработки и outbox.",
            "paired_index.json": "Индекс соответствия original ID -> case ids.",
        },
        "id_rule": "original_emails.jsonl export_id == processed/folders parent_export_id/export_id. Example: original ID1 <-> processed ID1.",
        "counts": {
            "original_emails": original_count,
            "processed_rows": processed_count,
            "raw_emails_with_cases": sum(1 for p in pairs if p["case_ids"]),
            "raw_emails_without_cases": sum(1 for p in pairs if not p["case_ids"]),
        },
        # Полнота: сумма по корзинам = всего писем.
        "buckets": bucket_counts,
        "buckets_sum": sum(bucket_counts.values()),
        "completeness_ok": sum(bucket_counts.values()) == original_count,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    paired_path.write_text(json.dumps(pairs, ensure_ascii=False, indent=2), encoding="utf-8")
    folders_report_path.write_text(
        json.dumps(
            {
                "created_at": utcnow(),
                "folders": sorted(folders_report.values(), key=lambda x: str(x.get("folder") or "")),
                "configured_folders": settings.folders,
            },
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        ),
        encoding="utf-8",
    )

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as z:
        for path in (original_path, processed_path, pattern_path, full_trace_path, folders_report_path, paired_path, manifest_path):
            z.write(path, arcname=path.name)

    return FileResponse(path=str(zip_path), media_type="application/zip", filename=zip_path.name)


@app.get("/api/export/source-archive/download")
def download_source_archive() -> FileResponse:
    """Download a .tar.gz archive of the project source code WITHOUT data/,
    including configs, scripts, and code."""
    import io
    import tarfile

    buf = io.BytesIO()

    # In Docker: /app/app/main.py → parent.parent = /app (project root)
    # On host:   backend/app/main.py   → parent.parent.parent = project root
    # Use /app directly for Docker context, which has code + configs + scripts
    project_root = Path(__file__).resolve().parent.parent  # /app in Docker

    exclude_dirs = {"data", "__pycache__"}
    exclude_files = {".DS_Store"}

    with tarfile.open(fileobj=buf, mode="w:gz") as tar:
        for path in project_root.rglob("*"):
            try:
                rel = path.relative_to(project_root)
            except ValueError:
                continue
            parts = rel.parts
            if any(p in exclude_dirs for p in parts):
                continue
            if path.is_file():
                base = path.name
                if base in exclude_files:
                    continue
                # Skip hidden files/dirs except .env.example
                if base.startswith(".") and base != ".env.example":
                    continue
                tar.add(path, arcname=rel)

    buf.seek(0)
    tmp = Path(f"/tmp/readmail_source_archive_{int(time.time())}.tar.gz")
    tmp.write_bytes(buf.getvalue())
    return FileResponse(
        path=str(tmp),
        media_type="application/gzip",
        filename="readmail_project_source.tar.gz",
    )


@app.get("/api/audit/pack")
def audit_pack(limit: int = Query(default=80, ge=10, le=500)) -> dict[str, Any]:
    """A compact bundle that can be pasted/sent to a stronger external model for audit."""
    apply_runtime_settings()
    with connect() as con:
        stats_data = stats()
        quality_rows = [row_to_dict(r) for r in con.execute(
            """
            SELECT c.id, c.buyer_code, c.buyer_name, c.event_type, c.claim_kind, c.status, c.priority,
                   c.confidence, c.deadline_at, c.strong_key, c.weak_key, c.ready_for_export, c.needs_review,
                   c.state, c.fields_json, c.missing_json, c.quality_json, e.direction, e.subject, e.from_addr, e.received_at, e.snippet
            FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
            WHERE c.needs_review=1 OR c.ready_for_export=1 OR c.state IN ('needs_link','linked_event')
            ORDER BY c.updated_at DESC, c.id DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()]
        learning = {}  # v2.1 AI-only: обучения нет
        ai = con.execute("SELECT COUNT(*) c, COALESCE(SUM(prompt_chars),0) p, COALESCE(SUM(response_chars),0) r FROM ai_usage").fetchone()
    prompt = (
        "Ты аудитируешь систему Project Readmail New для обработки писем по возвратам автозапчастей. "
        "Проверь: 1) не попадают ли follow-up/напоминания в new_return; 2) не является ли part_number мусором; "
        "3) хватает ли strong_key для экспорта в 1С; 4) где дедлайн/приоритет опасны; "
        "5) хватает ли доказательств по типу причины: брак=фото+акт/заказ-наряд/заключение, "
        "некондиция/недовоз/пересорт/некомплект=фото или возвратная ссылка; "
        "6) какие правила/валидаторы надо усилить. Отвечай списком конкретных багов и патчей."
    )
    return {
        "schema": "readmail-new-audit-pack-v1",
        "generated_at": utcnow(),
        "suggested_audit_prompt": prompt,
        "settings_summary": {
            "folders": settings.folders,
            "configured_folders_are_customer": settings.configured_folders_are_customer,
            "enable_ai": settings.enable_ai,
            "ai_provider": settings.ai_provider,
            "ai_model": settings.ai_model,
            "scan_interval_seconds": settings.scan_interval_seconds,
            "one_c_export_mode": settings.one_c_export_mode,
            "auto_queue_control_events": settings.auto_queue_control_events,
        },
        "stats": stats_data,
        "ai_usage": {"total": ai["c"], "prompt_chars": ai["p"], "response_chars": ai["r"]},
        "learning_counts": learning.get("counts", {}),
        "cases_sample": quality_rows,
    }




@app.post("/api/test/demo-data")
def api_test_demo_data() -> dict[str, Any]:
    apply_runtime_settings()
    return generate_demo_data(queue=True)


@app.post("/api/admin/compact")
def compact() -> dict[str, Any]:
    return {"ok": True, **compact_db()}


@app.get("/api/ai/discover")
def api_ai_discover() -> dict[str, Any]:
    """Scan local ports for running AI servers (Ollama, MLX, vLLM)."""
    import socket
    ports_str = getattr(settings, "ai_auto_discover_ports", "11434,8080,8010,1337")
    ports = [int(p.strip()) for p in ports_str.split(",") if p.strip().isdigit()]
    results = []
    for port in ports:
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.settimeout(2)
            r = s.connect_ex(("host.docker.internal", port))
            s.close()
            if r == 0:
                kind = {11434: "ollama", 8080: "mlx", 8010: "vllm", 1337: "custom", 11435: "ollama"}.get(port, "unknown")
                results.append({"port": port, "status": "open", "kind": kind})
            else:
                results.append({"port": port, "status": "closed"})
        except Exception:
            results.append({"port": port, "status": "error"})
    
    # Also check RouterAI
    routerai_ok = False
    try:
        import httpx
        r = httpx.get(settings.routerai_base_url.rstrip("/") + "/v1/models", timeout=5)
        routerai_ok = r.status_code < 400
    except Exception:
        pass
    
    return {
        "ok": True,
        "local": results,
        "routerai": {"status": "ok" if routerai_ok else "unreachable", "url": settings.routerai_base_url},
        "current_provider": settings.ai_provider,
        "current_model": settings.ai_model,
        "discovery_enabled": getattr(settings, "ai_auto_discover_enabled", True),
    }


@app.get("/api/ai/journal")
def api_ai_journal(limit: int = Query(default=20, ge=1, le=100)) -> dict[str, Any]:
    """Return recent AI decisions log."""
    with connect() as con:
        rows = con.execute(
            "SELECT * FROM ai_suggestions ORDER BY created_at DESC LIMIT ?",
            (limit,)
        ).fetchall()
        items = [dict(r) for r in rows]
        for item in items:
            if isinstance(item.get("response_json"), str):
                try:
                    item["response_parsed"] = json.loads(item["response_json"])
                except: pass
    return {"ok": True, "items": items}



@app.get("/api/config/buyer/{code}")
def api_get_buyer_config(code: str) -> dict[str, Any]:
    """Return raw buyer config YAML content."""
    cfg_path = settings.buyer_config_dir / f"{code}.yml"
    if not cfg_path.exists():
        raise HTTPException(404, f"Buyer config '{code}' not found")
    try:
        return {"ok": True, "code": code, "content": cfg_path.read_text(encoding="utf-8"), "path": str(cfg_path)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/config/buyer/{code}")
def api_update_buyer_config(code: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Update buyer config YAML. Accepts 'content' with full YAML string."""
    content = (payload or {}).get("content", "").strip()
    if not content:
        raise HTTPException(400, "content is required")
    cfg_path = settings.buyer_config_dir / f"{code}.yml"
    try:
        cfg_path.write_text(content, encoding="utf-8")
        _log("config", f"Buyer config '{code}' updated via API", level="ok", details={"path": str(cfg_path)})
        return {"ok": True, "code": code, "message": f"Config '{code}' saved", "size": len(content)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/ai/context")
def api_get_ai_context() -> dict[str, Any]:
    """Return AI_CONTEXT.md content."""
    from pathlib import Path as _Path
    base = Path(__file__).resolve().parent.parent.parent
    for candidate in [base / "AI_CONTEXT.md", base / ".." / "AI_CONTEXT.md", Path("/app/AI_CONTEXT.md")]:
        try:
            if candidate.exists():
                return {"ok": True, "content": candidate.read_text(encoding="utf-8"), "path": str(candidate.resolve())}
        except Exception:
            continue
    return {"ok": False, "error": "AI_CONTEXT.md not found"}


@app.post("/api/ai/context")
def api_update_ai_context(payload: dict[str, Any]) -> dict[str, Any]:
    """Update AI_CONTEXT.md content."""
    content = (payload or {}).get("content", "").strip()
    if not content:
        raise HTTPException(400, "content is required")
    from pathlib import Path as _Path
    base = Path(__file__).resolve().parent.parent.parent
    for candidate in [base / "AI_CONTEXT.md", Path("/app/AI_CONTEXT.md")]:
        try:
            candidate.parent.mkdir(parents=True, exist_ok=True)
            candidate.write_text(content, encoding="utf-8")
            _log("config", "AI_CONTEXT.md updated via API", level="ok", details={"path": str(candidate.resolve())})
            return {"ok": True, "message": "AI_CONTEXT.md updated"}
        except (OSError, IOError):
            continue
    return {"ok": False, "error": "Could not write AI_CONTEXT.md"}


@app.post("/api/reload/config")
def api_reload_config() -> dict[str, Any]:
    """Reload buyer configs and runtime settings without restart."""
    try:
        from .classifier import load_buyer_rules as _reload_rules
        from .runtime_settings import apply_runtime_settings as _reload_settings
        _reload_settings()
        result = {"ok": True, "message": "Configs reloaded", "settings_applied": True}
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════════════════
# READMAIL v2 — новые эндпоинты
# ═══════════════════════════════════════════════════════════

import threading as _threading
from .archiver import run_archive_cleanup, get_archive_stats
from .ai_client import get_token_stats

# ── Хранилище undo-токенов (в памяти, TTL 60 сек) ──
_undo_store: dict[int, float] = {}
_undo_lock = _threading.Lock()

def _register_undo(case_id: int, seconds: int = 60) -> None:
    import time
    with _undo_lock:
        _undo_store[case_id] = time.time() + seconds

def _can_undo(case_id: int) -> bool:
    import time
    with _undo_lock:
        exp = _undo_store.get(case_id)
        return exp is not None and time.time() < exp

def _clear_undo(case_id: int) -> None:
    with _undo_lock:
        _undo_store.pop(case_id, None)


# ── /api/import (алиас для нового UI) ──
_IMPORT_BG_STATE: dict[str, Any] = {"running": False, "started_at": None, "finished_at": None, "imported": 0, "classified": 0, "result": None, "error": None, "start_emails": None}


def _background_import() -> None:
    from .imap_importer import import_from_imap_raw
    started = utcnow()
    start_emails = 0
    try:
        with connect() as con:
            start_emails = int(con.execute("SELECT COUNT(*) FROM raw_emails").fetchone()[0] or 0)
    except Exception:
        start_emails = 0
    _IMPORT_BG_STATE.update({"running": True, "stage": "import", "started_at": started, "finished_at": None, "imported": 0, "classified": 0, "result": None, "error": None, "start_emails": start_emails})
    if not _IMPORT_LOCK.acquire(blocking=False):
        _IMPORT_BG_STATE.update({
            "running": False,
            "finished_at": utcnow(),
            "error": "import_already_running",
            "result": {"ok": False, "skipped": True, "reason": "import_already_running"},
        })
        return
    try:
        apply_runtime_settings()
        result = import_from_imap_raw()
        classified = 0
        if result.get("ok", True):
            result["classified"] = 0
            result["manual_flow"] = "import_only"
        _IMPORT_BG_STATE.update({
            "imported": result.get("imported", 0),
            "classified": classified,
            "result": {k: result.get(k) for k in ("ok", "imported", "skipped", "classified", "manual_flow", "errors", "total_on_server", "folders_processed", "folders")},
        })
        global _LAST_IMPORT_RESULT
        _LAST_IMPORT_RESULT = {**result, "finished_at": utcnow()}
    except Exception as exc:
        _IMPORT_BG_STATE["error"] = str(exc)
    finally:
        _IMPORT_LOCK.release()
        _IMPORT_BG_STATE["running"] = False
        _IMPORT_BG_STATE["stage"] = "done"
        _IMPORT_BG_STATE["finished_at"] = utcnow()


@app.post("/api/import")
def api_import_v2() -> dict[str, Any]:
    """Запустить импорт IMAP в фоне — возвращается сразу, прогресс через /api/import/progress."""
    if _IMPORT_BG_STATE.get("running") or _IMPORT_LOCK.locked() or _PATTERNS_STATE.get("running"):
        return {"ok": True, "already_running": True, "reason": "pipeline_busy", "state": dict(_IMPORT_BG_STATE), "patterns": dict(_PATTERNS_STATE)}
    _IMPORT_BG_STATE.update({"running": True, "stage": "starting", "started_at": utcnow(), "finished_at": None, "imported": 0, "classified": 0, "result": None, "error": None, "start_emails": None})
    th = threading.Thread(target=_background_import, name="readmail-manual-import", daemon=True)
    th.start()
    return {"ok": True, "started": True, "background": True, "state": dict(_IMPORT_BG_STATE)}


@app.get("/api/import/progress")
def api_import_progress() -> dict[str, Any]:
    """Текущий прогресс фонового импорта + сколько писем уже в базе."""
    emails = None
    db_error = None
    try:
        with connect() as con:
            emails = con.execute("SELECT COUNT(*) c FROM raw_emails").fetchone()["c"]
    except Exception as exc:
        db_error = str(exc)
    state = dict(_IMPORT_BG_STATE)
    state["running"] = bool(state.get("running") or _IMPORT_LOCK.locked())
    if emails is not None and state.get("running") and state.get("start_emails") is not None:
        state["imported"] = max(
            int(state.get("imported") or 0),
            max(0, int(emails or 0) - int(state.get("start_emails") or 0)),
        )
    return {"ok": db_error is None, "state": state, "emails_in_db": emails, "db_error": db_error}


# ── /api/emails (с фильтрами для нового UI) ──
@app.get("/api/emails")
def api_emails_v2(
    filter: str = "all",
    buyer: str = "",
    q: str = "",
    sort: str = "date_desc",
    page: int = 1,
    limit: int = 50,
) -> dict[str, Any]:
    try:
        with connect() as con:
            conditions = []
            params: list[Any] = []

            if filter == "new":
                conditions.append("c.event_type = 'new_return'")
            elif filter == "reminder":
                conditions.append("c.event_type = 'followup_reminder'")
            elif filter == "quarantine":
                conditions.append("c.state = 'link_quarantine'")
            elif filter == "unprocessed":
                conditions.append("COALESCE(c.state, e.status) IN ('needs_review', 'pending')")

            if buyer:
                conditions.append("c.buyer_code = ?")
                params.append(buyer)

            if q:
                conditions.append("(e.subject LIKE ? OR e.from_addr LIKE ? OR c.fields_json LIKE ?)")
                params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

            total = con.execute(
                f"SELECT COUNT(*) FROM raw_emails e LEFT JOIN cases c ON c.raw_email_id = e.id {where}",
                params
            ).fetchone()[0]

            # Сортировка
            sort_map = {
                "date_desc": "e.received_at DESC",
                "date_asc": "e.received_at ASC",
                "client": "c.buyer_name ASC, e.received_at DESC",
                "subject": "e.subject ASC, e.received_at DESC",
                "priority": "CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END, e.received_at DESC",
                "status": "COALESCE(c.state, e.status) ASC, e.received_at DESC",
            }
            order_clause = sort_map.get(sort, "e.received_at DESC")

            offset = (page - 1) * limit
            rows = con.execute(
                f"""SELECT e.id, c.id AS case_id, c.buyer_code, c.buyer_name, c.event_type, c.claim_kind,
                    COALESCE(c.state, e.status) AS state, c.priority, c.confidence, c.deadline_at, c.fields_json,
                    e.from_addr, e.subject, e.received_at,
                    (CASE WHEN c.payload_json LIKE '%"processing_source":"ai"%' OR c.payload_json LIKE '%ai_overlay%' THEN 1 ELSE 0 END) AS ai_processed,
                    (SELECT COUNT(*) FROM attachments a WHERE a.raw_email_id = e.id) as has_attachments
                    FROM raw_emails e LEFT JOIN cases c ON c.raw_email_id = e.id
                    {where}
                    ORDER BY {order_clause}
                    LIMIT ? OFFSET ?""",
                params + [limit, offset]
            ).fetchall()

            import json as _json

            emails = []
            for r in rows:
                d = dict(r)
                fld = {}
                try:
                    fld = _json.loads(d.pop("fields_json") or "{}")
                except Exception:
                    d.pop("fields_json", None)
                d["part_number"] = fld.get("part_number") or ""
                d["document_number"] = fld.get("document_number") or ""
                d["brand"] = fld.get("brand") or ""
                d["claim_number"] = fld.get("claim_number") or ""
                d["fields"] = fld
                emails.append(d)

            # Список уникальных клиентов для фильтра
            buyer_rows = con.execute(
                "SELECT DISTINCT buyer_code as code, buyer_name as name FROM cases WHERE buyer_code IS NOT NULL ORDER BY buyer_name"
            ).fetchall()
            buyers = [dict(r) for r in buyer_rows]

            return {"ok": True, "total": total, "emails": emails, "buyers": buyers}
    except Exception as e:
        return {"ok": False, "error": str(e), "emails": [], "total": 0, "buyers": []}


# ── /api/emails/{id} ──
@app.get("/api/emails/{email_id}")
def api_email_detail_v2(email_id: int) -> dict[str, Any]:
    try:
        with connect() as con:
            row = con.execute(
                "SELECT * FROM raw_emails WHERE id = ?", (email_id,)
            ).fetchone()
            if not row:
                return {"ok": False, "error": "not found"}
            data = dict(row)
            atts = con.execute(
                "SELECT id, filename, content_type, size_bytes FROM attachments WHERE raw_email_id = ?",
                (email_id,)
            ).fetchall()
            data["attachments"] = [dict(a) for a in atts]
            return {"ok": True, **data}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── /api/emails/{id}/ai ──
@app.post("/api/emails/{email_id}/ai")
def api_email_ai_v2(email_id: int) -> dict[str, Any]:
    try:
        with connect() as con:
            case_row = con.execute(
                "SELECT id FROM cases WHERE raw_email_id = ? LIMIT 1", (email_id,)
            ).fetchone()
            if not case_row:
                return {"ok": False, "error": "case not found for email"}
        return api_ai_suggest_case(case_row["id"])
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── /api/cases (расширенный с новыми фильтрами) ──
@app.get("/api/cases")
def api_cases_v2(
    state: str = "needs_review",
    priority: str = "",
    event_type: str = "",
    date_from: str = "",
    date_to: str = "",
    page: int = 1,
    limit: int = 50,
) -> dict[str, Any]:
    try:
        with connect() as con:
            conditions = []
            params: list[Any] = []

            STATE_MAP = {
                "needs_review": ["needs_review"],
                "needs_link": ["needs_link"],
                "waiting_docs": ["waiting_docs", "evidence_requested"],
                "overdue": None,  # особый случай
                "ready_to_1c": ["ready_to_export", "export_queued"],
                "sent": ["delivered"],
            }

            if state == "overdue":
                conditions.append("c.deadline_at < datetime('now') AND c.status NOT IN ('delivered','closed')")
            elif state in STATE_MAP and STATE_MAP[state]:
                placeholders = ",".join("?" * len(STATE_MAP[state]))
                conditions.append(f"c.status IN ({placeholders})")
                params.extend(STATE_MAP[state])
            elif state not in ("all", ""):
                conditions.append("c.status = ?")
                params.append(state)

            if priority:
                conditions.append("c.priority = ?")
                params.append(priority)

            if event_type:
                conditions.append("c.event_type = ?")
                params.append(event_type)

            if date_from:
                conditions.append("e.received_at >= ?")
                params.append(date_from)

            if date_to:
                conditions.append("e.received_at <= ?")
                params.append(date_to + "T23:59:59")

            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

            total = con.execute(
                f"SELECT COUNT(*) FROM cases c JOIN raw_emails e ON c.raw_email_id = e.id {where}",
                params
            ).fetchone()[0]

            offset = (page - 1) * limit
            rows = con.execute(
                f"""SELECT c.id, c.buyer_code, c.buyer_name, c.event_type, c.claim_kind,
                    c.status, c.priority, c.confidence, c.deadline_at, c.fields_json,
                    c.quality_json, c.missing_json,
                    e.from_addr, e.subject, e.received_at,
                    ob.status as outbox_status
                    FROM cases c
                    JOIN raw_emails e ON c.raw_email_id = e.id
                    LEFT JOIN outbox ob ON ob.case_id = c.id
                    {where}
                    ORDER BY c.priority DESC, e.received_at DESC
                    LIMIT ? OFFSET ?""",
                params + [limit, offset]
            ).fetchall()

            cases = []
            for r in rows:
                d = dict(r)
                import json as _json
                d["fields"] = _json.loads(d.pop("fields_json") or "{}")
                d["quality"] = _json.loads(d.pop("quality_json") or "[]")
                d["missing"] = _json.loads(d.pop("missing_json") or "[]")
                cases.append(d)

            return {"ok": True, "total": total, "cases": cases}
    except Exception as e:
        return {"ok": False, "error": str(e), "cases": [], "total": 0}


# ── /api/cases/{id}/confirm-with-undo  (двойное подтверждение с undo) ──
@app.post("/api/cases/{case_id}/confirm-with-undo")
def api_confirm_case_v2(case_id: int) -> dict[str, Any]:
    try:
        undo_secs = int(settings.confirm_undo_seconds)
        _register_undo(case_id, undo_secs)
        with connect() as con:
            con.execute(
                "UPDATE cases SET status = 'confirmed_pending_undo', updated_at = ? WHERE id = ?",
                (utcnow(), case_id)
            )
        # Через undo_secs финализируем если не отменили
        def _finalize():
            import time
            time.sleep(undo_secs + 1)
            if not _can_undo(case_id):
                return
            _clear_undo(case_id)
            with connect() as con:
                con.execute(
                    "UPDATE cases SET status = 'ready_to_export', updated_at = ? WHERE id = ? AND status = 'confirmed_pending_undo'",
                    (utcnow(), case_id)
                )
        _threading.Thread(target=_finalize, daemon=True).start()
        return {"ok": True, "case_id": case_id, "undo_seconds": undo_secs}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class TrainingCorrection(BaseModel):
    fields: dict[str, Any]
    ai_generate_patterns: bool = True
    # Мультипозиция: оператор правит/добавляет позиции карточками в Сверке. Каждая —
    # {part_number, brand, product_name, quantity, price}. Если задано — пишем в
    # payload.table_items, а позицию 1 синхронизируем с плоскими fields.
    items: list[dict[str, Any]] | None = None


@app.post("/api/cases/{case_id}/train")
def api_case_train(case_id: int, correction: TrainingCorrection) -> dict[str, Any]:
    """Operator corrects fields manually → system learns regex patterns + optionally asks AI."""
    from .classifier import quality_check, _normalize_date
    from .db import record_learning_event

    try:
        with connect() as con:
            row = con.execute(
                """SELECT c.*, e.body_text, e.subject, e.from_addr, e.received_at, e.snippet
                   FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id WHERE c.id=?""",
                (case_id,),
            ).fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="case not found")
            data = row_to_dict(row) or {}

        before_fields = data.get("fields") or {}
        after_fields = dict(before_fields)

        # Merge operator corrections, normalize dates
        for k, v in correction.fields.items():
            if v not in (None, ""):
                val = str(v).strip()
                if k == "document_date":
                    val = _normalize_date(val) or val
                after_fields[k] = val
            elif k in after_fields:
                del after_fields[k]

        # Мультипозиция: оператор отредактировал/добавил позиции карточками. Чистим, пишем
        # в payload.table_items (build_export_json выгрузит все), позицию 1 кладём в fields.
        payload = dict(data.get("payload") or {})
        if correction.items is not None:
            POS_KEYS = ("part_number", "brand", "product_name", "quantity", "price")
            clean_items: list[dict[str, Any]] = []
            for it in correction.items:
                row_it: dict[str, Any] = {}
                for pk in POS_KEYS:
                    pv = (it or {}).get(pk)
                    if pv not in (None, ""):
                        row_it[pk] = str(pv).strip()
                if row_it.get("part_number") or row_it.get("product_name"):
                    clean_items.append(row_it)
            payload["table_items"] = clean_items if len(clean_items) > 1 else None
            payload["multi_item_count"] = len(clean_items)
            data["payload"] = payload
            # Позицию 1 синхронизируем с плоскими полями (источник для одиночной выдачи).
            if clean_items:
                first = clean_items[0]
                for pk in POS_KEYS:
                    if first.get(pk):
                        after_fields[pk] = first[pk]
                    elif pk in after_fields and pk not in ("document_number", "document_date"):
                        after_fields.pop(pk, None)

        # Re-run quality_check with corrected fields
        direction = (data.get("payload") or {}).get("direction", "inbound_customer")
        buyer_code = data.get("buyer_code")
        event_type = data.get("event_type") or "new_return"
        claim_kind = data.get("claim_kind")
        from .classifier import make_strong_key
        strong_key = make_strong_key(buyer_code, after_fields)
        missing, quality_issues = quality_check(
            event_type, claim_kind, after_fields, strong_key, direction, buyer_code
        )
        hard_errors = [q for q in quality_issues if q.get("level") == "error"]

        ready_for_export = (
            event_type == "new_return"
            and not missing
            and not hard_errors
            and direction == "inbound_customer"
        )
        new_state = "ready_to_1c" if ready_for_export else data.get("state", "needs_review")

        # Save corrected case
        with connect() as con:
            corrected = {
                **data,
                "id": case_id,
                "fields": after_fields,
                "missing": missing,
                "quality": quality_issues,
                "payload": payload,
                "ready_for_export": ready_for_export,
                "needs_review": not ready_for_export,
                "state": new_state,
                "strong_key": strong_key,
            }
            corrected["export"] = _case_export_from_row(corrected, after_fields)
            save_case(
                con,
                int(data["raw_email_id"]),
                corrected,
                item_index=int(data.get("item_index") or 0),
            )
            saved = con.execute(
                "SELECT state, ready_for_export, payload_json FROM cases WHERE id=?",
                (case_id,),
            ).fetchone()
            new_state = str(saved["state"] or "needs_review")
            ready_for_export = bool(saved["ready_for_export"])
            saved_payload = loads(saved["payload_json"], {})
            evidence_gate = saved_payload.get("evidence_gate") or {}
            if ready_for_export and new_state == "ready_to_1c":
                queue_case_event(con, case_id)
            record_learning_event(
                con,
                kind="operator_correction",
                source="manual_train",
                case_id=case_id,
                confidence=0.95,
                payload={"before": before_fields, "after": after_fields, "buyer_code": buyer_code},
            )

        # v2.1 AI-only: генерация regex-паттернов из правок и AI-тренировка убраны.

        # v2.1 AI-only: AI-тренировка паттернов, автопромоут в YAML и перепрогон убраны.
        return {
            "ok": True,
            "case_id": case_id,
            "buyer_code": buyer_code,
            "before_fields": before_fields,
            "after_fields": after_fields,
            "state": new_state,
            "ready_for_export": ready_for_export,
            "evidence_gate": evidence_gate,
            "missing": missing,
        }
    except HTTPException:
        raise
    except Exception as exc:
        _log("learning", "Ручное обучение: ошибка сохранения", level="error", case_id=case_id, details={"error": str(exc)})
        return {"ok": False, "case_id": case_id, "error": str(exc)}


# v2.1 AI-only: эндпоинт promote-patterns удалён.


# ── /api/cases/{id}/undo-confirm ──
@app.post("/api/cases/{case_id}/undo-confirm")
def api_undo_confirm_v2(case_id: int) -> dict[str, Any]:
    if not _can_undo(case_id):
        return {"ok": False, "error": "undo window expired"}
    _clear_undo(case_id)
    try:
        with connect() as con:
            con.execute(
                "UPDATE cases SET status = 'needs_review', updated_at = ? WHERE id = ? AND status = 'confirmed_pending_undo'",
                (utcnow(), case_id)
            )
        return {"ok": True, "case_id": case_id}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── /api/cases/{id}/ai ──
@app.post("/api/cases/{case_id}/ai")
def api_case_ai_v2(case_id: int) -> dict[str, Any]:
    return api_ai_suggest_case(case_id)


def api_ai_suggest_case(case_id: int) -> dict[str, Any]:
    """Запустить AI на одном кейсе — через правильный пайплайн (run_ai_suggestion + overlay)."""
    try:
        return _apply_ai_to_case_id(case_id, purpose="manual_single_ai")
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── /api/cases/{id}/export ──
@app.post("/api/cases/{case_id}/export")
def api_case_export_v2(case_id: int) -> dict[str, Any]:
    try:
        with connect() as con:
            con.execute(
                "UPDATE cases SET status = 'ready_to_export', updated_at = ? WHERE id = ?",
                (utcnow(), case_id)
            )
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── /api/ai/run-batch ──
_AI_BATCH_STATE: dict[str, Any] = {"running": False, "total": 0, "processed": 0, "resolved": 0, "started_at": None, "finished_at": None, "error": None, "current": None}


def _ai_batch_worker(limit: int, target: str, force: bool = False, order: str = "new") -> None:
    _AI_BATCH_STATE.update({"running": True, "total": 0, "processed": 0, "resolved": 0, "started_at": utcnow(), "finished_at": None, "error": None, "current": None})
    try:
        apply_runtime_settings()
        if target == "unknown":
            event_filter = "AND c.event_type = 'unknown'"
        elif target == "returns":
            event_filter = "AND c.event_type = 'new_return'"
        else:
            event_filter = "AND c.event_type IN ('new_return','unknown')"
        # force=True — прогнать AI заново даже по кейсам с уже принятой подсказкой
        # (нужно после улучшения промта, чтобы переобработать застрявшие).
        accepted_filter = "" if force else "AND NOT EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=c.id AND s.accepted=1)"
        with connect() as con:
            rows = con.execute(
                f"""
                SELECT c.id FROM cases c
                WHERE c.state IN ('needs_review', 'needs_link')
                  {event_filter}
                  {accepted_filter}
                ORDER BY
                  CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
                  c.id """ + ("ASC" if order == "old" else "DESC") + """
                LIMIT ?
                """,
                (int(limit),),
            ).fetchall()
        ids = [int(r["id"]) for r in rows]
        _AI_BATCH_STATE["total"] = len(ids)
        for cid in ids:
            if _AI_BATCH_STOP.is_set():
                break
            _AI_BATCH_STATE["current"] = cid
            try:
                result = _apply_ai_to_case_id(
                    cid,
                    purpose="manual_ai_batch",  # v2.1: полный промт (компактного больше нет — вход дешевле выхода)
                    queue_ready=False,
                    manual_review_gate=True,
                )
                if result.get("ok"):
                    _AI_BATCH_STATE["processed"] += 1
                    if result.get("ready_for_export"):
                        _AI_BATCH_STATE["resolved"] += 1
            except Exception as exc:
                _log("ai", f"AI-батч: ошибка кейса {cid}: {exc}", level="error", case_id=cid)
    except Exception as exc:
        _AI_BATCH_STATE["error"] = str(exc)
    finally:
        _AI_BATCH_STATE["running"] = False
        _AI_BATCH_STATE["current"] = None
        _AI_BATCH_STATE["finished_at"] = utcnow()


def _ai_full_batch_worker(limit: int) -> None:
    """Фоновый полный ИИ-прогон (Автопилот ИИ, ручной режим): ИИ по N свежим возвратам подряд.
    Использует общий _AI_BATCH_STATE/_AI_BATCH_STOP → прогресс/стоп/лог работают как у run-batch."""
    _AI_BATCH_STATE.update({"running": True, "total": 0, "processed": 0, "resolved": 0,
                            "started_at": utcnow(), "finished_at": None, "error": None, "current": None})
    try:
        apply_runtime_settings()
        ids = _select_full_ai_ids(int(limit))
        _AI_BATCH_STATE["total"] = len(ids)
        for cid in ids:
            if _AI_BATCH_STOP.is_set():
                break
            _AI_BATCH_STATE["current"] = cid
            try:
                result = _apply_ai_to_case_id(cid, purpose="manual_full_ai")
                if result.get("ok"):
                    _AI_BATCH_STATE["processed"] += 1
                    if result.get("ready_for_export"):
                        _AI_BATCH_STATE["resolved"] += 1
            except Exception as exc:
                _log("ai", f"Полный ИИ: ошибка кейса {cid}: {exc}", level="error", case_id=cid)
    except Exception as exc:
        _AI_BATCH_STATE["error"] = str(exc)
    finally:
        _AI_BATCH_STATE["running"] = False
        _AI_BATCH_STATE["current"] = None
        _AI_BATCH_STATE["finished_at"] = utcnow()


@app.post("/api/ai/run-batch")
def api_ai_run_batch(limit: int = 200, target: str = "all", force: bool = False, order: str = "new") -> dict[str, Any]:
    """Запустить AI-батч в ФОНЕ (не блокирует UI/БД). Прогресс — /api/ai/batch-progress.
    order='new' — свежие первыми, 'old' — старые (бэклог) первыми."""
    apply_runtime_settings()
    if not settings.enable_ai:
        return {"ok": False, "error": "AI выключен в настройках (ENABLE_AI=false)"}
    if _AI_BATCH_STATE.get("running"):
        return {"ok": True, "already_running": True, "state": dict(_AI_BATCH_STATE)}
    _AI_BATCH_STOP.clear()
    th = threading.Thread(target=_ai_batch_worker, args=(int(limit), target, bool(force), order), name="readmail-ai-batch", daemon=True)
    th.start()
    return {"ok": True, "started": True, "background": True}


@app.get("/api/cases/{case_id}/links")
def api_case_links(case_id: int) -> dict[str, Any]:
    """Все ссылки письма (для кликабельного показа в панели)."""
    from .link_fetcher import extract_links_from_email
    with connect() as con:
        row = con.execute(
            "SELECT e.subject, e.body_text, e.body_html FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id WHERE c.id=?",
            (case_id,),
        ).fetchone()
    if not row:
        return {"ok": False, "error": "case_not_found", "links": []}
    return {"ok": True, "links": extract_links_from_email(row["subject"], row["body_text"], row["body_html"])}


@app.post("/api/cases/{case_id}/read-links")
def api_read_case_links(case_id: int) -> dict[str, Any]:
    """Прочитать ВСЕ ссылки письма (1,2,3 в одном) → поля + фото + документы. Для брака
    документы/фото со страницы — доказательная база. На забаненном IP вернёт blocked>0."""
    from .link_fetcher import read_email_links
    with connect() as con:
        row = con.execute(
            "SELECT e.subject, e.body_text, e.body_html FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id WHERE c.id=?",
            (case_id,),
        ).fetchone()
    if not row:
        return {"ok": False, "error": "case_not_found"}
    res = read_email_links(row["subject"], row["body_text"], row["body_html"])
    res["case_id"] = case_id
    if res.get("blocked") and not res.get("processed"):
        res["hint"] = "IP сервера забанен источником (403). Нужен прокси/VPN-выход или сохранённая страница."
    return res
def api_ai_read_attachments(case_id: int) -> dict[str, Any]:
    """ИИ дочитывает вложения (Excel-акт/doc) и добивает поля (бренд/имя/артикул из акта брака)."""
    apply_runtime_settings()
    if not settings.enable_ai:
        return {"ok": False, "error": "AI выключен в настройках"}
    with connect() as con:
        ex = con.execute("SELECT raw_email_id FROM cases WHERE id=?", (case_id,)).fetchone()
    if not ex:
        return {"ok": False, "error": f"Кейс #{case_id} не найден"}
    has_text = bool(_case_attachment_text(int(ex["raw_email_id"])).strip())
    if not has_text:
        return {"ok": False, "error": "В вложениях нет читаемого текста (Excel/doc). Для фото — Vision."}
    try:
        res = _apply_ai_to_case_id(case_id, purpose="manual_full_ai", queue_ready=False,
                                   manual_review_gate=True, read_attachments=True)
        return {"ok": bool(res.get("ok")), "case_id": case_id, "applied": res.get("applied"),
                "ready_for_export": res.get("ready_for_export"), "error": res.get("error")}
    except Exception as exc:
        return {"ok": False, "case_id": case_id, "error": str(exc)}


@app.post("/api/ai/run-one")
def api_ai_run_one(case_id: int = Query(...)) -> dict[str, Any]:
    """ИИ по одному кейсу (оператор даёт номер). Результат сразу в Сверку."""
    apply_runtime_settings()
    if not settings.enable_ai:
        return {"ok": False, "error": "AI выключен в настройках"}
    with connect() as con:
        ex = con.execute("SELECT id FROM cases WHERE id=?", (case_id,)).fetchone()
    if not ex:
        return {"ok": False, "error": f"Кейс #{case_id} не найден"}
    try:
        res = _apply_ai_to_case_id(case_id, purpose="manual_full_ai", queue_ready=False, manual_review_gate=True)
        return {"ok": bool(res.get("ok")), "case_id": case_id, "applied": res.get("applied"),
                "ready_for_export": res.get("ready_for_export"), "error": res.get("error")}
    except Exception as exc:
        return {"ok": False, "case_id": case_id, "error": str(exc)}


@app.get("/api/ai/batch-progress")
def api_ai_batch_progress() -> dict[str, Any]:
    return {"ok": True, "state": dict(_AI_BATCH_STATE)}


@app.get("/api/ai/live-log")
def api_ai_live_log(limit: int = 30) -> dict[str, Any]:
    """Живой AI-лог «от запроса до вывода» — последние вызовы (для режима обучения)."""
    from .ai_client import get_ai_live_log
    return {"ok": True, "items": get_ai_live_log(limit)}


@app.get("/api/ai/usage-by-mode")
def api_ai_usage_by_mode() -> dict[str, Any]:
    """Токены за сегодня раздельно по режимам (pattern vs full_ai) и типам (text/vision) — для 2 овалов."""
    today = utcnow()[:10]
    agg: dict[str, Any] = {
        "pattern": {"prompt_tokens": 0, "completion_tokens": 0, "text": {"pt": 0, "ct": 0}, "vision": {"pt": 0, "ct": 0}, "n": 0},
        "full_ai": {"prompt_tokens": 0, "completion_tokens": 0, "text": {"pt": 0, "ct": 0}, "vision": {"pt": 0, "ct": 0}, "n": 0},
    }
    with connect() as con:
        rows = con.execute(
            """SELECT COALESCE(mode,'pattern') mode, COALESCE(kind,'text') kind,
                      COALESCE(SUM(prompt_tokens),0) pt, COALESCE(SUM(completion_tokens),0) ct, COUNT(*) n
               FROM ai_usage WHERE substr(created_at,1,10)=? GROUP BY mode, kind""",
            (today,),
        ).fetchall()
    for r in rows:
        m = r["mode"] if r["mode"] in ("pattern", "full_ai") else "pattern"
        k = "vision" if r["kind"] == "vision" else "text"
        d = agg[m]
        d["prompt_tokens"] += r["pt"]; d["completion_tokens"] += r["ct"]; d["n"] += r["n"]
        d[k]["pt"] += r["pt"]; d[k]["ct"] += r["ct"]
    return {"ok": True, "today": today, "modes": agg}


@app.get("/api/ai/usage-timeline")
def api_ai_usage_timeline(period: str = "day", limit: int = 14) -> dict[str, Any]:
    """Аналитика токенов по дням/неделям/месяцам, раздельно режим (pattern/full_ai) и тип (text/vision)."""
    grp = "substr(created_at,1,7)" if period == "month" else "strftime('%Y-W%W', created_at)" if period == "week" else "substr(created_at,1,10)"
    with connect() as con:
        rows = con.execute(
            f"""SELECT {grp} period, COALESCE(mode,'pattern') mode, COALESCE(kind,'text') kind,
                       COALESCE(SUM(prompt_tokens),0) pt, COALESCE(SUM(completion_tokens),0) ct
                FROM ai_usage GROUP BY period, mode, kind ORDER BY period DESC""",
        ).fetchall()
    # Сворачиваем в {period: {pattern:{text:{pt,ct},vision:{...}}, full_ai:{...}}}
    out: dict[str, Any] = {}
    for r in rows:
        p = r["period"] or "—"
        m = r["mode"] if r["mode"] in ("pattern", "full_ai") else "pattern"
        k = "vision" if r["kind"] == "vision" else "text"
        pd = out.setdefault(p, {"pattern": {"text": {"pt": 0, "ct": 0}, "vision": {"pt": 0, "ct": 0}},
                                "full_ai": {"text": {"pt": 0, "ct": 0}, "vision": {"pt": 0, "ct": 0}}})
        pd[m][k]["pt"] += r["pt"]; pd[m][k]["ct"] += r["ct"]
    periods = list(out.keys())[:limit]
    return {"ok": True, "period": period, "periods": [{"period": p, **out[p]} for p in periods]}


@app.get("/api/ai/token-report")
def api_ai_token_report() -> dict[str, Any]:
    """Сводка токенов для сверки: по двум режимам (паттерн+ИИ / полный ИИ) — текст вход/выход,
    визуал вход/выход, итого и СРЕДНЕЕ на 1 письмо (по всем запросам кейса). Read-only.
    Письмо = уникальный case_id; среднее = (вход+выход) / число писем в режиме."""
    apply_runtime_settings()

    def _empty() -> dict[str, Any]:
        return {"text": {"in": 0, "out": 0, "calls": 0}, "vision": {"in": 0, "out": 0, "calls": 0},
                "total": {"in": 0, "out": 0, "calls": 0}, "emails": 0, "avg_tokens_per_email": 0.0}

    labels = {"pattern": "Паттерн + ИИ", "full_ai": "Полный ИИ", "untagged": "Без режима"}
    report: dict[str, Any] = {k: _empty() for k in ("pattern", "full_ai", "untagged")}
    with connect() as con:
        rows = con.execute(
            """SELECT COALESCE(mode,'untagged') m,
                      CASE WHEN kind='vision' OR model LIKE '%vl%' OR model LIKE '%vision%'
                           THEN 'vision' ELSE 'text' END k,
                      COUNT(*) n, COALESCE(SUM(prompt_tokens),0) pin, COALESCE(SUM(completion_tokens),0) pout
               FROM ai_usage GROUP BY m, k""",
        ).fetchall()
        emails = con.execute(
            "SELECT COALESCE(mode,'untagged') m, COUNT(DISTINCT case_id) e FROM ai_usage "
            "WHERE case_id IS NOT NULL GROUP BY m",
        ).fetchall()
    email_by_mode = {r["m"]: int(r["e"]) for r in emails}
    for r in rows:
        m = r["m"] if r["m"] in report else "untagged"
        k = r["k"]
        report[m][k] = {"in": int(r["pin"]), "out": int(r["pout"]), "calls": int(r["n"])}
    total = _empty()
    for m, b in report.items():
        b["total"] = {"in": b["text"]["in"] + b["vision"]["in"],
                      "out": b["text"]["out"] + b["vision"]["out"],
                      "calls": b["text"]["calls"] + b["vision"]["calls"]}
        b["emails"] = email_by_mode.get(m, 0)
        toks = b["total"]["in"] + b["total"]["out"]
        b["emails_label"] = labels.get(m, m)
        b["avg_tokens_per_email"] = round(toks / b["emails"], 1) if b["emails"] else 0.0
        for kk in ("text", "vision", "total"):
            total[kk] = {"in": total[kk]["in"] + b[kk]["in"], "out": total[kk]["out"] + b[kk]["out"],
                         "calls": total[kk]["calls"] + b[kk]["calls"]}
        total["emails"] += b["emails"]
    gt = total["total"]["in"] + total["total"]["out"]
    total["avg_tokens_per_email"] = round(gt / total["emails"], 1) if total["emails"] else 0.0
    return {"ok": True, "labels": labels, "modes": report, "total": total}


# ── /api/ai/stop-batch ──
@app.post("/api/ai/stop-batch")
def api_ai_stop_batch() -> dict[str, Any]:
    """Остановка батчевой обработки AI."""
    _AI_BATCH_STOP.set()
    return {"ok": True, "message": "Остановка AI..."}


# ── /api/ai/token-stats ──
@app.get("/api/ai/token-stats")
def api_token_stats_v2() -> dict[str, Any]:
    try:
        with connect() as con:
            return get_token_stats(con)
    except Exception as e:
        return {"error": str(e)}


# ── /api/imap/folders ──
@app.get("/api/imap/folders")
def api_imap_folders_v2() -> dict[str, Any]:
    try:
        from .imap_importer import list_imap_folders
        apply_runtime_settings()
        result = list_imap_folders()
        result["selected_folders"] = settings.imap_folders
        return result
    except Exception as e:
        return {"ok": False, "error": str(e), "folders": [], "items": []}


# ── /api/archive/stats ──
@app.get("/api/archive/stats")
def api_archive_stats_v2() -> dict[str, Any]:
    return get_archive_stats()


# ── /api/archive/cleanup ──
@app.post("/api/archive/cleanup")
def api_archive_cleanup_v2() -> dict[str, Any]:
    return run_archive_cleanup()


# ── /api/outbox/deliver (алиас для нового UI) ──
@app.post("/api/outbox/deliver")
def api_outbox_deliver_v2() -> dict[str, Any]:
    try:
        apply_runtime_settings()
        with connect() as con:
            result = deliver_outbox_events(con, limit=500)
        delivered = result.get("delivered", 0)
        return {"ok": True, "delivered": delivered, **result}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════
# НОВЫЕ API ДЛЯ СТАБИЛИЗАЦИИ UI v2.0
# ═══════════════════════════════════════════════

# ── Системный статус ──
@app.get("/api/v2/system/status")
def api_v2_system_status() -> dict[str, Any]:
    """Понятный статус сервера, почты, AI, импорта, outbox."""
    try:
        apply_runtime_settings()
        with connect() as con:
            total_emails = con.execute("SELECT COUNT(*) FROM raw_emails").fetchone()[0]
            total_cases = con.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
            outbox_new = con.execute("SELECT COUNT(*) FROM outbox WHERE status='new'").fetchone()[0]
            outbox_sent = con.execute("SELECT COUNT(*) FROM outbox WHERE status='sent'").fetchone()[0]
            outbox_error = con.execute("SELECT COUNT(*) FROM outbox WHERE status='error'").fetchone()[0]
            att_size = con.execute("SELECT COALESCE(SUM(size_bytes),0) FROM attachments").fetchone()[0]
        
        # Import status
        import_busy = bool(_IMPORT_LOCK.locked() or _IMPORT_BG_STATE.get("running"))
        last_import = _LAST_IMPORT_RESULT or {}
        
        # AI status
        ai_enabled = bool(getattr(settings, "enable_ai", False))
        ai_model = getattr(settings, "ai_model", "") or ""
        ai_configured = bool(getattr(settings, "ai_api_key", "")) and ai_enabled
        
        autopilot_thread_alive = _AUTOPILOT_THREAD and _AUTOPILOT_THREAD.is_alive() if hasattr(_AUTOPILOT_THREAD, 'is_alive') else False
        
        return {
            "ok": True,
            "server": "running",
            "autopilot_running": autopilot_thread_alive,
            "import": {
                "status": "running" if import_busy else ("done" if last_import.get("ok") else ("error" if last_import.get("error") else "idle")),
                "imported": last_import.get("imported", 0),
                "classified": last_import.get("classified", 0),
                "error": last_import.get("error"),
                "finished_at": last_import.get("finished_at"),
            },
            "ai": {
                "enabled": ai_enabled,
                "configured": ai_configured,
                "model": ai_model,
                "provider": getattr(settings, "ai_provider", ""),
                "busy": autopilot_thread_alive,
            },
            "mail": {
                "configured": bool(getattr(settings, "imap_username", "")),
                "host": getattr(settings, "imap_host", ""),
                "folders_count": len((getattr(settings, "imap_folders", "") or "").split(",")) if getattr(settings, "imap_folders", "") else 0,
            },
            "outbox": {
                "new": outbox_new,
                "sent": outbox_sent,
                "error": outbox_error,
            },
            "stats": {
                "total_emails": total_emails,
                "total_cases": total_cases,
                "attachments_mb": round(att_size / (1024 * 1024), 1),
            },
        }
    except Exception as e:
        return {"ok": False, "server": "error", "error": str(e)}


# ── Статус импорта (сводка БД; live heartbeat см. /api/v2/import/status) ──
@app.get("/api/v2/import/status-summary")
def api_v2_import_status() -> dict[str, Any]:
    """Детальный статус импорта с прогрессом + cumulative данные из БД."""
    try:
        import_busy = bool(_IMPORT_LOCK.locked() or _IMPORT_BG_STATE.get("running"))
        last_import = _LAST_IMPORT_RESULT or {}
        
        # Реальные cumulative данные из БД
        with connect() as con:
            total_emails = con.execute("SELECT COUNT(*) FROM raw_emails").fetchone()[0]
            total_cases = con.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
            classified_count = con.execute("SELECT COUNT(*) FROM cases WHERE claim_kind IS NOT NULL AND claim_kind != ''").fetchone()[0]
            att_size = con.execute("SELECT COALESCE(SUM(size_bytes),0) FROM attachments").fetchone()[0]
        
        return {
            "ok": True,
            "status": "running" if import_busy else ("done" if last_import.get("ok") else ("error" if last_import.get("error") else "idle")),
            "running": import_busy,
            "result": {
                "imported": last_import.get("imported", 0),
                "classified": last_import.get("classified", 0),
                "skipped": last_import.get("skipped", 0),
                "errors": last_import.get("errors", 0),
                "folders_processed": last_import.get("folders_processed", 0),
                "total_folders": last_import.get("total_folders", 0),
                "current_folder": last_import.get("current_folder", ""),
                "new_emails": last_import.get("new", 0),
                "attachments_mb": last_import.get("attachments_mb", 0),
                "started_at": last_import.get("started_at") or _IMPORT_BG_STATE.get("started_at"),
                "finished_at": last_import.get("finished_at"),
                "error": last_import.get("error"),
                "total_on_server": last_import.get("total_on_server", 0),
                "folders_detail": last_import.get("folders", []),
            },
            "db": {
                "total_emails": total_emails,
                "total_cases": total_cases,
                "classified": classified_count,
                "attachments_mb": round(att_size / (1024 * 1024), 1),
            },
        }
    except Exception as e:
        return {"ok": False, "status": "error", "error": str(e)}


# ── Остановить импорт ──
@app.post("/api/v2/import/stop")
def api_v2_import_stop() -> dict[str, Any]:
    """Остановка текущего импорта."""
    global _IMPORT_LOCK
    try:
        from .imap_importer import request_import_stop
        request_import_stop()
        _IMPORT_STOP.set()
        return {"ok": True, "message": "Остановка импорта..."}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Тест почтового подключения ──
@app.post("/api/v2/mail/test")
def api_v2_mail_test() -> dict[str, Any]:
    """Тест подключения к IMAP: проверка host/port/login/password, получение папок."""
    try:
        from .imap_importer import _open_imap, decode_imap_utf7
        apply_runtime_settings()
        
        # Пробуем открыть IMAP
        imap = _open_imap()
        
        # Получаем папки в формате (raw_name, display_name)
        _, folder_data = imap.list()
        folders_display = []
        folder_raw_map: dict[str, str] = {}   # display_name → raw_name
        folder_counts: dict[str, int] = {}
        for item in folder_data or []:
            if isinstance(item, bytes):
                parts = item.decode("utf-8", errors="replace").split('"/"')
                raw_name = parts[-1].strip().strip('"')
                display_name = decode_imap_utf7(raw_name)
                folders_display.append(display_name)
                folder_raw_map[display_name] = raw_name
                try:
                    # status() должен получать raw IMAP имя, а не декодированное
                    status = imap.status(raw_name, '(MESSAGES)')
                    if status and status[1] and status[1][0]:
                        count_match = re.search(r'MESSAGES\s+(\d+)', str(status[1][0]))
                        if count_match:
                            folder_counts[display_name] = int(count_match.group(1))
                except Exception:
                    pass
        
        imap.logout()
        
        selected = (getattr(settings, "imap_folders", "") or "").split(",")
        selected = [f.strip() for f in selected if f.strip()]
        selected_folders_counts = {}
        for sf in selected:
            display_sf = decode_imap_utf7(sf)
            if display_sf in folder_counts:
                selected_folders_counts[sf] = folder_counts[display_sf]
        
        return {
            "ok": True,
            "connection": "ok",
            "host": getattr(settings, "imap_host", ""),
            "user": getattr(settings, "imap_username", ""),
            "folders_found": len(folders_display),
            "folders": sorted(folders_display),
            "folders_with_counts": folder_counts,
            "selected_folders": selected,
            "selected_folders_count": len(selected),
            "selected_with_counts": selected_folders_counts,
            "capabilities": "ok",
        }
    except imaplib.IMAP4.error as e:
        return {"ok": False, "error": f"IMAP ошибка: {e}", "error_type": "auth"}
    except ConnectionRefusedError:
        return {"ok": False, "error": f"Сервер {getattr(settings,'imap_host','?')}:{getattr(settings,'imap_port','?')} недоступен", "error_type": "connection"}
    except Exception as e:
        return {"ok": False, "error": str(e), "error_type": "unknown"}


# ── Тест AI (аналогично btn-test-ai, но отдельный) ──
@app.post("/api/v2/ai/test")
def api_v2_ai_test() -> dict[str, Any]:
    """Тест AI: проверка ключа, base_url, списка моделей, выбранной модели, короткий запрос."""
    try:
        from .ai_client import test_ai_connection, list_ai_models
        apply_runtime_settings()
        
        result = {"ok": True, "checks": {}}
        
        # 1. Проверка ключа
        has_key = bool(getattr(settings, "ai_api_key", ""))
        result["checks"]["api_key"] = "ok" if has_key else "missing"
        
        # 2. Проверка base_url
        base_url = getattr(settings, "ai_base_url", "")
        result["checks"]["base_url"] = base_url if base_url else "not_set"
        
        # 3. Проверка провайдера
        provider = getattr(settings, "ai_provider", "")
        result["checks"]["provider"] = provider if provider else "not_set"
        
        # 4. Попытка получить список моделей
        try:
            models_result = list_ai_models()
            models = models_result if isinstance(models_result, list) else (models_result.get("models") or models_result.get("data") or [])
            result["checks"]["models_api"] = f"ok, {len(models)} models" if models else "no_models"
            result["models_count"] = len(models) if isinstance(models, list) else 0
            result["model_ids"] = [m.get("id", str(m)) for m in (models if isinstance(models, list) else [])][:50]
        except Exception as e:
            result["checks"]["models_api"] = f"error: {e}"
        
        # 5. Проверка выбранной модели
        model = getattr(settings, "ai_model", "")
        result["checks"]["selected_model"] = model if model else "not_selected"
        
        # 6. Короткий chat completion
        if has_key and base_url and model:
            try:
                test_res = test_ai_connection()
                result["checks"]["chat_completion"] = "ok" if test_res.get("ok") else str(test_res.get("error", "unknown"))
                if test_res.get("duration_ms"):
                    result["response_time_ms"] = test_res["duration_ms"]
            except Exception as e:
                result["checks"]["chat_completion"] = f"error: {e}"
        else:
            result["checks"]["chat_completion"] = "skipped (missing config)"
        
        all_ok = all(v == "ok" or v.startswith("ok") for v in result["checks"].values())
        result["ok"] = all_ok
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Статус обработки писем ──
@app.get("/api/v2/processing/status")
def api_v2_processing_status() -> dict[str, Any]:
    """Статус обработки писем AI."""
    try:
        with connect() as con:
            pending = con.execute("SELECT COUNT(*) FROM cases WHERE status IN ('needs_review','pending')").fetchone()[0]
            processed = con.execute("SELECT COUNT(*) FROM cases WHERE status IN ('ready_to_export','delivered','closed')").fetchone()[0]
            error_count = con.execute("SELECT COUNT(*) FROM ai_usage WHERE ok=0").fetchone()[0]
            ai_ok = con.execute("SELECT COUNT(*) FROM ai_usage WHERE ok=1").fetchone()[0]
        
        autopilot_running = _AUTOPILOT_THREAD and _AUTOPILOT_THREAD.is_alive() if hasattr(_AUTOPILOT_THREAD, 'is_alive') else False
        
        return {
            "ok": True,
            "pending": pending,
            "processed": processed,
            "ai_ok": ai_ok,
            "ai_errors": error_count,
            "autopilot_running": autopilot_running,
            "autopilot_enabled": _AUTOPILOT_STATE.get("enabled", False),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Outbox статус ──
@app.get("/api/v2/outbox/status")
def api_v2_outbox_status() -> dict[str, Any]:
    """Статус outbox/1С."""
    try:
        with connect() as con:
            rows = con.execute("""
                SELECT status, COUNT(*) cnt,
                       MAX(created_at) as last_created
                FROM outbox GROUP BY status
            """).fetchall()
        
        statuses = {r["status"]: {"count": r["cnt"], "last": r["last_created"]} for r in rows}
        
        latest_error = None
        try:
            with connect() as con:
                row = con.execute(
                    "SELECT status, error, created_at FROM outbox WHERE error IS NOT NULL ORDER BY id DESC LIMIT 1"
                ).fetchone()
                if row:
                    latest_error = {"status": row["status"], "error": str(row["error"] or "")[:200], "at": row["created_at"]}
        except Exception:
            pass
        
        return {
            "ok": True,
            "statuses": statuses,
            "total": sum(s["count"] for s in statuses.values()),
            "latest_error": latest_error,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── Получить настройки импорта (режим импорта, лимиты) ──
@app.get("/api/v2/import/settings")
def api_v2_import_settings() -> dict[str, Any]:
    """Настройки импорта писем."""
    try:
        return {
            "ok": True,
            "import_mode": getattr(settings, "import_mode", "new"),
            "import_search_query": getattr(settings, "import_search_query", ""),
            "import_limit_per_folder": getattr(settings, "imap_limit", 200),
            "import_total_limit": getattr(settings, "imap_total_limit", 2000),
            "import_max_attachment_mb": getattr(settings, "import_max_attachment_mb", 10),
            "import_download_attachments": getattr(settings, "import_download_attachments", True),
            "import_save_body": getattr(settings, "import_save_body", True),
            "import_skip_duplicates": getattr(settings, "import_skip_duplicates", True),
            "auto_import_enabled": getattr(settings, "auto_import_enabled", False),
            "auto_process_enabled": getattr(settings, "auto_process_enabled", False),
            "require_confirmation_before_case": getattr(settings, "require_confirmation_before_case", True),
            "require_confirmation_before_outbox": getattr(settings, "require_confirmation_before_outbox", True),
            "confidence_threshold": getattr(settings, "confidence_threshold", 0.85),
            "processing_mode": getattr(settings, "processing_mode", "manual"),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _run_sanity_pass() -> None:
    """Повторная санитарная проверка кейсов в ready_to_1c перед outbox."""
    from .classifier import sanity_check_fields
    with connect() as con:
        rows = con.execute("""
            SELECT id, fields_json FROM cases
            WHERE state = 'ready_to_1c'
              AND outbox_validated = 0
        """).fetchall()
        for row in rows:
            fields = loads(row["fields_json"] or "{}")
            cleaned, errors = sanity_check_fields(fields)
            if errors:
                con.execute("""
                    UPDATE cases SET state='needs_review',
                    quality_json=?, outbox_validated=1, updated_at=?
                    WHERE id=?
                """, (dumps([{"level": "error", "code": "sanity_fail", "message": e} for e in errors]),
                      utcnow(), row["id"]))
            else:
                con.execute("""
                    UPDATE cases SET fields_json=?, outbox_validated=1, updated_at=?
                    WHERE id=?
                """, (dumps(cleaned), utcnow(), row["id"]))


# ── /api/outbox/validate ──
@app.get("/api/outbox/validate")
def api_validate_outbox() -> dict[str, Any]:
    """Проверить все pending-записи outbox перед отправкой в 1С."""
    issues: list[dict[str, Any]] = []
    with connect() as con:
        rows = con.execute("""
            SELECT ob.id, ob.case_id, ob.payload_json
            FROM outbox ob
            WHERE ob.status IN ('new', 'pending')
        """).fetchall()
        for row in rows:
            payload = loads(row["payload_json"] or "{}")
            items = payload.get("items") or []
            row_issues: list[str] = []
            for item in items:
                if item.get("price") is not None and not isinstance(item["price"], (int, float)):
                    row_issues.append(f"price не число: {item['price']}")
                if item.get("quantity") is not None and not isinstance(item["quantity"], (int, float)):
                    row_issues.append(f"quantity не число: {item['quantity']}")
                document = payload.get("document", {}) or {}
                claim = payload.get("claim", {}) or {}
                if not item.get("part_number"):
                    row_issues.append("нет артикула")
                has_number = bool(
                    document.get("number")
                    or claim.get("number")
                    or claim.get("claim_number")
                    or claim.get("client_request_number")
                    or claim.get("return_number")
                )
                if not has_number:
                    row_issues.append("нет номера документа/обращения")
                if not document.get("date"):
                    row_issues.append("нет даты документа")
                if not claim.get("kind"):
                    row_issues.append("нет причины/типа претензии")
            if row_issues:
                issues.append({
                    "outbox_id": row["id"],
                    "case_id": row["case_id"],
                    "issues": row_issues,
                })
    return {"ok": True, "total_checked": len(rows), "issues_count": len(issues), "issues": issues}


# ── /api/v2/pipeline/run ──
@app.post("/api/v2/pipeline/run")
def api_pipeline_run() -> dict[str, Any]:
    """Полный пайплайн: импорт → паттерны → AI на unknown → quality gate → outbox."""
    apply_runtime_settings()
    results: dict[str, Any] = {"ok": True, "steps": {}}
    
    # Step 1: Import
    try:
        imp = _safe_import_cycle()
        results["steps"]["import"] = {"ok": imp.get("ok", True), "imported": imp.get("imported", 0), "errors": imp.get("errors", 0)}
    except Exception as e:
        results["steps"]["import"] = {"ok": False, "error": str(e)}
    
    # Step 2: Patterns
    try:
        buyer_rules = load_buyer_rules()
        pat_count = 0
        with connect() as con:
            rows = con.execute(
                "SELECT r.* FROM raw_emails r LEFT JOIN cases c ON c.raw_email_id=r.id WHERE c.id IS NULL ORDER BY r.id"
            ).fetchall()
            learned = load_buyer_identities(con)
            for row in rows:
                email_data = row_to_dict(row) or {}
                email_data["attachments"] = _load_attachments_with_text(con, int(row["id"]))
                email_data["visible_text"] = email_data.get("body_text") or email_data.get("snippet") or ""
                case_data = classify_email(email_data, buyer_rules, learned_identities=learned)
                case_id = save_case(con, int(row["id"]), case_data)
                case_data["export"]["case_id"] = case_id
                save_case(con, int(row["id"]), case_data)
                # v2.1 AI-only: наблюдение/промоция паттернов убраны.
                pat_count += 1
        results["steps"]["patterns"] = {"ok": True, "processed": pat_count}
    except Exception as e:
        results["steps"]["patterns"] = {"ok": False, "error": str(e)}
    
    # Step 3: AI on needs_review / unknown only
    try:
        ai_processed = 0
        ai_errors = 0
        with connect() as con:
            ai_candidates = con.execute(
                """
                SELECT c.id FROM cases c
                WHERE c.state IN ('needs_review','needs_link')
                  AND c.event_type IN ('new_return','unknown','correction_request','marking_request')
                  AND NOT EXISTS (
                    SELECT 1 FROM ai_suggestions s
                    WHERE s.case_id=c.id AND s.accepted=1
                  )
                ORDER BY c.id DESC
                LIMIT 50
                """
            ).fetchall()
        if getattr(settings, "enable_ai", False):
            for r in ai_candidates:
                try:
                    res = _apply_ai_to_case_id(r["id"], purpose="pipeline_ai")
                    if res.get("applied"):
                        ai_processed += 1
                    else:
                        ai_errors += 1
                except Exception:
                    ai_errors += 1
        results["steps"]["ai"] = {"ok": True, "enabled": bool(getattr(settings, "enable_ai", False)), "processed": ai_processed, "errors": ai_errors, "candidates": len(ai_candidates)}
    except Exception as e:
        results["steps"]["ai"] = {"ok": False, "error": str(e)}
    
    # Step 4: Quality gate + queue to outbox
    try:
        qc_passed = 0
        qc_failed = 0
        with connect() as con:
            ready_cases = con.execute(
                "SELECT c.id, c.ready_for_export, c.confidence, c.state, c.fields_json, c.quality_json FROM cases c WHERE c.state IN ('needs_review','needs_link','ready_to_1c') AND c.ready_for_export=1 AND c.confidence >= 0.8"
            ).fetchall()
            for c in ready_cases:
                fields = loads(c["fields_json"], {})
                quality = loads(c["quality_json"], [])
                # Quality check: price fields must be numeric
                has_price_error = False
                for k, v in fields.items():
                    if k in ("price", "sum", "amount", "quantity") and v is not None:
                        if isinstance(v, str) and not re.search(r'^\d+[.,]?\d*$', v.replace(" ", "")):
                            has_price_error = True
                if has_price_error or any(q.get("level") == "error" for q in quality):
                    qc_failed += 1
                    con.execute("UPDATE cases SET state='needs_review', updated_at=? WHERE id=?", (utcnow(), c["id"]))
                else:
                    qc_passed += 1
                    queue_case_event(con, c["id"])
        results["steps"]["quality_gate"] = {"ok": True, "passed": qc_passed, "failed": qc_failed}
    except Exception as e:
        results["steps"]["quality_gate"] = {"ok": False, "error": str(e)}
    
    # Step 5: Auto-deliver outbox if enabled
    try:
        if getattr(settings, "auto_deliver_outbox", False):
            with connect() as con:
                delivery = deliver_outbox_events(con, limit=500)
            results["steps"]["delivery"] = {"ok": True, "delivered": delivery.get("delivered", 0)}
        else:
            results["steps"]["delivery"] = {"ok": True, "skipped": True}
    except Exception as e:
        results["steps"]["delivery"] = {"ok": False, "error": str(e)}
    
    results["ok"] = all(s.get("ok", False) for s in results["steps"].values())
    return results


# ── /api/v2/pipeline/pipeline-status ──
@app.get("/api/v2/pipeline/status")
def api_pipeline_status() -> dict[str, Any]:
    """Статус пайплайна: счётчики по каждому этапу обработки."""
    try:
        apply_runtime_settings()
        with connect() as con:
            total_emails = con.execute("SELECT COUNT(*) c FROM raw_emails").fetchone()["c"]

            # Этап 1: письма без кейса — ещё не обработаны паттернами
            needs_pattern = con.execute(
                "SELECT COUNT(*) c FROM raw_emails r LEFT JOIN cases c ON c.raw_email_id=r.id WHERE c.id IS NULL"
            ).fetchone()["c"]

            # Этап 2: обработаны паттернами, готовы к 1С (без AI)
            pattern_ready = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE state='ready_to_1c' AND ready_for_export=1
                   AND NOT EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=cases.id AND s.accepted=1)"""
            ).fetchone()["c"]

            # Этап 3: требуют AI (паттерны не справились)
            needs_ai = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE state IN ('needs_review','needs_link')
                   AND event_type IN ('new_return','unknown')
                   AND (needs_ai = 1 OR event_type = 'unknown')
                   AND NOT EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=cases.id AND s.accepted=1)"""
            ).fetchone()["c"]

            # Этап 3b: AI применён, готово к 1С
            ai_ready = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE state='ready_to_1c' AND ready_for_export=1
                   AND EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=cases.id AND s.accepted=1)"""
            ).fetchone()["c"]

            review_count = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE event_type='new_return'
                     AND state IN ('ready_to_1c','needs_review')
                     AND state NOT IN ('ignored_internal','context_sent','ignored_info_only','ignored_spam_promo')"""
            ).fetchone()["c"]

            # Этап 4: полностью неразобранные (ни паттерны, ни AI не дали результата)
            unprocessed = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE state IN ('needs_review','needs_link')
                   AND event_type IN ('unknown','')
                   AND NOT EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=cases.id AND s.accepted=1)"""
            ).fetchone()["c"]

            # Счётчик для ВКЛАДКИ «Неразобранные» — совпадает с by-method=unprocessed
            # (всё в needs_review/unknown, кроме followup/решений; вкл. уже тронутые AI).
            unprocessed_tab = con.execute(
                """SELECT COUNT(*) c FROM cases c
                   WHERE c.state IN ('needs_review','unknown')
                     AND c.event_type NOT IN ('followup_reminder','followup_dialog','supplier_decision')
                     AND (
                        EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=c.id)
                        OR ((c.buyer_code IS NULL OR c.buyer_code='') AND (c.claim_kind IS NULL OR c.claim_kind=''))
                        OR c.event_type='unknown'
                     )"""
            ).fetchone()["c"]

            links_count = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE (state='needs_link' OR link_quarantine=1)
                     AND event_type NOT IN ('info_only','spam_promo','unknown')"""
            ).fetchone()["c"]

            offtopic = con.execute(
                """SELECT COUNT(*) c FROM cases
                   WHERE state IN ('ignored_info_only','ignored_spam_promo')
                      OR event_type IN ('info_only','spam_promo')"""
            ).fetchone()["c"]

            # Outbox
            in_outbox = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='new'").fetchone()["c"]
            sent_outbox = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='sent'").fetchone()["c"]
            err_outbox = con.execute("SELECT COUNT(*) c FROM outbox WHERE status='error'").fetchone()["c"]

            case_states = {r["state"]: r["c"] for r in con.execute(
                "SELECT COALESCE(state,'?') state, COUNT(*) c FROM cases GROUP BY state"
            ).fetchall()}

            patterns_total = int(_PATTERNS_STATE.get("total") or 0)
            patterns_processed = int(_PATTERNS_STATE.get("processed") or 0)
            patterns_busy = bool(
                _PATTERNS_STATE.get("running")
                and (patterns_total == 0 or patterns_processed < patterns_total)
            )

            return {
                "ok": True,
                "total_emails": total_emails,
                "needs_pattern": needs_pattern,
                "pattern_ready": pattern_ready,
                "needs_ai": needs_ai,
                "ai_ready": ai_ready,
                "links_count": links_count,
                "unprocessed": unprocessed,
                "unprocessed_tab": unprocessed_tab,
                "processed_hidden": _processed_hidden_count(con),
                "offtopic": offtopic,
                "review_count": review_count,
                "case_states": case_states,
                "ready_to_1c": pattern_ready + ai_ready,
                "outbox_new": in_outbox,
                "outbox_sent": sent_outbox,
                "outbox_errors": err_outbox,
                "total_cases": con.execute("SELECT COUNT(*) c FROM cases").fetchone()["c"],
                "import_busy": bool(_IMPORT_BG_STATE.get("running") or _IMPORT_LOCK.locked()),
                "patterns_busy": patterns_busy,
                "ai_busy": _AI_BATCH_LOCK.locked(),
                "imap_configured": bool(getattr(settings, "imap_username", None)),
                "ai_configured": getattr(settings, "enable_ai", False),
                "autopilot_running": _AUTOPILOT_THREAD is not None and _AUTOPILOT_THREAD.is_alive(),
            }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ── /api/v2/pipeline/reset ──
@app.post("/api/v2/pipeline/reset")
def api_pipeline_reset() -> dict[str, Any]:
    """Сброс обработанных данных (письма на почте не трогаем, только локальную БД)."""
    try:
        with connect() as con:
            result = reset_processing_data(con, keep_settings=True, keep_learning=True, keep_process_events=False)
            record_process_event(con, stage="operator", level="warn", message="Локальные данные сброшены для переимпорта", details=result)
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/v2/pipeline/reset-work")
def api_pipeline_reset_work() -> dict[str, Any]:
    """Сбросить только результаты обработки, сохранив загруженные письма и вложения."""
    try:
        if _IMPORT_LOCK.locked() or _IMPORT_BG_STATE.get("running"):
            return {"ok": False, "error": "import_running", "message": "Сначала остановите или дождитесь завершения импорта."}
        if _PATTERNS_STATE.get("running") or _AI_BATCH_STATE.get("running"):
            return {"ok": False, "error": "processing_running", "message": "Сначала остановите паттерны/AI."}
        with connect() as con:
            result = reset_processed_work_data(con, keep_process_events=False)
            record_process_event(con, stage="operator", level="warn", message="Обработка обнулена, письма сохранены", details=result)
        return result
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════
# КЛИЕНТЫ — v2.1
# ═══════════════════════════════════════════════

@app.get("/api/v2/clients")
def api_v2_clients() -> dict[str, Any]:
    """Список клиентов из YAML-конфигов + статистика из БД."""
    try:
        rules = load_buyer_rules()
        with connect() as con:
            # Статистика по клиентам
            stats_rows = con.execute("""
                SELECT buyer_code,
                       COUNT(*) total,
                       SUM(CASE WHEN state='ready_to_1c' THEN 1 ELSE 0 END) ready,
                       SUM(CASE WHEN state='needs_review' THEN 1 ELSE 0 END) review,
                       SUM(CASE WHEN state IN ('delivered','exported','closed') THEN 1 ELSE 0 END) done,
                       MAX(c.updated_at) last_at
                FROM cases c
                GROUP BY buyer_code
            """).fetchall()
            stats_by_code = {r["buyer_code"]: dict(r) for r in stats_rows}

            # AI применён?
            ai_rows = con.execute("""
                SELECT c.buyer_code, COUNT(*) ai_applied
                FROM cases c
                JOIN ai_suggestions s ON s.case_id=c.id AND s.accepted=1
                GROUP BY c.buyer_code
            """).fetchall()
            ai_by_code = {r["buyer_code"]: r["ai_applied"] for r in ai_rows}

        clients = []
        for r in rules:
            s = stats_by_code.get(r.code, {})
            # Срок возврата из конфига
            deadline_days = getattr(r, "return_deadline_days", None)
            if deadline_days is None:
                # Попробуем прочитать из YAML напрямую
                try:
                    cfg_path = settings.buyer_config_dir / f"{r.code}.yml"
                    if cfg_path.exists():
                        import yaml as _yaml
                        raw = _yaml.safe_load(cfg_path.read_text(encoding="utf-8"))
                        deadline_days = (raw.get("buyer") or {}).get("return_deadline_days", 45)
                except Exception:
                    deadline_days = 45
            clients.append({
                "code": r.code,
                "name": r.name,
                "enabled": getattr(r, "enabled", True),
                "domains": r.domains[:5],
                "senders": r.senders[:5],
                "folders": getattr(r, "folders", [])[:5],
                "return_deadline_days": deadline_days or 45,
                "stats": {
                    "total": s.get("total", 0),
                    "ready": s.get("ready", 0),
                    "review": s.get("review", 0),
                    "done": s.get("done", 0),
                    "ai_applied": ai_by_code.get(r.code, 0),
                    "last_at": s.get("last_at"),
                },
            })

        # Клиенты из БД но не из YAML (неизвестные)
        known_codes = {r.code for r in rules}
        for code, s in stats_by_code.items():
            if code and code not in known_codes:
                clients.append({
                    "code": code,
                    "name": code,
                    "enabled": False,
                    "domains": [],
                    "senders": [],
                    "folders": [],
                    "return_deadline_days": 45,
                    "unknown": True,
                    "stats": {
                        "total": s.get("total", 0),
                        "ready": s.get("ready", 0),
                        "review": s.get("review", 0),
                        "done": s.get("done", 0),
                        "ai_applied": ai_by_code.get(code, 0),
                        "last_at": s.get("last_at"),
                    },
                })
        clients.sort(key=lambda x: (-x["stats"]["total"], x["name"].lower()))
        return {"ok": True, "count": len(clients), "items": clients}
    except Exception as e:
        return {"ok": False, "error": str(e), "items": []}


@app.patch("/api/v2/clients/{code}/deadline")
def api_v2_client_deadline(code: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Обновить срок возврата клиента в YAML."""
    days = payload.get("return_deadline_days")
    if not isinstance(days, int) or days < 1 or days > 3650:
        raise HTTPException(400, "return_deadline_days must be int 1..3650")
    cfg_path = settings.buyer_config_dir / f"{code}.yml"
    if not cfg_path.exists():
        raise HTTPException(404, f"Buyer config '{code}' not found")
    try:
        import yaml as _yaml
        raw = _yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
        if "buyer" not in raw:
            raw["buyer"] = {}
        raw["buyer"]["return_deadline_days"] = days
        cfg_path.write_text(_yaml.dump(raw, allow_unicode=True, sort_keys=False), encoding="utf-8")
        _log("config", f"Срок возврата клиента '{code}' обновлён: {days} дней", level="ok")
        return {"ok": True, "code": code, "return_deadline_days": days}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.get("/api/v2/cases/by-method")
def api_v2_cases_by_method(
    method: str = "pattern",
    limit: int = 100,
    page: int = 1,
    q: str = "",
) -> dict[str, Any]:
    """Кейсы разбитые по этапу обработки: pattern / ai / unprocessed / offtopic.

    pattern     — возвраты на стадии паттернов (ещё не трогал AI): и готовые, и проблемные.
                  Это рабочая поверхность для проверки и обучения.
    ai          — кейсы, которые прошли через AI (есть любое AI-предложение).
    unprocessed — совсем непонятные: нет ни клиента, ни типа претензии.
    """
    try:
        offset = (page - 1) * limit
        params: list[Any] = []
        with connect() as con:
            if method == "pattern":
                # ЭТАП 1: паттерны СПРАВИЛИСЬ — кейс готов, AI не трогал.
                # Отсюда кейсы идут в Сверку. Тут можно проверить и поправить.
                where = """
                    c.event_type = 'new_return'
                    AND c.state = 'ready_to_1c'
                    AND NOT EXISTS (
                        SELECT 1 FROM ai_suggestions s WHERE s.case_id=c.id
                    )
                """
            elif method == "ai":
                # ЭТАП 2: паттерны НЕ справились, AI ещё не запускался.
                # После AI: либо готово → Сверка, либо не смог → Неразобранные.
                where = """
                    c.state IN ('needs_review', 'needs_link')
                    AND c.event_type IN ('new_return','unknown')
                    AND (c.needs_ai = 1 OR c.event_type = 'unknown')
                    AND NOT EXISTS (
                        SELECT 1 FROM ai_suggestions s WHERE s.case_id=c.id
                    )
                """
            elif method == "links":
                # Письма, которые нужно связать с исходным кейсом:
                # корректировки, готово к выдаче, напоминания и продолжения диалога.
                where = """
                    (
                        c.state = 'needs_link'
                        OR c.link_quarantine = 1
                    )
                    AND c.event_type NOT IN ('info_only','spam_promo','unknown')
                """
            elif method == "unprocessed":
                # ЭТАП 3: AI попробовал и НЕ смог довести до готовности.
                # Тут оператор обучает вручную. Плюс совсем пустые письма.
                # ВАЖНО: диалоги/продолжения (followup_*) и решения поставщика — НЕ сюда,
                # они идут в свой поток связывания, а не в «неразобранные».
                where = """
                    c.state IN ('needs_review', 'unknown')
                    AND c.event_type NOT IN ('followup_reminder','followup_dialog','supplier_decision')
                    AND (
                        EXISTS (SELECT 1 FROM ai_suggestions s WHERE s.case_id=c.id)
                        OR (
                            (c.buyer_code IS NULL OR c.buyer_code = '')
                            AND (c.claim_kind IS NULL OR c.claim_kind = '')
                        )
                        OR c.event_type = 'unknown'
                    )
                """
            elif method == "problem_notice":
                # Уведомления о проблеме (принят с дефектом, не запрос на возврат).
                where = "(c.event_type='problem_notice' OR c.state='problem_notice')"
            else:
                # ЭТАП 4: письма не по задаче возвратов/претензий.
                # Их можно открыть и вручную переобучить/переназначить, если классификация ошиблась.
                where = """
                    (
                        c.state IN ('ignored_info_only','ignored_spam_promo')
                        OR c.event_type IN ('info_only','spam_promo')
                    )
                """

            if q:
                where += " AND (e.subject LIKE ? OR c.fields_json LIKE ? OR e.from_addr LIKE ?)"
                params.extend([f"%{q}%", f"%{q}%", f"%{q}%"])

            total = con.execute(
                f"SELECT COUNT(*) c FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id WHERE {where}",
                params,
            ).fetchone()["c"]

            rows = con.execute(
                f"""
                SELECT c.id, c.buyer_code, c.buyer_name, c.event_type, c.claim_kind,
                       c.state, c.priority, c.confidence, c.deadline_at,
                       c.fields_json, c.missing_json, c.quality_json, c.thread_key,
                       c.raw_email_id,
                       e.from_addr, e.subject, e.received_at, e.snippet,
                       e.body_text, e.body_html,
                       (SELECT COUNT(*) FROM attachments a WHERE a.raw_email_id=e.id) has_att
                FROM cases c JOIN raw_emails e ON e.id=c.raw_email_id
                WHERE {where}
                ORDER BY CASE c.priority WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
                         e.received_at DESC
                LIMIT ? OFFSET ?
                """,
                params + [limit, offset],
            ).fetchall()

            cases = []
            for r in rows:
                d = row_to_dict(r) or {}
                cases.append(d)

        return {"ok": True, "method": method, "total": total, "page": page, "limit": limit, "cases": cases}
    except Exception as e:
        return {"ok": False, "error": str(e), "cases": [], "total": 0}


@app.get("/api/v2/thread/{thread_key}")
def api_v2_thread(thread_key: str) -> dict[str, Any]:
    """Все письма одного треда (диалога) по thread_key."""
    try:
        with connect() as con:
            rows = con.execute(
                """
                SELECT c.id, c.event_type, c.claim_kind, c.state, c.priority,
                       c.is_followup, c.confidence, c.fields_json, c.thread_key,
                       e.subject, e.from_addr, e.received_at, e.snippet, e.body_text
                FROM cases c
                JOIN raw_emails e ON e.id = c.raw_email_id
                WHERE c.thread_key = ?
                ORDER BY e.received_at ASC
                """,
                (thread_key,)
            ).fetchall()
        items = [row_to_dict(r) or {} for r in rows]
        followups = [i for i in items if i.get("is_followup") or i.get("event_type") in
                     ("followup_reminder", "followup_dialog", "supplier_decision")]
        return {"ok": True, "thread_key": thread_key, "count": len(items),
                "followups": len(followups), "items": items}
    except Exception as e:
        return {"ok": False, "error": str(e), "items": []}


# ═══════════════════════════════════════════════
# САМООБУЧЕНИЕ ПАТТЕРНОВ — API
# ═══════════════════════════════════════════════

@app.get("/api/v2/learning/pattern-candidates")
def api_v2_pattern_candidates(
    buyer_code: str = "",
    min_seen: int = Query(default=1, ge=1),
    limit: int = Query(default=200, ge=1, le=1000),
) -> dict[str, Any]:
    """Список кандидатов на паттерны, накопленных из AI-результатов."""
    try:
        with connect() as con:
            conditions = []
            params: list[Any] = []
            if buyer_code:
                conditions.append("buyer_code=?")
                params.append(buyer_code)
            if min_seen > 1:
                conditions.append("seen_count >= ?")
                params.append(min_seen)
            where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
            rows = con.execute(
                f"""
                SELECT id, buyer_code, field_name, value_sample, pattern_regex,
                       context_before, context_after, source, confidence,
                       seen_count, accepted, rejected, promoted_at, updated_at
                FROM field_pattern_candidates
                {where}
                ORDER BY confidence DESC, seen_count DESC, updated_at DESC
                LIMIT ?
                """,
                params + [limit],
            ).fetchall()
            total = con.execute(
                f"SELECT COUNT(*) c FROM field_pattern_candidates {where}", params
            ).fetchone()["c"]

            # Группируем по клиенту
            by_buyer: dict[str, list[dict]] = {}
            for r in rows:
                d = dict(r)
                bc = d.get("buyer_code") or "unknown"
                by_buyer.setdefault(bc, []).append(d)

        return {
            "ok": True, "total": total, "limit": limit,
            "by_buyer": by_buyer,
            "items": [dict(r) for r in rows],
        }
    except Exception as e:
        return {"ok": False, "error": str(e), "items": []}


# v2.1 AI-only: v2-эндпоинты обучения (promote / promote-all / reject-candidate) удалены.


# ── Очистка outbox от некорректно попавших записей ──
@app.post("/api/v2/outbox/cleanup-non-ready")
def api_v2_outbox_cleanup() -> dict[str, Any]:
    """Удалить из outbox записи, кейсы которых НЕ готовы к 1С.

    Используется однократно после исправления логики пайплайна.
    Удаляет только 'new' и 'error' записи — отправленные ('sent') не трогает.
    """
    try:
        with connect() as con:
            rows = con.execute("""
                SELECT o.id, o.case_id, c.state, c.ready_for_export
                FROM outbox o
                LEFT JOIN cases c ON c.id = o.case_id
                WHERE o.status IN ('new', 'error')
                  AND (
                    c.id IS NULL
                    OR c.state != 'ready_to_1c'
                    OR COALESCE(c.ready_for_export, 0) = 0
                  )
            """).fetchall()
            ids = [r["id"] for r in rows]
            if ids:
                con.execute(f"DELETE FROM outbox WHERE id IN ({','.join('?' * len(ids))})", ids)
            _log("operator", f"Очистка outbox: удалено {len(ids)} некорректных записей", level="ok",
                 details={"deleted_ids": ids[:50]})
        return {"ok": True, "deleted": len(ids),
                "message": f"Удалено {len(ids)} записей без готовых кейсов"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/v2/outbox/cleanup-empty-fields")
def api_v2_outbox_cleanup_empty() -> dict[str, Any]:
    """Убрать из outbox записи, у кейсов которых нет минимальных полей (артикул + номер + дата).

    Такие записи перемещаются обратно в needs_review для доработки паттернами/AI.
    """
    try:
        deleted = 0
        demoted = 0
        with connect() as con:
            rows = con.execute("""
                SELECT o.id as outbox_id, o.case_id,
                       c.fields_json, c.state
                FROM outbox o
                JOIN cases c ON c.id = o.case_id
                WHERE o.status IN ('new', 'error')
            """).fetchall()

            for r in rows:
                fields = loads(r["fields_json"] or "{}", {})
                has_doc = bool(
                    fields.get("document_number") or
                    fields.get("claim_number") or
                    fields.get("return_number")
                )
                has_part = bool(fields.get("part_number"))
                has_date = bool(fields.get("document_date"))

                if not (has_doc and has_part):
                    # Удаляем запись из outbox
                    con.execute("DELETE FROM outbox WHERE id=?", (r["outbox_id"],))
                    deleted += 1
                    # Кейс возвращаем в needs_review
                    if r["state"] in ("ready_to_1c", "ready_to_export"):
                        con.execute(
                            "UPDATE cases SET state='needs_review', ready_for_export=0, needs_review=1, updated_at=? WHERE id=?",
                            (utcnow(), r["case_id"])
                        )
                        demoted += 1

        _log("operator",
             f"Очистка 1С: удалено {deleted} записей без min-полей, возвращено в review: {demoted}",
             level="ok")
        return {
            "ok": True,
            "deleted_outbox": deleted,
            "demoted_to_review": demoted,
            "message": f"Удалено {deleted} пустых записей 1С, {demoted} кейсов → на доработку",
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}


# ═══════════════════════════════════════════════
# ВЛОЖЕНИЯ — скачать, распарсить Excel, Vision AI
# ═══════════════════════════════════════════════

@app.get("/api/attachments/{att_id}/download")
def api_attachment_download(att_id: int):
    """Скачать файл вложения."""
    from fastapi.responses import FileResponse as _FileResponse
    with connect() as con:
        row = con.execute(
            "SELECT filename, content_type, file_path FROM attachments WHERE id=?", (att_id,)
        ).fetchone()
    if not row:
        raise HTTPException(404, "Attachment not found")
    fp = row["file_path"]
    if not fp or not Path(fp).exists():
        raise HTTPException(404, "File not saved on disk (imported before v2.1)")
    return _FileResponse(
        path=fp,
        filename=row["filename"] or "attachment",
        media_type=row["content_type"] or "application/octet-stream",
    )


@app.get("/api/attachments/{att_id}/preview")
def api_attachment_preview(att_id: int) -> dict[str, Any]:
    """Распарсить Excel/CSV вложение и вернуть таблицу для отображения в UI."""
    with connect() as con:
        row = con.execute(
            "SELECT filename, content_type, file_path, size_bytes FROM attachments WHERE id=?", (att_id,)
        ).fetchone()
    if not row:
        raise HTTPException(404, "Attachment not found")
    fp = row["file_path"]
    filename = (row["filename"] or "").lower()
    if not fp or not Path(fp).exists():
        return {"ok": False, "error": "Файл не сохранён (импорт до v2.1 — переимпортируйте письмо)"}

    ext = filename.rsplit(".", 1)[-1] if "." in filename else ""
    try:
        if ext in {"xlsx", "xls", "xlsm"}:
            import openpyxl
            wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
            sheets = []
            for ws in list(wb.worksheets)[:3]:  # первые 3 листа
                rows = []
                for i, row_data in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 200:
                        break
                    rows.append([str(c) if c is not None else "" for c in row_data])
                sheets.append({"name": ws.title, "rows": rows})
            return {"ok": True, "type": "excel", "sheets": sheets, "filename": row["filename"]}

        elif ext == "csv":
            import csv, io
            text = Path(fp).read_text(encoding="utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = [r for i, r in enumerate(reader) if i < 200]
            return {"ok": True, "type": "csv", "sheets": [{"name": "CSV", "rows": rows}], "filename": row["filename"]}

        elif ext == "pdf":
            return {"ok": False, "error": "PDF — используйте Vision AI для анализа", "use_vision": True}

        elif ext in {"jpg", "jpeg", "png", "gif", "webp"}:
            return {"ok": False, "error": "Изображение — используйте Vision AI", "use_vision": True,
                    "download_url": f"/api/attachments/{att_id}/download"}

        elif ext == "zip":
            # «Заглянуть внутрь»: список файлов + парсинг первого Excel/CSV (внутри ПИТСТОП-актов).
            import zipfile, io
            entries = []
            sheets = []
            with zipfile.ZipFile(fp) as zf:
                for zi in zf.infolist():
                    if zi.is_dir():
                        continue
                    entries.append({"name": zi.filename, "size": zi.file_size})
                # первый Excel/CSV/TXT внутри — распарсим в таблицу/строки
                inner = next((zi for zi in zf.infolist()
                              if not zi.is_dir() and zi.filename.lower().rsplit(".", 1)[-1] in {"xlsx", "xls", "xlsm", "csv", "txt"}), None)
                if inner is not None:
                    iext = inner.filename.lower().rsplit(".", 1)[-1]
                    raw = zf.read(inner.filename)
                    def _dec(b: bytes) -> str:
                        # Русские прайс-листы ПИТСТОП обычно в windows-1251, не utf-8.
                        for enc in ("utf-8-sig", "cp1251", "latin-1"):
                            try:
                                return b.decode(enc)
                            except Exception:
                                continue
                        return b.decode("utf-8", errors="replace")
                    if iext == "txt":
                        text = _dec(raw)
                        # Прайс-лист — таблица с табами; режем по \t для читаемых колонок.
                        rows = [(ln.split("\t") if "\t" in ln else [ln]) for ln in text.splitlines()[:120]]
                        sheets.append({"name": inner.filename, "rows": rows})
                    elif iext == "csv":
                        text = _dec(raw)
                        import csv as _csv
                        rows = [r for i, r in enumerate(_csv.reader(io.StringIO(text))) if i < 200]
                        sheets.append({"name": inner.filename, "rows": rows})
                    else:
                        import openpyxl
                        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
                        for ws in list(wb.worksheets)[:2]:
                            rows = []
                            for i, rd in enumerate(ws.iter_rows(values_only=True)):
                                if i >= 200:
                                    break
                                rows.append([str(c) if c is not None else "" for c in rd])
                            sheets.append({"name": f"{inner.filename}:{ws.title}", "rows": rows})
            return {"ok": True, "type": "zip", "entries": entries, "sheets": sheets,
                    "filename": row["filename"], "download_url": f"/api/attachments/{att_id}/download"}

        else:
            return {"ok": False, "error": f"Неподдерживаемый формат: {ext}"}
    except ImportError:
        return {"ok": False, "error": "openpyxl не установлен: pip install openpyxl"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/attachments/{att_id}/vision")
def api_attachment_vision(att_id: int) -> dict[str, Any]:
    """Запустить Vision AI на изображении или PDF-вложении."""
    from .ai_client import run_vision_extraction
    with connect() as con:
        row = con.execute(
            """SELECT a.filename, a.content_type, a.file_path,
                      e.body_text, e.snippet, c.buyer_code
               FROM attachments a
               JOIN raw_emails e ON e.id = a.raw_email_id
               LEFT JOIN cases c ON c.raw_email_id = e.id
               WHERE a.id=?""",
            (att_id,)
        ).fetchone()
    if not row:
        raise HTTPException(404, "Attachment not found")
    fp = row["file_path"]
    if not fp or not Path(fp).exists():
        return {"ok": False, "error": "Файл не сохранён на диске"}
    try:
        raw_bytes = Path(fp).read_bytes()
        ctype = row["content_type"] or "image/jpeg"
        hint = row["body_text"] or row["snippet"] or ""
        result = run_vision_extraction(raw_bytes, content_type=ctype, hint_text=hint)
        return {**result, "attachment_id": att_id, "filename": row["filename"]}
    except Exception as e:
        return {"ok": False, "error": str(e)}


# Промт классификации документа брака (3 документа + фото детали). «Хороший промт» под доки.
DEFECT_DOC_PROMPT = (
    "Это вложение из претензии по БРАКУ автозапчасти. Определи ТИП документа на изображении. "
    "Ответь ТОЛЬКО валидным JSON без markdown: "
    '{"doc_type": "install_order|removal_order|service_act|part_photo|other", '
    '"confidence": 0.0, "reason": "кратко что видно"}. '
    "install_order = заказ-наряд/акт на УСТАНОВКУ детали; "
    "removal_order = заказ-наряд/акт на СНЯТИЕ или демонтаж детали; "
    "service_act = акт или заключение сервиса о неисправности, дефектовка, диагностика, экспертиза; "
    "part_photo = фотография самой детали (НЕ документ); other = иное. "
    "Если на фото скан документа — классифицируй по ТИПУ документа, не part_photo."
)


@app.post("/api/cases/{case_id}/check-defect-docs")
def api_check_defect_docs(case_id: int) -> dict[str, Any]:
    """ИИ-прогон файлов брака: сначала ДОКУМЕНТЫ (PDF/doc), потом ФОТО (если документов не все).
    Каждый файл классифицируется vision-моделью на 3 типа документа брака; стоп, как только
    набраны все 3. Результат пишется в payload.defect_doc_flag (mode=ai)."""
    from .ai_client import run_vision_extraction
    with connect() as con:
        case = con.execute("SELECT id, raw_email_id, claim_kind, payload_json FROM cases WHERE id=?", (case_id,)).fetchone()
        if not case:
            raise HTTPException(404, "Case not found")
        atts = con.execute(
            "SELECT id, filename, content_type, file_path FROM attachments WHERE raw_email_id=? ORDER BY id",
            (case["raw_email_id"],),
        ).fetchall()
    if not atts:
        return {"ok": True, "state": "absent", "found": {}, "n_present": 0, "attachments": [], "note": "вложений нет"}

    # Текст-извлекаемые документы (PDF/Excel/csv/txt/архив) — классифицируем по тексту, без vision.
    _TEXTDOC_RE = re.compile(r"\.(pdf|docx?|xlsx?|xlsm|csv|txt|zip|rar|7z)$", re.I)

    def _is_doc(a: Any) -> bool:
        ct = (a["content_type"] or "").lower()
        fn = (a["filename"] or "").lower()
        return "pdf" in ct or bool(_TEXTDOC_RE.search(fn))

    def _is_pdf(a: Any) -> bool:
        return "pdf" in (a["content_type"] or "").lower() or (a["filename"] or "").lower().endswith(".pdf")

    docs = [a for a in atts if _is_doc(a)]
    photos = [a for a in atts if not _is_doc(a)]
    found: dict[str, Any] = {"install_order": None, "removal_order": None, "service_act": None}
    per_att: list[dict[str, Any]] = []

    def _classify_doc_text(text: str) -> str:
        """Классификация документа по ИЗВЛЕЧЁННОМУ ТЕКСТУ (быстро, без vision)."""
        t = (text or "").lower()
        if not t.strip():
            return ""
        if ("наряд" in t or "заказ" in t) and ("сняти" in t or "демонтаж" in t):
            return "removal_order"
        if ("наряд" in t or "заказ" in t) and ("установк" in t or "монтаж" in t):
            return "install_order"
        if any(w in t for w in ("заключени", "дефектовк", "диагност", "экспертиз")) or ("акт" in t and "сервис" in t):
            return "service_act"
        if "акт" in t:
            return "service_act"
        return ""

    def _doc_text(a: Any) -> str:
        """Текст документа: PDF→pdfminer, Excel/csv/zip/архив→_extract_attachment_text."""
        fp = a["file_path"]
        if _is_pdf(a):
            try:
                from .email_parser import _extract_pdf_text
                return _extract_pdf_text(Path(fp).read_bytes()) or ""
            except Exception:
                return ""
        try:
            return _extract_attachment_text(fp, a["filename"]) or ""
        except Exception:
            return ""

    def _process(a: Any, group: str) -> None:
        fp = a["file_path"]
        if not fp or not Path(fp).exists():
            per_att.append({"id": a["id"], "filename": a["filename"], "group": group, "error": "файл не на диске"})
            return
        # 1) Текст-документы (PDF/Excel/zip/csv) — классифицируем по тексту, без vision.
        if _is_doc(a):
            txt = _doc_text(a)
            dt = _classify_doc_text(txt)
            if dt:
                per_att.append({"id": a["id"], "filename": a["filename"], "group": group,
                                "doc_type": dt, "reason": "текст вложения", "via": "text"})
                if dt in found and not found[dt]:
                    found[dt] = a["id"]
                return
            # Excel/csv/архив без признаков документа → дальше vision НЕ зовём (это не картинка).
            if not _is_pdf(a):
                per_att.append({"id": a["id"], "filename": a["filename"], "group": group,
                                "doc_type": "", "reason": "текст без признаков документа", "via": "text"})
                return
            # Скан-PDF без текстового слоя → проваливаемся в vision ниже.
        # 2) Картинки и скан-PDF → vision.
        try:
            # Учёт токенов vision пишется внутри run_vision_extraction (kind=vision, case_id).
            res = run_vision_extraction(Path(fp).read_bytes(), content_type=(a["content_type"] or "image/png"),
                                        prompt_text=DEFECT_DOC_PROMPT, case_id=case_id)
            resp = res.get("response") if isinstance(res, dict) else None
            if isinstance(resp, str):
                try:
                    resp = json.loads(resp)
                except Exception:
                    resp = {}
            resp = resp or {}
            dt = str(resp.get("doc_type") or "").strip()
            per_att.append({"id": a["id"], "filename": a["filename"], "group": group,
                            "doc_type": dt, "reason": resp.get("reason"), "confidence": resp.get("confidence"),
                            "skipped": res.get("skipped"), "error": res.get("error"), "via": "vision"})
            if dt in found and not found[dt]:
                found[dt] = a["id"]
        except Exception as e:
            per_att.append({"id": a["id"], "filename": a["filename"], "group": group, "error": str(e)})

    def _photo_probe_order(items: list[Any], mode: str) -> list[Any]:
        """Порядок чтения фото (vision дорогой). first_last_then_inner: первый, последний,
        второй, предпоследний… — как описана пробивка (нет явного PDF → первый и последний;
        конец документ → предпоследний). 'in_attachment_order'/'first_then_inner' → по порядку."""
        if mode in ("in_attachment_order", "first_then_inner"):
            return list(items)
        order: list[Any] = []
        i, j = 0, len(items) - 1
        while i <= j:
            order.append(items[i])
            if i != j:
                order.append(items[j])
            i += 1
            j -= 1
        return order

    # 1) Документы (PDF/Excel/zip) первыми — по тексту, дёшево (read_pdf_first).
    for a in docs:
        if all(found.values()):
            break
        _process(a, "document")
    # 2) Фото — vision, порядок first_last_then_inner, лимит max_defect_images_per_case.
    if not all(found.values()) and photos:
        order_mode = getattr(settings, "defect_read_images_order", "first_last_then_inner")
        max_imgs = max(0, int(getattr(settings, "max_defect_images_per_case", 2) or 0))
        probe = _photo_probe_order(photos, order_mode)
        if max_imgs:
            probe = probe[:max_imgs]
        for a in probe:
            if all(found.values()):
                break
            _process(a, "photo")

    n = sum(1 for v in found.values() if v)
    state = "complete" if n >= 3 else ("partial" if n > 0 else "absent")
    flag = {"state": state, "present": {k: bool(v) for k, v in found.items()}, "found": found,
            "n_present": n, "mode": "ai", "attachments": per_att,
            "docs_scanned": len(docs), "photos_scanned": sum(1 for x in per_att if x.get("group") == "photo")}
    # Сохранить в payload кейса
    with connect() as con:
        row = con.execute("SELECT payload_json FROM cases WHERE id=?", (case_id,)).fetchone()
        payload = {}
        try:
            payload = json.loads(row["payload_json"] or "{}") if row else {}
        except Exception:
            payload = {}
        payload["defect_doc_flag"] = flag
        con.execute("UPDATE cases SET payload_json=?, updated_at=? WHERE id=?",
                    (json.dumps(payload, ensure_ascii=False), utcnow(), case_id))
        con.commit()
    return {"ok": True, **flag}


# ── Гарантированный запуск scan_loop при любом импорте модуля ──
# WatchFiles перезагружает код без повторного вызова @app.on_event("startup"),
# поэтому поток _scan_loop не стартует. Вызов на уровне модуля решает проблему.
_start_scan_thread_once()
